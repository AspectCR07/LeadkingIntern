import os
import sys
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import threading
import urllib.request
import json

import pdf_parser

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(base_path, relative_path)
    if not os.path.exists(path):
        path = os.path.join(os.path.abspath("."), relative_path)
    return path

def num_to_words(num):
    """ Convert numeric amount to USD words format """
    units = ["", "ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN", "EIGHT", "NINE", "TEN", 
             "ELEVEN", "TWELVE", "THIRTEEN", "FOURTEEN", "FIFTEEN", "SIXTEEN", "SEVENTEEN", "EIGHTEEN", "NINETEEN"]
    tens = ["", "", "TWENTY", "THIRTY", "FORTY", "FIFTY", "SIXTY", "SEVENTY", "EIGHTY", "NINETY"]
    
    def helper(n):
        if n < 20:
            return units[n]
        elif n < 100:
            return tens[n // 10] + (" " + units[n % 10] if n % 10 != 0 else "")
        elif n < 1000:
            return units[n // 100] + " HUNDRED" + (" AND " + helper(n % 100) if n % 100 != 0 else "")
        elif n < 100000:
            return helper(n // 1000) + " THOUSAND" + (" " + helper(n % 1000) if n % 1000 != 0 else "")
        else:
            return str(n)
            
    if num == 0:
        return "SAY USD ZERO ONLY"
        
    parts = f"{num:.2f}".split(".")
    dollars = int(parts[0])
    cents = int(parts[1])
    
    words = "SAY USD " + helper(dollars)
    if cents > 0:
        words += " AND CENTS " + helper(cents) + " ONLY"
    else:
        words += " ONLY"
    return words

def write_address_block(ws, start_row, col_idx, text, max_rows=10):
    lines = [line.strip() for line in str(text).split("\n") if line.strip()]
    for i in range(max_rows):
        r = start_row + i
        if i < len(lines):
            ws.cell(row=r, column=col_idx).value = lines[i]
        else:
            ws.cell(row=r, column=col_idx).value = None

def write_wrapped_text_to_rows(ws, start_row, col_idx, text, max_rows, limit=65):
    raw_lines = [line.strip() for line in str(text).split("\n") if line.strip()]
    wrapped_lines = []
    for line in raw_lines:
        if len(line) <= limit:
            wrapped_lines.append(line)
        else:
            words = line.split()
            current = []
            for w in words:
                if len(" ".join(current + [w])) <= limit:
                    current.append(w)
                else:
                    wrapped_lines.append(" ".join(current))
                    current = [w]
            if current:
                wrapped_lines.append(" ".join(current))
    for i in range(max_rows):
        r = start_row + i
        if i < len(wrapped_lines):
            ws.cell(row=r, column=col_idx).value = wrapped_lines[i]
        else:
            ws.cell(row=r, column=col_idx).value = None

def clean_weight(val):
    if not val:
        return None
    cleaned = re.sub(r'(?i)kgs|kg|packages|packages|pkg|pkgs', '', str(val)).replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return val

def insert_row_copy_style(ws, row_idx, copy_from_row):
    """ Insert a row in openpyxl and copy cell styles from a reference row """
    ws.insert_rows(row_idx)
    for col_idx in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=copy_from_row, column=col_idx)
        dest_cell = ws.cell(row=row_idx, column=col_idx)
        if src_cell.has_style:
            dest_cell.font = Font(
                name=src_cell.font.name,
                size=src_cell.font.size,
                bold=src_cell.font.bold,
                italic=src_cell.font.italic,
                color=src_cell.font.color,
                underline=src_cell.font.underline
            )
            dest_cell.border = Border(
                left=src_cell.border.left,
                right=src_cell.border.right,
                top=src_cell.border.top,
                bottom=src_cell.border.bottom
            )
            dest_cell.fill = PatternFill(
                fill_type=src_cell.fill.fill_type,
                start_color=src_cell.fill.start_color,
                end_color=src_cell.fill.end_color
            )
            dest_cell.alignment = Alignment(
                horizontal=src_cell.alignment.horizontal,
                vertical=src_cell.alignment.vertical,
                wrap_text=src_cell.alignment.wrap_text
            )
            dest_cell.number_format = src_cell.number_format

def find_value_by_keywords(ws, keywords, all_keywords=None):
    if all_keywords is None:
        all_keywords = keywords
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val:
                val_str = str(val).lower().strip()
                if any(kw in val_str for kw in keywords):
                    # Check cells to the right
                    for col_offset in range(1, 4):
                        if c + col_offset <= ws.max_column:
                            right_val = ws.cell(row=r, column=c+col_offset).value
                            if right_val:
                                rv_str = str(right_val).strip()
                                if rv_str and not any(ak in rv_str.lower() for ak in all_keywords):
                                    return rv_str
                    # Check cells below
                    for row_offset in range(1, 3):
                        if r + row_offset <= ws.max_row:
                            down_val = ws.cell(row=r+row_offset, column=c).value
                            if down_val:
                                dv_str = str(down_val).strip()
                                if dv_str and not any(ak in dv_str.lower() for ak in all_keywords):
                                    return dv_str
    return None

def find_block_below_keyword(ws, keyword, num_rows=4):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and keyword in str(val).lower():
                lines = []
                for row_offset in range(1, num_rows + 1):
                    if r + row_offset <= ws.max_row:
                        v = ws.cell(row=r+row_offset, column=c).value
                        if v and str(v).strip():
                            lines.append(str(v).strip())
                if lines:
                    return "\n".join(lines)
    return None

class XlsSheetWrapper:
    def __init__(self, xls_sheet):
        self.sheet = xls_sheet
        
    @property
    def max_row(self):
        return self.sheet.nrows
        
    @property
    def max_column(self):
        return self.sheet.ncols
        
    def cell(self, row, column, value=None):
        r_idx = row - 1
        c_idx = column - 1
        if r_idx < 0 or r_idx >= self.sheet.nrows or c_idx < 0 or c_idx >= self.sheet.ncols:
            class EmptyCell:
                value = None
            return EmptyCell()
        val = self.sheet.cell_value(r_idx, c_idx)
        class CellWrapper:
            def __init__(self, v):
                self.value = v
        return CellWrapper(val)
        
    def __getitem__(self, key):
        try:
            from openpyxl.utils import coordinate_to_tuple
            r, c = coordinate_to_tuple(key)
            return self.cell(row=r, column=c)
        except:
            class EmptyCell:
                value = None
            return EmptyCell()

class XlsWorkbookWrapper:
    def __init__(self, filepath):
        import xlrd
        self.wb = xlrd.open_workbook(filepath)
        
    @property
    def sheetnames(self):
        return self.wb.sheet_names()
        
    @property
    def active(self):
        return XlsSheetWrapper(self.wb.sheet_by_index(0))
        
    def __getitem__(self, name):
        return XlsSheetWrapper(self.wb.sheet_by_name(name))

DEFAULT_CHARGES = [
    {"desc": "O/F  - USD20  x Ex.rate Rs.98.50", "amount": 1970.0},
    {"desc": "FUEL SURCHARGE USD65 x Rs.98.50", "amount": 6403.0},
    {"desc": "BL CHARGES", "amount": 4500.0},
    {"desc": "surrender BL Charges", "amount": 3500.0},
    {"desc": "SEAL CHARGES/WEIGHTMENT CHARGES", "amount": 1250.0},
    {"desc": "SSR CHARGES", "amount": 1500.0},
    {"desc": "CONTAINER TRANSPORTATION CHARGES", "amount": 9500.0},
    {"desc": "LOADING INTO CONT.CHARGES", "amount": 1000.0},
    {"desc": "CFS CHARGES - GATEWAY", "amount": 5782.0},
    {"desc": "SURVEY CHARGES", "amount": 650.0},
    {"desc": "DOCUMENTATION CHARAGES", "amount": 500.0},
    {"desc": "EIR AMENDMENT CHARGES", "amount": 250.0},
    {"desc": "HANDLING CHARGES", "amount": 4000.0},
    {"desc": "TRANSHIPMENT", "amount": 750.0},
    {"desc": "SERVICE CHARGES", "amount": 1620.0}
]

class UnifiedGeneratorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Unified Logistics & Billing Suite")
        self.geometry("1200x850")
        self.configure(bg="#F1F5F9")
        
        # Styles Setup
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure(".", background="#F1F5F9", font=("Segoe UI", 10))
        self.style.configure("TLabel", background="#F1F5F9", foreground="#334155")
        self.style.configure("Header.TLabel", font=("Segoe UI", 16, "bold"), foreground="#1E293B", background="#F1F5F9")
        self.style.configure("Card.TFrame", background="#FFFFFF", relief="flat", borderwidth=1)
        self.style.configure("CardTitle.TLabel", font=("Segoe UI", 12, "bold"), foreground="#1E293B", background="#FFFFFF")
        
        self.style.configure("TNotebook", background="#F1F5F9", borderwidth=0)
        self.style.configure("TNotebook.Tab", background="#CBD5E1", foreground="#475569", padding=[15, 6], font=("Segoe UI", 10, "bold"))
        self.style.map("TNotebook.Tab", background=[("selected", "#FFFFFF")], foreground=[("selected", "#2563EB")])
        
        self.style.configure("Primary.TButton", background="#2563EB", foreground="#FFFFFF", borderwidth=0, font=("Segoe UI", 10, "bold"), padding=[12, 6])
        self.style.map("Primary.TButton", background=[("active", "#1D4ED8"), ("pressed", "#1E40AF")])
        
        self.style.configure("Secondary.TButton", background="#64748B", foreground="#FFFFFF", borderwidth=0, font=("Segoe UI", 10), padding=[10, 4])
        self.style.map("Secondary.TButton", background=[("active", "#475569"), ("pressed", "#334155")])
        
        self.style.configure("Danger.TButton", background="#EF4444", foreground="#FFFFFF", borderwidth=0, font=("Segoe UI", 9, "bold"), padding=[4, 2])
        self.style.map("Danger.TButton", background=[("active", "#DC2626"), ("pressed", "#B91C1C")])

        # State Variables
        self.inputs = {}
        self.billing_vars = {}
        self.container_rows = []
        self.charge_rows = []
        
        # Files Selection
        self.invoice_path_var = tk.StringVar()
        self.packing_path_var = tk.StringVar()
        self.checklist_path_var = tk.StringVar()
        self.export_path_var = tk.StringVar()
        
        app_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        default_export = os.path.join(app_dir, "Generated")
        self.export_path_var.set(default_export)
        
        # Summary variables
        self.total_container_pkgs = tk.StringVar(value="0 BAGS")
        self.total_container_net = tk.DoubleVar(value=0.0)
        self.total_container_gross = tk.DoubleVar(value=0.0)
        self.total_container_cbm = tk.DoubleVar(value=0.0)
        
        self.total_taxable_var = tk.DoubleVar(value=0.0)
        self.total_cgst_var = tk.DoubleVar(value=0.0)
        self.total_sgst_var = tk.DoubleVar(value=0.0)
        self.grand_total_var = tk.DoubleVar(value=0.0)
        self.tds_var = tk.DoubleVar(value=0.0)
        self.total_payable_var = tk.DoubleVar(value=0.0)
        
        self.fmt_taxable = tk.StringVar(value="Rs. 0.00")
        self.fmt_cgst = tk.StringVar(value="Rs. 0.00")
        self.fmt_sgst = tk.StringVar(value="Rs. 0.00")
        self.fmt_grand = tk.StringVar(value="Rs. 0.00")
        self.fmt_tds = tk.StringVar(value="Rs. 0.00")
        self.fmt_payable = tk.StringVar(value="Rs. 0.00")
        
        # Documents Generation selection
        self.gen_bl = tk.BooleanVar(value=True)
        self.bl_template_var = tk.StringVar(value="Standard (CPS 001)")
        self.gen_bill = tk.BooleanVar(value=True)
        self.gen_awb = tk.BooleanVar(value=False)
        
        self.init_variables()
        self.create_widgets()
        self.load_defaults()
        self.recalculate_totals()
        
        # Start background thread to fetch live exchange rate
        threading.Thread(target=self.fetch_live_exchange_rate, daemon=True).start()
        
    def init_variables(self):
        # General Fields
        fields = [
            "shipper", "consignee", "notify", "also_notify", "cha", "forwarding_agent",
            "vessel_voyage", "port_of_loading", "port_of_discharge", "final_destination",
            "country_of_origin", "country_of_dest", "terms_delivery_payment",
            "invoice_no_date", "buyer_order_no_date", "pre_carriage", "place_of_receipt",
            "sb_no_date", "hs_code", "payment_mode", "dimensions", "gross_weight",
            "net_weight", "special_instructions", "rbi_code", "ie_code", "account",
            "gr_no_date", "other_ref", "buyer_if_other", "marks_nos", "contents",
            "bank_ac_drawback", "type_of_packing", "total_pkgs", "cargo_desc", "invoice_value"
        ]
        for f in fields:
            self.inputs[f] = tk.StringVar()
            
        # Billing Metadata fields
        billing_fields = [
            "Shipper Name", "JOB NO:", "BOOKING REF.  NO", "MBL ", "DESTINATION",
            "NO OF PACKS", "GROSS WEIGHT", "CHARGABLE WEIGHT", "DESCRIPTION",
            "INVOICE NO:&  DATE", "SHIPPING LINE", "SB. NO", "Account Name",
            "Vessel Details", "CONTAINER NUMBER ", "CBM", "Remarks"
        ]
        for bf in billing_fields:
            self.billing_vars[bf] = tk.StringVar()
            
    def create_widgets(self):
        # Top Title Bar
        title_frame = tk.Frame(self, bg="#1E293B", height=60)
        title_frame.pack(fill="x", side="top")
        title_frame.pack_propagate(False)
        
        title_lbl = tk.Label(title_frame, text="Logistics & Billing Suite", font=("Segoe UI", 18, "bold"), fg="#F8FAFC", bg="#1E293B")
        title_lbl.pack(side="left", padx=20, pady=12)
        
        subtitle_lbl = tk.Label(title_frame, text="Integrated parser & document compiler", font=("Segoe UI", 10), fg="#94A3B8", bg="#1E293B")
        subtitle_lbl.pack(side="left", padx=10, pady=18)
        
        # Tabs container
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=20, pady=15)
        
        self.tab_upload = ttk.Frame(self.notebook)
        self.tab_general = ttk.Frame(self.notebook)
        self.tab_containers = ttk.Frame(self.notebook)
        self.tab_billing = ttk.Frame(self.notebook)
        self.tab_export = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab_upload, text="1. Upload Documents")
        self.notebook.add(self.tab_general, text="2. General Details")
        self.notebook.add(self.tab_containers, text="3. Container Details")
        self.notebook.add(self.tab_billing, text="4. Billing & Charges")
        self.notebook.add(self.tab_export, text="5. Generate Output")
        
        self.build_upload_tab()
        self.build_general_tab()
        self.build_containers_tab()
        self.build_billing_tab()
        self.build_export_tab()
        
    def build_upload_tab(self):
        card = ttk.Frame(self.tab_upload, style="Card.TFrame")
        card.pack(fill="both", expand=True, padx=20, pady=20)
        
        lbl_title = ttk.Label(card, text="Upload Shipment Documents (Checklist, Invoice, Packing List)", style="CardTitle.TLabel")
        lbl_title.pack(anchor="w", padx=20, pady=(20, 10))
        
        lbl_desc = tk.Label(card, text="Upload any checklist (PDF), invoice, or packing list (Excel or PDF) to auto-fill the shipment forms. All are optional—uploading even one is enough to extract details.", font=("Segoe UI", 10), fg="#64748B", bg="#FFFFFF", justify="left")
        lbl_desc.pack(anchor="w", padx=20, pady=(0, 20))
        
        # Invoice selection
        inv_frame = tk.Frame(card, bg="#FFFFFF")
        inv_frame.pack(fill="x", padx=20, pady=8)
        lbl_inv = tk.Label(inv_frame, text="Select Invoice (.xlsx/.xls/.pdf):", font=("Segoe UI", 10, "bold"), fg="#475569", bg="#FFFFFF", width=28, anchor="w")
        lbl_inv.pack(side="left")
        ent_inv = ttk.Entry(inv_frame, textvariable=self.invoice_path_var, font=("Segoe UI", 10), width=65)
        ent_inv.pack(side="left", padx=10)
        btn_inv = ttk.Button(inv_frame, text="Browse...", style="Secondary.TButton", command=self.upload_invoice)
        btn_inv.pack(side="left")
        
        # Packing list selection
        pl_frame = tk.Frame(card, bg="#FFFFFF")
        pl_frame.pack(fill="x", padx=20, pady=8)
        lbl_pl = tk.Label(pl_frame, text="Select Packing List (.xlsx/.xls/.pdf):", font=("Segoe UI", 10, "bold"), fg="#475569", bg="#FFFFFF", width=28, anchor="w")
        lbl_pl.pack(side="left")
        ent_pl = ttk.Entry(pl_frame, textvariable=self.packing_path_var, font=("Segoe UI", 10), width=65)
        ent_pl.pack(side="left", padx=10)
        btn_pl = ttk.Button(pl_frame, text="Browse...", style="Secondary.TButton", command=self.upload_packing_list)
        btn_pl.pack(side="left")

        # Checklist selection
        chk_frame = tk.Frame(card, bg="#FFFFFF")
        chk_frame.pack(fill="x", padx=20, pady=8)
        lbl_chk = tk.Label(chk_frame, text="Select Checklist (.pdf):", font=("Segoe UI", 10, "bold"), fg="#475569", bg="#FFFFFF", width=28, anchor="w")
        lbl_chk.pack(side="left")
        ent_chk = ttk.Entry(chk_frame, textvariable=self.checklist_path_var, font=("Segoe UI", 10), width=65)
        ent_chk.pack(side="left", padx=10)
        btn_chk = ttk.Button(chk_frame, text="Browse...", style="Secondary.TButton", command=self.upload_checklist)
        btn_chk.pack(side="left")
        
        # Status Box
        self.lbl_status_box = tk.Label(card, text="Ready. Choose files to extract details.", font=("Segoe UI", 11, "bold"), fg="#2563EB", bg="#EFF6FF", relief="solid", bd=1, padx=15, pady=10)
        self.lbl_status_box.pack(fill="x", padx=20, pady=25)
        
        # Button to continue
        btn_cont = ttk.Button(card, text="Review Details & Addresses →", style="Primary.TButton", command=lambda: self.notebook.select(1))
        btn_cont.pack(anchor="e", padx=20, pady=15)
        
    def build_general_tab(self):
        canvas_frame = tk.Frame(self.tab_general, bg="#F1F5F9")
        canvas_frame.pack(fill="both", expand=True)
        
        canvas = tk.Canvas(canvas_frame, bg="#F1F5F9", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#F1F5F9")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=20, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        
        card = ttk.Frame(scrollable_frame, style="Card.TFrame")
        card.pack(fill="both", expand=True, padx=5, pady=5)
        
        lbl = ttk.Label(card, text="Parsed Entities & Addresses (Fully Editable)", style="CardTitle.TLabel")
        lbl.grid(row=0, column=0, columnspan=4, sticky="w", padx=20, pady=15)
        
        # Address Details
        address_fields = [
            ("Shipper Details (Invoice B3:B6)", "shipper"),
            ("Consignee Details (Invoice B8:B13)", "consignee"),
            ("Notify Party Details (Notify)", "notify"),
            ("Also Notify Details", "also_notify"),
            ("CHA Address/Details", "cha"),
            ("Forwarding Agent", "forwarding_agent")
        ]
        
        self.address_textboxes = {}
        for idx, (label, key) in enumerate(address_fields):
            r = 1 + (idx // 2)
            c = (idx % 2) * 2
            
            lbl_w = tk.Label(card, text=label, font=("Segoe UI", 10, "bold"), fg="#475569", bg="#FFFFFF", anchor="w")
            lbl_w.grid(row=r*2-1, column=c, padx=20, pady=(10, 2), sticky="w")
            
            txt = tk.Text(card, font=("Segoe UI", 10), width=48, height=4, relief="solid", bd=1)
            txt.grid(row=r*2, column=c, columnspan=2, padx=20, pady=(2, 10), sticky="ew")
            self.address_textboxes[key] = txt
            
        # Meta grid
        lbl_meta = ttk.Label(card, text="Voyage, Ports & Shipment References", style="CardTitle.TLabel")
        lbl_meta.grid(row=7, column=0, columnspan=4, sticky="w", padx=20, pady=(20, 10))
        
        entries_config = [
            ("Vessel / Voyage Flight:", "vessel_voyage", 8, 0),
            ("Port of Loading:", "port_of_loading", 8, 2),
            ("Port of Discharge:", "port_of_discharge", 9, 0),
            ("Final Destination:", "final_destination", 9, 2),
            ("Country of Origin:", "country_of_origin", 10, 0),
            ("Country of Final Dest:", "country_of_dest", 10, 2),
            ("Invoice No & Date:", "invoice_no_date", 11, 0),
            ("Buyer's Order No & Date:", "buyer_order_no_date", 11, 2),
            ("Pre-Carriage by:", "pre_carriage", 12, 0),
            ("Place of Receipt:", "place_of_receipt", 12, 2),
            ("Shipping Bill No & Date:", "sb_no_date", 13, 0),
            ("HS Code:", "hs_code", 13, 2),
            ("Payment Mode (Prepaid/Collect):", "payment_mode", 14, 0),
            ("Gross Weight (e.g. 500 KGS):", "gross_weight", 14, 2),
            ("Net Weight (e.g. 480 KGS):", "net_weight", 15, 0),
            ("Dimensions (CMS):", "dimensions", 15, 2),
            ("Other Reference(s):", "other_ref", 16, 0),
            ("Type of Packing:", "type_of_packing", 16, 2),
            ("Total Packages (Text):", "total_pkgs", 17, 0)
        ]
        
        for label_text, key, r, c in entries_config:
            lbl_w = ttk.Label(card, text=label_text, background="#FFFFFF", font=("Segoe UI", 10, "bold"), width=28, anchor="w")
            lbl_w.grid(row=r, column=c, padx=(20, 5), pady=6, sticky="w")
            
            ent = ttk.Entry(card, textvariable=self.inputs[key], font=("Segoe UI", 10), width=32)
            ent.grid(row=r, column=c + 1, padx=(5, 20), pady=6, sticky="ew")
            
        # Large multi-line items
        textareas = [
            ("Terms of Delivery & Payment:", "terms_delivery_payment", 18),
            ("Special Instructions:", "special_instructions", 19),
            ("Bank Account / Drawback Details:", "bank_ac_drawback", 20),
            ("Buyer (If other than consignee):", "buyer_if_other", 21),
            ("Marks & Numbers:", "marks_nos", 22),
            ("Cargo Description:", "cargo_desc", 23)
        ]
        
        self.multiline_entries = {}
        for idx, (label_text, key, r) in enumerate(textareas):
            lbl_w = tk.Label(card, text=label_text, font=("Segoe UI", 10, "bold"), fg="#475569", bg="#FFFFFF", anchor="w")
            lbl_w.grid(row=r, column=0, padx=20, pady=(10, 2), sticky="w")
            
            txt = tk.Text(card, font=("Segoe UI", 10), width=85, height=3, relief="solid", bd=1)
            txt.grid(row=r, column=1, columnspan=3, padx=20, pady=(2, 10), sticky="ew")
            self.multiline_entries[key] = txt
            
        btn = ttk.Button(card, text="Continue to Container Details →", style="Primary.TButton", command=lambda: self.notebook.select(2))
        btn.grid(row=24, column=3, sticky="e", padx=20, pady=20)
        
    def build_containers_tab(self):
        card = ttk.Frame(self.tab_containers, style="Card.TFrame")
        card.pack(fill="both", expand=True, padx=20, pady=20)
        
        top_ctrl = tk.Frame(card, bg="#FFFFFF")
        top_ctrl.pack(fill="x", padx=15, pady=10)
        
        lbl = ttk.Label(top_ctrl, text="Container & Package List (For BL Instructions)", style="CardTitle.TLabel")
        lbl.pack(side="left")
        
        btn_add = ttk.Button(top_ctrl, text="+ Add Container Row", style="Primary.TButton", command=self.add_blank_container_row)
        btn_add.pack(side="right", padx=5)
        
        # Table Header
        headers_frame = tk.Frame(card, bg="#F8FAFC", height=30)
        headers_frame.pack(fill="x", padx=15, pady=(5, 0))
        
        headers = [
            ("Sl.No", 50, "center"),
            ("CONTAINER NUMBER", 180, "w"),
            ("SEAL NUMBER", 180, "w"),
            ("NO. OF PKGS (e.g. 400 BAGS)", 180, "w"),
            ("NET WT (KGS)", 110, "e"),
            ("GROSS WT (KGS)", 110, "e"),
            ("CBM", 100, "e"),
            ("ACTION", 80, "center")
        ]
        
        for name, width, anchor in headers:
            lbl = tk.Label(headers_frame, text=name, font=("Segoe UI", 9, "bold"), fg="#475569", bg="#F8FAFC", width=width, anchor=anchor)
            lbl.pack(side="left", padx=5, fill="y")
            lbl.pack_propagate(False)
            lbl.config(width=int(width/8))
            
        # Scrollable container
        canvas_frame = tk.Frame(card, bg="#FFFFFF")
        canvas_frame.pack(fill="both", expand=True, padx=15, pady=5)
        
        self.container_canvas = tk.Canvas(canvas_frame, bg="#FFFFFF", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.container_canvas.yview)
        
        self.container_scroll_frame = tk.Frame(self.container_canvas, bg="#FFFFFF")
        self.container_scroll_frame.bind(
            "<Configure>",
            lambda e: self.container_canvas.configure(scrollregion=self.container_canvas.bbox("all"))
        )
        
        self.container_window = self.container_canvas.create_window((0, 0), window=self.container_scroll_frame, anchor="nw")
        self.container_canvas.configure(yscrollcommand=scrollbar.set)
        
        self.container_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.container_canvas.bind('<Configure>', lambda e: self.container_canvas.itemconfig(self.container_window, width=e.width))
        
        # Totals Bar
        self.container_totals_frame = tk.Frame(card, bg="#F8FAFC", height=40)
        self.container_totals_frame.pack(fill="x", padx=15, pady=(5, 15))
        
        self.lbl_container_totals = tk.Label(
            self.container_totals_frame, 
            text="Total Containers: 0 | Total Pkgs: 0 BAGS | Total Net: 0.00 kg | Total Gross: 0.00 kg | Total CBM: 0.000", 
            font=("Segoe UI", 10, "bold"), 
            fg="#1E293B", 
            bg="#F8FAFC", 
            anchor="e"
        )
        self.lbl_container_totals.pack(fill="both", expand=True, padx=20)
        
        self.bind_all("<KeyRelease>", lambda e: self.recalculate_totals())
        
        btn_cont = ttk.Button(card, text="Continue to Billing & Charges →", style="Primary.TButton", command=lambda: self.notebook.select(3))
        btn_cont.pack(anchor="e", padx=20, pady=10)
        
    def add_blank_container_row(self):
        self.add_container_row()
        
    def add_container_row(self, c_no="", s_no="", pkgs="", net_wt=0.0, gross_wt=0.0, cbm=0.0):
        row_frame = tk.Frame(self.container_scroll_frame, bg="#FFFFFF", pady=4)
        row_frame.pack(fill="x", expand=True)
        
        idx = len(self.container_rows) + 1
        
        # Row variables
        c_no_var = tk.StringVar(value=c_no)
        s_no_var = tk.StringVar(value=s_no)
        pkgs_var = tk.StringVar(value=pkgs)
        
        widgets = []
        
        # Sl.No
        lbl_idx = tk.Label(row_frame, text=str(idx), font=("Segoe UI", 10), bg="#FFFFFF", fg="#64748B", width=6, anchor="center")
        lbl_idx.pack(side="left", padx=5)
        widgets.append(lbl_idx)
        
        # Container No
        ent_c_no = ttk.Entry(row_frame, textvariable=c_no_var, font=("Segoe UI", 10), width=20)
        ent_c_no.pack(side="left", padx=5)
        widgets.append(ent_c_no)
        
        # Seal No
        ent_s_no = ttk.Entry(row_frame, textvariable=s_no_var, font=("Segoe UI", 10), width=20)
        ent_s_no.pack(side="left", padx=5)
        widgets.append(ent_s_no)
        
        # Pkgs
        ent_pkgs = ttk.Entry(row_frame, textvariable=pkgs_var, font=("Segoe UI", 10), width=20)
        ent_pkgs.pack(side="left", padx=5)
        widgets.append(ent_pkgs)
        
        # Net
        ent_net = ttk.Entry(row_frame, font=("Segoe UI", 10), width=12, justify="right")
        ent_net.insert(0, f"{net_wt:.2f}")
        ent_net.pack(side="left", padx=5)
        widgets.append(ent_net)
        
        # Gross
        ent_gross = ttk.Entry(row_frame, font=("Segoe UI", 10), width=12, justify="right")
        ent_gross.insert(0, f"{gross_wt:.2f}")
        ent_gross.pack(side="left", padx=5)
        widgets.append(ent_gross)
        
        # CBM
        ent_cbm = ttk.Entry(row_frame, font=("Segoe UI", 10), width=10, justify="right")
        ent_cbm.insert(0, f"{cbm:.3f}")
        ent_cbm.pack(side="left", padx=5)
        widgets.append(ent_cbm)
        
        # Remove button
        btn_del = ttk.Button(row_frame, text="Remove", style="Danger.TButton", command=lambda: self.remove_container_row(row_frame))
        btn_del.pack(side="left", padx=5)
        widgets.append(btn_del)
        
        row_data = {
            "frame": row_frame,
            "widgets": widgets,
            "c_no_var": c_no_var,
            "s_no_var": s_no_var,
            "pkgs_var": pkgs_var,
            "net_ent": ent_net,
            "gross_ent": ent_gross,
            "cbm_ent": ent_cbm,
            "lbl_idx": lbl_idx
        }
        self.container_rows.append(row_data)
        self.recalculate_totals()
        
    def remove_container_row(self, frame):
        for idx, row in enumerate(self.container_rows):
            if row["frame"] == frame:
                for w in row["widgets"]:
                    w.destroy()
                row["frame"].destroy()
                self.container_rows.pop(idx)
                break
                
        for idx, row in enumerate(self.container_rows):
            row["lbl_idx"].config(text=str(idx+1))
            
        self.recalculate_totals()
        
    def clear_container_rows(self):
        for row in self.container_rows:
            for w in row["widgets"]:
                w.destroy()
            row["frame"].destroy()
        self.container_rows.clear()
        self.recalculate_totals()
        
    def build_billing_tab(self):
        canvas_frame = tk.Frame(self.tab_billing, bg="#F1F5F9")
        canvas_frame.pack(fill="both", expand=True)
        
        canvas = tk.Canvas(canvas_frame, bg="#F1F5F9", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        self.billing_scrollable_frame = tk.Frame(canvas, bg="#F1F5F9")
        
        self.billing_scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.billing_scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=20, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        
        # 1. Billing Metadata Card
        meta_card = ttk.Frame(self.billing_scrollable_frame, style="Card.TFrame")
        meta_card.pack(fill="x", padx=5, pady=5)
        
        lbl_meta = ttk.Label(meta_card, text="Billing Invoice Metadata (Taxes & GST Layout)", style="CardTitle.TLabel")
        lbl_meta.grid(row=0, column=0, columnspan=4, sticky="w", padx=20, pady=15)
        
        fields_left = [
            ("Shipper Name", "Shipper Name"),
            ("JOB NO:", "JOB NO:"),
            ("BOOKING REF.  NO", "BOOKING REF.  NO"),
            ("MBL ", "MBL "),
            ("DESTINATION", "DESTINATION"),
            ("NO OF PACKS", "NO OF PACKS"),
            ("GROSS WEIGHT", "GROSS WEIGHT"),
            ("CHARGABLE WEIGHT", "CHARGABLE WEIGHT")
        ]
        
        fields_right = [
            ("DESCRIPTION", "DESCRIPTION"),
            ("INVOICE NO:&  DATE", "INVOICE NO:&  DATE"),
            ("SHIPPING LINE", "SHIPPING LINE"),
            ("SB. NO", "SB. NO"),
            ("Account Name", "Account Name"),
            ("Vessel Details", "Vessel Details"),
            ("CONTAINER NUMBER", "CONTAINER NUMBER "),
            ("CBM", "CBM"),
            ("Remarks", "Remarks")
        ]
        
        for idx, (label, key) in enumerate(fields_left):
            lbl_w = ttk.Label(meta_card, text=label, background="#FFFFFF", font=("Segoe UI", 9, "bold"), width=22, anchor="w")
            lbl_w.grid(row=idx+1, column=0, padx=(20, 5), pady=5, sticky="w")
            ent = ttk.Entry(meta_card, textvariable=self.billing_vars[key], font=("Segoe UI", 9), width=28)
            ent.grid(row=idx+1, column=1, padx=(5, 20), pady=5, sticky="ew")
            
        for idx, (label, key) in enumerate(fields_right):
            lbl_w = ttk.Label(meta_card, text=label, background="#FFFFFF", font=("Segoe UI", 9, "bold"), width=22, anchor="w")
            lbl_w.grid(row=idx+1, column=2, padx=(20, 5), pady=5, sticky="w")
            ent = ttk.Entry(meta_card, textvariable=self.billing_vars[key], font=("Segoe UI", 9), width=28)
            ent.grid(row=idx+1, column=3, padx=(5, 20), pady=5, sticky="ew")
            
        # 1.5 Ocean Freight Converter Card
        of_card = ttk.Frame(self.billing_scrollable_frame, style="Card.TFrame")
        of_card.pack(fill="x", padx=5, pady=5)
        
        of_inner = tk.Frame(of_card, bg="#FFFFFF", pady=10, padx=15)
        of_inner.pack(fill="x")
        
        title_of = ttk.Label(of_inner, text="Ocean Freight (O/F) USD Converter", font=("Segoe UI", 11, "bold"), background="#FFFFFF", foreground="#1E293B")
        title_of.pack(side="left", padx=(0, 20))
        
        lbl_usd = ttk.Label(of_inner, text="O/F (USD):", background="#FFFFFF")
        lbl_usd.pack(side="left", padx=5)
        
        self.of_usd_var = tk.StringVar(value="20.00")
        ent_usd = ttk.Entry(of_inner, textvariable=self.of_usd_var, font=("Segoe UI", 10), width=10, justify="right")
        ent_usd.pack(side="left", padx=5)
        
        lbl_ex = ttk.Label(of_inner, text="Exchange Rate (INR):", background="#FFFFFF")
        lbl_ex.pack(side="left", padx=(15, 5))
        
        self.of_ex_rate_var = tk.StringVar(value="98.50")
        ent_ex = ttk.Entry(of_inner, textvariable=self.of_ex_rate_var, font=("Segoe UI", 10), width=10, justify="right")
        ent_ex.pack(side="left", padx=5)
        
        lbl_eq = ttk.Label(of_inner, text="=", background="#FFFFFF", font=("Segoe UI", 10, "bold"))
        lbl_eq.pack(side="left", padx=10)
        
        self.of_inr_val_var = tk.StringVar(value="Rs. 1,970.00")
        lbl_inr_val = tk.Label(of_inner, textvariable=self.of_inr_val_var, font=("Segoe UI", 10, "bold"), fg="#16A34A", bg="#FFFFFF", anchor="w")
        lbl_inr_val.pack(side="left", padx=5)
        
        # Trace updates to sync converter to table
        self.of_usd_var.trace_add("write", lambda *args: self.update_ocean_freight_from_usd())
        self.of_ex_rate_var.trace_add("write", lambda *args: self.update_ocean_freight_from_usd())

        # 2. Charges Card
        charges_card = ttk.Frame(self.billing_scrollable_frame, style="Card.TFrame")
        charges_card.pack(fill="x", padx=5, pady=15)
        
        top_ctrl = tk.Frame(charges_card, bg="#FFFFFF")
        top_ctrl.pack(fill="x", padx=15, pady=10)
        
        lbl_c = ttk.Label(top_ctrl, text="Billing Charges", style="CardTitle.TLabel")
        lbl_c.pack(side="left")
        
        btn_reset = ttk.Button(top_ctrl, text="Reset Defaults", style="Secondary.TButton", command=self.reset_charges)
        btn_reset.pack(side="right", padx=5)
        
        btn_add = ttk.Button(top_ctrl, text="+ Add Charge Row", style="Primary.TButton", command=self.add_blank_charge_row)
        btn_add.pack(side="right", padx=5)
        
        # Table Header
        self.headers_frame = tk.Frame(charges_card, bg="#F8FAFC", height=30)
        self.headers_frame.pack(fill="x", padx=15, pady=(5, 0))
        
        headers = [
            ("Si.No", 50, "center"),
            ("CONTENT / CHARGE NAME", 300, "w"),
            ("AMOUNT", 110, "e"),
            ("CGST %", 70, "center"),
            ("CGST AMT", 100, "e"),
            ("SGST %", 70, "center"),
            ("SGST AMT", 100, "e"),
            ("TOTAL", 120, "e"),
            ("ACTION", 80, "center")
        ]
        
        for name, width, anchor in headers:
            lbl = tk.Label(self.headers_frame, text=name, font=("Segoe UI", 9, "bold"), fg="#475569", bg="#F8FAFC", width=width, anchor=anchor)
            lbl.pack(side="left", padx=5, fill="y")
            lbl.pack_propagate(False)
            lbl.config(width=int(width/8))
            
        # Charges List Frame (Grid packed inside)
        self.charges_list_frame = tk.Frame(charges_card, bg="#FFFFFF")
        self.charges_list_frame.pack(fill="x", padx=15, pady=5)
        
        # Billing Dashboard / Calculations Panel
        dash_card = ttk.Frame(self.billing_scrollable_frame, style="Card.TFrame")
        dash_card.pack(fill="x", padx=5, pady=5)
        
        lbl_dash = ttk.Label(dash_card, text="Billing Calculation Summary", style="CardTitle.TLabel")
        lbl_dash.pack(anchor="w", padx=20, pady=(15, 10))
        
        dash_grid = tk.Frame(dash_card, bg="#FFFFFF")
        dash_grid.pack(fill="x", padx=20, pady=5)
        
        def add_dash_cell(parent, label, var, r, c, is_bold=False, val_color="#1E293B"):
            f_lbl = tk.Frame(parent, bg="#FFFFFF")
            f_lbl.grid(row=r, column=c, padx=15, pady=5, sticky="ew")
            
            lbl_w = tk.Label(f_lbl, text=label, font=("Segoe UI", 10, "bold" if is_bold else "normal"), fg="#475569" if not is_bold else "#1E293B", bg="#FFFFFF")
            lbl_w.pack(side="left")
            val_w = tk.Label(f_lbl, textvariable=var, font=("Segoe UI", 11, "bold"), fg=val_color, bg="#FFFFFF")
            val_w.pack(side="right", padx=10)
            
        add_dash_cell(dash_grid, "Total Taxable Amount:", self.fmt_taxable, 0, 0)
        add_dash_cell(dash_grid, "CGST Amount:", self.fmt_cgst, 0, 1)
        add_dash_cell(dash_grid, "SGST Amount:", self.fmt_sgst, 0, 2)
        add_dash_cell(dash_grid, "Grand Total with GST:", self.fmt_grand, 1, 0, is_bold=True)
        add_dash_cell(dash_grid, "TDS Deducted (2.0%):", self.fmt_tds, 1, 1, val_color="#EF4444")
        add_dash_cell(dash_grid, "Total Payable Amount:", self.fmt_payable, 1, 2, is_bold=True, val_color="#16A34A")
        
        btn_cont = ttk.Button(dash_card, text="Continue to Generate Output →", style="Primary.TButton", command=lambda: self.notebook.select(4))
        btn_cont.pack(anchor="e", padx=20, pady=15)
        
    def add_blank_charge_row(self):
        self.add_charge_row("", 0.0, 9.0, 9.0)
        self.recalculate_billing()
        
    def on_charge_desc_change(self, row_data):
        desc = row_data["desc_var"].get().strip().lower()
        is_of = desc.startswith("o/f") or desc.startswith("of ") or desc.startswith("of-") or desc == "of" or "ocean freight" in desc
        rate = 2.5 if is_of else 9.0
        
        row_data["cgst_rate_ent"].delete(0, tk.END)
        row_data["cgst_rate_ent"].insert(0, f"{rate:.1f}")
        row_data["sgst_rate_ent"].delete(0, tk.END)
        row_data["sgst_rate_ent"].insert(0, f"{rate:.1f}")
        
        self.recalculate_billing()
        
    def update_ocean_freight_from_usd(self):
        try:
            usd_str = self.of_usd_var.get().strip()
            usd = float(usd_str) if usd_str else 0.0
        except ValueError:
            usd = 0.0
            
        try:
            ex_str = self.of_ex_rate_var.get().strip()
            ex_rate = float(ex_str) if ex_str else 0.0
        except ValueError:
            ex_rate = 0.0
            
        inr_amt = round(usd * ex_rate, 2)
        self.of_inr_val_var.set(f"Rs. {inr_amt:,.2f}")
        
        found = False
        for row in self.charge_rows:
            desc = row["desc_var"].get().strip()
            desc_lower = desc.lower()
            if desc_lower.startswith("o/f") or desc_lower.startswith("of ") or desc_lower.startswith("of-") or desc_lower == "of" or "ocean freight" in desc_lower:
                row["amount_ent"].delete(0, tk.END)
                row["amount_ent"].insert(0, f"{inr_amt:.2f}")
                row["desc_var"].set(f"O/F  - USD{usd_str}  x Ex.rate Rs.{ex_str}")
                found = True
            elif "fuel surcharge" in desc_lower and "usd" in desc_lower:
                match = re.search(r'usd\s*(\d+(?:\.\d+)?)', desc_lower)
                if match:
                    try:
                        fs_usd = float(match.group(1))
                        fs_inr = round(fs_usd * ex_rate, 2)
                        row["amount_ent"].delete(0, tk.END)
                        row["amount_ent"].insert(0, f"{fs_inr:.2f}")
                        row["desc_var"].set(f"FUEL SURCHARGE USD{match.group(1)} x Rs.{ex_str}")
                    except ValueError:
                        pass
                
        if not found:
            self.add_charge_row(f"O/F  - USD{usd_str}  x Ex.rate Rs.{ex_str}", inr_amt, 2.5, 2.5)
            
        self.recalculate_billing()
        
    def add_charge_row(self, description, amount, cgst_rate, sgst_rate):
        row_frame = tk.Frame(self.charges_list_frame, bg="#FFFFFF", pady=2)
        row_frame.pack(fill="x", expand=True)
        
        idx = len(self.charge_rows) + 1
        
        # Widgets variables
        desc_var = tk.StringVar(value=description)
        amount_var = tk.DoubleVar(value=amount)
        cgst_rate_var = tk.DoubleVar(value=cgst_rate)
        sgst_rate_var = tk.DoubleVar(value=sgst_rate)
        
        cgst_amt_var = tk.DoubleVar(value=0.0)
        sgst_amt_var = tk.DoubleVar(value=0.0)
        total_var = tk.DoubleVar(value=0.0)
        
        widgets = []
        
        # 1. Si.No
        lbl_idx = tk.Label(row_frame, text=str(idx), font=("Segoe UI", 10), bg="#FFFFFF", fg="#64748B", width=6, anchor="center")
        lbl_idx.pack(side="left", padx=5)
        widgets.append(lbl_idx)
        
        # 2. CONTENT / Charge Name Entry
        ent_desc = ttk.Entry(row_frame, textvariable=desc_var, font=("Segoe UI", 10), width=36)
        ent_desc.pack(side="left", padx=5)
        widgets.append(ent_desc)
        
        # 3. AMOUNT Entry
        ent_amount = ttk.Entry(row_frame, font=("Segoe UI", 10), width=12, justify="right")
        ent_amount.insert(0, f"{amount:.2f}")
        ent_amount.pack(side="left", padx=5)
        widgets.append(ent_amount)
        
        # 4. CGST % Entry
        ent_cgst_rate = ttk.Entry(row_frame, font=("Segoe UI", 10), width=8, justify="center")
        ent_cgst_rate.insert(0, f"{cgst_rate:.1f}")
        ent_cgst_rate.pack(side="left", padx=5)
        widgets.append(ent_cgst_rate)
        
        # 5. CGST Amt Label
        lbl_cgst_amt = tk.Label(row_frame, text="0.00", font=("Segoe UI", 10), bg="#FFFFFF", fg="#EF4444", width=12, anchor="e")
        lbl_cgst_amt.pack(side="left", padx=5)
        widgets.append(lbl_cgst_amt)
        
        # 6. SGST % Entry
        ent_sgst_rate = ttk.Entry(row_frame, font=("Segoe UI", 10), width=8, justify="center")
        ent_sgst_rate.insert(0, f"{sgst_rate:.1f}")
        ent_sgst_rate.pack(side="left", padx=5)
        widgets.append(ent_sgst_rate)
        
        # 7. SGST Amt Label
        lbl_sgst_amt = tk.Label(row_frame, text="0.00", font=("Segoe UI", 10), bg="#FFFFFF", fg="#EF4444", width=12, anchor="e")
        lbl_sgst_amt.pack(side="left", padx=5)
        widgets.append(lbl_sgst_amt)
        
        # 8. Total Amount Label
        lbl_total = tk.Label(row_frame, text="0.00", font=("Segoe UI", 10, "bold"), bg="#FFFFFF", fg="#0F172A", width=14, anchor="e")
        lbl_total.pack(side="left", padx=5)
        widgets.append(lbl_total)
        
        # 9. Delete Button
        btn_del = ttk.Button(row_frame, text="Remove", style="Danger.TButton", command=lambda: self.remove_charge_row(row_frame))
        btn_del.pack(side="left", padx=5)
        widgets.append(btn_del)
        
        row_data = {
            "frame": row_frame,
            "widgets": widgets,
            "desc_var": desc_var,
            "amount_ent": ent_amount,
            "cgst_rate_var": cgst_rate_var,
            "sgst_rate_var": sgst_rate_var,
            "cgst_rate_ent": ent_cgst_rate,
            "sgst_rate_ent": ent_sgst_rate,
            "cgst_amt_var": cgst_amt_var,
            "sgst_amt_var": sgst_amt_var,
            "total_var": total_var,
            "lbl_idx": lbl_idx,
            "lbl_cgst_amt": lbl_cgst_amt,
            "lbl_sgst_amt": lbl_sgst_amt,
            "lbl_total": lbl_total
        }
        
        ent_desc.bind("<KeyRelease>", lambda e, r=row_data: self.on_charge_desc_change(r))
        self.charge_rows.append(row_data)
        
    def remove_charge_row(self, frame):
        for idx, row in enumerate(self.charge_rows):
            if row["frame"] == frame:
                for w in row["widgets"]:
                    w.destroy()
                row["frame"].destroy()
                self.charge_rows.pop(idx)
                break
                
        for idx, row in enumerate(self.charge_rows):
            row["lbl_idx"].config(text=str(idx+1))
            
        self.recalculate_billing()
        
    def reset_charges(self):
        for row in self.charge_rows:
            for w in row["widgets"]:
                w.destroy()
            row["frame"].destroy()
        self.charge_rows.clear()
        
        for c in DEFAULT_CHARGES:
            desc = c["desc"].strip().lower()
            is_of = desc.startswith("o/f") or desc.startswith("of ") or desc.startswith("of-") or desc == "of" or "ocean freight" in desc
            cgst_rate = 2.5 if is_of else 9.0
            sgst_rate = 2.5 if is_of else 9.0
            self.add_charge_row(c["desc"], c["amount"], cgst_rate, sgst_rate)
            
        self.recalculate_billing()
        
    def build_export_tab(self):
        card = ttk.Frame(self.tab_export, style="Card.TFrame")
        card.pack(fill="both", expand=True, padx=20, pady=20)
        
        panel_left = tk.Frame(card, bg="#FFFFFF")
        panel_left.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        
        lbl_select = ttk.Label(panel_left, text="Compile Output Documents", style="CardTitle.TLabel")
        lbl_select.pack(anchor="w", pady=(0, 15))
        
        chk_style = {"background": "#FFFFFF", "font": ("Segoe UI", 11), "anchor": "w"}
        cb_bl = tk.Checkbutton(panel_left, text="1. Bill of Lading Instructions (BL INSTRUCTION)", variable=self.gen_bl, **chk_style)
        cb_bl.pack(fill="x", pady=8)
        
        # Add template dropdown
        tpl_frame = tk.Frame(panel_left, bg="#FFFFFF")
        tpl_frame.pack(fill="x", pady=(2, 10), padx=25)
        
        lbl_tpl = tk.Label(tpl_frame, text="Use Template:", font=("Segoe UI", 9, "bold"), fg="#475569", bg="#FFFFFF")
        lbl_tpl.pack(side="left", padx=(0, 10))
        
        self.tpl_combo = ttk.Combobox(
            tpl_frame, 
            textvariable=self.bl_template_var, 
            values=["Standard (CPS 001)", "Cotton Home", "PEMI", "Sri Sakthi"],
            state="readonly",
            width=25
        )
        self.tpl_combo.pack(side="left")
        
        cb_bill = tk.Checkbutton(panel_left, text="2. Billing Invoice / GST Invoice (SRI SAKTHI / Bill)", variable=self.gen_bill, **chk_style)
        cb_bill.pack(fill="x", pady=8)
        
        cb_awb = tk.Checkbutton(panel_left, text="3. Air Waybill Instructions (AWB Instructions)", variable=self.gen_awb, **chk_style)
        cb_awb.pack(fill="x", pady=8)
        
        lbl_dest = tk.Label(panel_left, text="Destination Folder Path:", font=("Segoe UI", 10, "bold"), fg="#1E293B", bg="#FFFFFF")
        lbl_dest.pack(anchor="w", pady=(30, 5))
        
        path_frame = tk.Frame(panel_left, bg="#FFFFFF")
        path_frame.pack(fill="x", pady=5)
        
        ent_path = ttk.Entry(path_frame, textvariable=self.export_path_var, font=("Segoe UI", 10))
        ent_path.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        btn_browse = ttk.Button(path_frame, text="Browse...", style="Secondary.TButton", command=self.browse_export_path)
        btn_browse.pack(side="right")
        
        btn_export = ttk.Button(panel_left, text="Generate Documents", style="Primary.TButton", command=self.generate_all)
        btn_export.pack(fill="x", pady=30)
        
        # Log Panel
        panel_right = tk.Frame(card, bg="#FFFFFF")
        panel_right.pack(side="right", fill="both", expand=True, padx=20, pady=20)
        
        lbl_log = ttk.Label(panel_right, text="Generation Progress Log", style="CardTitle.TLabel")
        lbl_log.pack(anchor="w", pady=(0, 10))
        
        self.log_txt = tk.Text(panel_right, font=("Consolas", 9), height=25, width=45, bg="#F8FAFC", relief="solid", bd=1)
        self.log_txt.pack(fill="both", expand=True)
        
    def browse_export_path(self):
        folder = filedialog.askdirectory(title="Select Destination Folder")
        if folder:
            self.export_path_var.set(folder)
            
    def load_defaults(self):
        defaults = {
            "shipper": "CREATIVE PACKAGING SYSTEMS\nNo.195-A, SERVICE ROAD , SOMJI BHAI NAGAR\nKALLIKUPPAM, CHENNAI - 600 053.\nPh: 044-6133 4666 MAIL: bm@creativepackaging.in",
            "consignee": "FOSHAN BOGAL PACKING MACHINERY CO.,LTD\nNO.65,Road 2,Rongxing Industrial Zone ,\nWuzhuang,Luocun Town,Naihai  District,Guangdong\npostcode:528000\nEMAIL ID:amy@sierac.cn\nTAX ID :91440605MA4W0ERU95",
            "notify": "GLOBAL OCEAN TRADING LIMITED,\nBLK 808 FRENCH ROAD,\n06-151 KITCHENER COMPLEX\nSINGAPORE 200208",
            "also_notify": "SAME AS CONSIGNEE",
            "cha": "CREATIVE CHA SERVICES",
            "forwarding_agent": "LEADKING AIR SERVICES PVT. LTD",
            "vessel_voyage": "HUA DA 617 V. 006W",
            "port_of_loading": "CHENNAI / INDIA",
            "port_of_discharge": "SHANGHAI",
            "final_destination": "SHANGHAI / CHINA",
            "country_of_origin": "INDIA",
            "country_of_dest": "CHINA",
            "terms_delivery_payment": "SAMPLE FOR TESTING THE PACKAGING MACHINERY\nNO COMMERCIAL VALUE & NOT FOR SALE\nNO FOREIGN EXCHANGE IS INVOLVED",
            "invoice_no_date": "CPS/001 DT. 08.04.2026",
            "buyer_order_no_date": "Order dt. 05.04.2026",
            "pre_carriage": "NA",
            "place_of_receipt": "NA",
            "sb_no_date": "2421435 DT. 16/04/2026",
            "hs_code": "10051000",
            "payment_mode": "PREPAID",
            "dimensions": "28 X 25.5 X 25.5 CMS",
            "gross_weight": "4.66 Kgs",
            "net_weight": "4.00 Kgs",
            "special_instructions": "DELIVER TO SHANGHAI SHED 3",
            "rbi_code": "RBI-12345",
            "ie_code": "IE-54321",
            "account": "A/C NO. 9876543210 - STATE BANK OF INDIA",
            "gr_no_date": "GR-777 DT. 08.04.2026",
            "other_ref": "REF-001",
            "buyer_if_other": "",
            "marks_nos": "1 CARTON BOX\nContainer no.",
            "contents": "PACKAGING MACHINERY SPARES",
            "bank_ac_drawback": "SBI BANK A/C: 9876543210, IFSC: SBIN0000001",
            "type_of_packing": "PP BAGS",
            "total_pkgs": "1600 PP BAGS",
            "cargo_desc": "HYBRID MAIZE SEEDS (C.P.811)",
            "invoice_value": "USD 120.00"
        }
        
        for k, v in defaults.items():
            if k in self.inputs:
                self.inputs[k].set(v)
            if k in self.address_textboxes:
                self.address_textboxes[k].delete("1.0", tk.END)
                self.address_textboxes[k].insert("1.0", v)
            if k in self.multiline_entries:
                self.multiline_entries[k].delete("1.0", tk.END)
                self.multiline_entries[k].insert("1.0", v)
                
        # Load default containers
        self.add_container_row("REGU3290956", "INMAA2605783", "400 BAGS", 20000.0, 20200.0, 20.0)
        self.add_container_row("CAIU3743948", "INMAA2605787", "400 BAGS", 20000.0, 20200.0, 20.0)
        self.add_container_row("TGBU3715873", "INMAA2605738", "400 BAGS", 20000.0, 20200.0, 20.0)
        self.add_container_row("REGU3265634", "INMAA2605754", "400 BAGS", 20000.0, 20200.0, 20.0)
        
        # Load default billing metadata
        billing_defaults = {
            "Shipper Name": "CREATIVE PACKAGING SYSTEMS",
            "JOB NO:": "2606SX124",
            "BOOKING REF.  NO": "1187698",
            "MBL ": "TALTSM03090091",
            "DESTINATION": "SHANGHAI",
            "NO OF PACKS": "1600 BAGS",
            "GROSS WEIGHT": "80800.000 KGS",
            "CHARGABLE WEIGHT": "",
            "DESCRIPTION": "HYBRID MAIZE SEEDS (C.P.811)",
            "INVOICE NO:&  DATE": "CPS/001 DT. 08.04.2026",
            "SHIPPING LINE": "TRANS ASIA LINE ",
            "SB. NO": "2421435 DT. 16/04/2026",
            "Account Name": "",
            "Vessel Details": "HUA DA 617 V. 006W",
            "CONTAINER NUMBER ": "REGU3290956, CAIU3743948, TGBU3715873, REGU3265634",
            "CBM": "80.000",
            "Remarks": "GATEWAY CFS"
        }
        for k, v in billing_defaults.items():
            if k in self.billing_vars:
                self.billing_vars[k].set(v)
                
        self.reset_charges()
        
    def apply_parsed_data(self, data):
        """ Auto-fill the UI fields using the parsed data dictionary """
        # Address textboxes
        for k in ["shipper", "consignee", "notify", "also_notify", "cha", "forwarding_agent"]:
            if k in data and data[k]:
                if k in self.address_textboxes:
                    self.address_textboxes[k].delete("1.0", tk.END)
                    self.address_textboxes[k].insert("1.0", data[k])
                    
        # Multiline entries
        for k in ["terms_delivery_payment", "special_instructions", "bank_ac_drawback", "buyer_if_other", "marks_nos", "cargo_desc"]:
            if k in data and data[k]:
                if k in self.multiline_entries:
                    self.multiline_entries[k].delete("1.0", tk.END)
                    self.multiline_entries[k].insert("1.0", data[k])
                    
        # Standard input StringVars
        for k, v in data.items():
            if k in self.inputs and v:
                self.inputs[k].set(str(v))
                
        # Handle Job No and File Ref No directly if parsed
        if "job_no" in data and data["job_no"]:
            if "job_no" not in self.inputs:
                self.inputs["job_no"] = tk.StringVar()
            self.inputs["job_no"].set(data["job_no"])
            self.billing_vars["JOB NO:"].set(data["job_no"])
            
        if "file_ref_no" in data and data["file_ref_no"]:
            if "file_ref_no" not in self.inputs:
                self.inputs["file_ref_no"] = tk.StringVar()
            self.inputs["file_ref_no"].set(data["file_ref_no"])
            
        # Containers list
        if "containers" in data and data["containers"]:
            self.clear_container_rows()
            for c in data["containers"]:
                self.add_container_row(
                    c_no=c.get("container_no", ""),
                    s_no=c.get("seal_no", ""),
                    pkgs=c.get("pkgs", ""),
                    net_wt=c.get("net_wt", 0.0),
                    gross_wt=c.get("gross_wt", 0.0),
                    cbm=c.get("cbm", 0.0)
                )
                
        # Auto-select Client Template based on Shipper Name
        shipper_name = data.get("shipper", "").upper()
        if "COTTON HOME" in shipper_name:
            self.log_to_box("  Auto-detected Exporter: COTTON HOME. Pre-selecting Cotton Home template.")
            self.bl_template_var.set("Cotton Home")
        elif "PEMI" in shipper_name:
            self.log_to_box("  Auto-detected Exporter: PEMI EXPORTS. Pre-selecting PEMI template.")
            self.bl_template_var.set("PEMI")
        elif "SAKTHI" in shipper_name:
            self.log_to_box("  Auto-detected Exporter: SRI SAKTHI VINAYAGA FOODS. Pre-selecting Sri Sakthi template.")
            self.bl_template_var.set("Sri Sakthi")
        else:
            self.bl_template_var.set("Standard (CPS 001)")

    def sync_billing_metadata(self):
        """ Sync values from General/Container tabs into the Billing tab metadata """
        # Shipper Name
        shipper_text = self.address_textboxes["shipper"].get("1.0", tk.END).strip()
        if shipper_text:
            self.billing_vars["Shipper Name"].set(shipper_text.split("\n")[0].strip())
            
        # Job No
        job_no = ""
        if "job_no" in self.inputs:
            job_no = self.inputs["job_no"].get().strip()
        if not job_no and "file_ref_no" in self.inputs:
            job_no = self.inputs["file_ref_no"].get().strip()
        if job_no:
            self.billing_vars["JOB NO:"].set(job_no)
            
        # Destination
        pod = self.inputs["port_of_discharge"].get().strip()
        fdest = self.inputs["final_destination"].get().strip()
        if fdest:
            self.billing_vars["DESTINATION"].set(fdest.split("/")[0].strip())
        elif pod:
            self.billing_vars["DESTINATION"].set(pod.split("/")[0].strip())
            
        # Packages
        pkgs = self.inputs["total_pkgs"].get().strip()
        if pkgs:
            self.billing_vars["NO OF PACKS"].set(pkgs)
            
        # Weights
        gw = self.inputs["gross_weight"].get().strip()
        if gw:
            self.billing_vars["GROSS WEIGHT"].set(gw)
            
        # Description
        desc = self.multiline_entries["cargo_desc"].get("1.0", tk.END).strip()
        if desc:
            self.billing_vars["DESCRIPTION"].set(desc.split("\n")[0].strip())
            
        # Invoice No
        inv = self.inputs["invoice_no_date"].get().strip()
        if inv:
            self.billing_vars["INVOICE NO:&  DATE"].set(inv)
            
        # Vessel
        vessel = self.inputs["vessel_voyage"].get().strip()
        if vessel:
            self.billing_vars["Vessel Details"].set(vessel)
            
        # Shipping Bill
        sb = self.inputs["sb_no_date"].get().strip()
        if sb:
            self.billing_vars["SB. NO"].set(sb)
            
        # Container numbers comma-separated
        container_nos = [r["c_no_var"].get().strip() for r in self.container_rows if r["c_no_var"].get().strip()]
        if container_nos:
            self.billing_vars["CONTAINER NUMBER "].set(", ".join(container_nos))
            
        # CBM
        total_cbm = self.total_container_cbm.get()
        if total_cbm > 0:
            self.billing_vars["CBM"].set(f"{total_cbm:.3f}")
            
    def upload_invoice(self):
        filepath = filedialog.askopenfilename(
            filetypes=[("All Supported Files", "*.xlsx *.xls *.pdf"), ("Excel Files", "*.xlsx *.xls"), ("PDF Files", "*.pdf"), ("All Files", "*.*")],
            title="Select Invoice Document"
        )
        if not filepath:
            return
            
        self.invoice_path_var.set(filepath)
        self.lbl_status_box.config(
            text=f"Loading and parsing invoice: {os.path.basename(filepath)}...",
            fg="#2563EB", bg="#EFF6FF"
        )
        self.update_idletasks()
        self.log_to_box(f"Uploaded Invoice: {os.path.basename(filepath)}")
        
        success = False
        if filepath.lower().endswith(".pdf"):
            try:
                data = pdf_parser.parse_pdf(filepath)
                if data:
                    self.apply_parsed_data(data)
                    success = True
            except Exception as e:
                self.log_to_box(f"  Error parsing PDF Invoice: {e}")
        else:
            success = self.parse_uploaded_invoice(filepath)
            
        if success:
            self.lbl_status_box.config(
                text=f"SUCCESS: Invoice details successfully extracted from:\n{os.path.basename(filepath)}",
                fg="#16A34A", bg="#F0FDF4"
            )
            self.sync_billing_metadata()
        else:
            self.lbl_status_box.config(
                text=f"FAILED: Could not parse Invoice file:\n{os.path.basename(filepath)}",
                fg="#EF4444", bg="#FEF2F2"
            )
        self.update_idletasks()

    def upload_checklist(self):
        filepath = filedialog.askopenfilename(
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")],
            title="Select Checklist PDF Document"
        )
        if not filepath:
            return
            
        self.checklist_path_var.set(filepath)
        self.lbl_status_box.config(
            text=f"Loading and parsing checklist: {os.path.basename(filepath)}...",
            fg="#2563EB", bg="#EFF6FF"
        )
        self.update_idletasks()
        self.log_to_box(f"Uploaded Checklist: {os.path.basename(filepath)}")
        
        success = False
        try:
            data = pdf_parser.parse_pdf(filepath)
            if data:
                self.apply_parsed_data(data)
                success = True
        except Exception as e:
            self.log_to_box(f"  Error parsing PDF Checklist: {e}")
            
        if success:
            self.lbl_status_box.config(
                text=f"SUCCESS: Checklist details successfully extracted from:\n{os.path.basename(filepath)}",
                fg="#16A34A", bg="#F0FDF4"
            )
            self.sync_billing_metadata()
        else:
            self.lbl_status_box.config(
                text=f"FAILED: Could not parse Checklist file:\n{os.path.basename(filepath)}",
                fg="#EF4444", bg="#FEF2F2"
            )
        self.update_idletasks()
            
    def parse_uploaded_invoice(self, filepath):
        try:
            if filepath.lower().endswith(".xls"):
                wb = XlsWorkbookWrapper(filepath)
            else:
                wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            for sheet in wb.sheetnames:
                if "invoice" in sheet.lower():
                    ws = wb[sheet]
                    break
            
            all_kws = ["invoice no", "inv no", "invoice number", "order no", "purchase order", 
                       "ref no", "reference", "vessel", "flight no", "voyage", "port of loading", 
                       "pol", "port of discharge", "pod", "final destination", "place of delivery", 
                       "gross weight", "gross wt", "net weight", "net wt", "dimensions", "dimension", 
                       "hs code", "h.s. code", "shipping bill", "sb no", "sb date"]
            
            # Shipper block
            shipper_lines = []
            for r in range(3, 7):
                v = ws.cell(row=r, column=2).value
                if v: shipper_lines.append(str(v).strip())
            
            shipper_text = None
            if shipper_lines and len(shipper_lines) > 1:
                shipper_text = "\n".join(shipper_lines)
            else:
                shipper_text = find_block_below_keyword(ws, "shipper", 4) or find_block_below_keyword(ws, "exporter", 4)
                
            if shipper_text:
                self.address_textboxes["shipper"].delete("1.0", tk.END)
                self.address_textboxes["shipper"].insert("1.0", shipper_text)
                first_line = shipper_text.split("\n")[0].strip()
                self.billing_vars["Shipper Name"].set(first_line)
                
            # Consignee block
            consignee_lines = []
            for r in range(8, 14):
                v = ws.cell(row=r, column=2).value
                if v: consignee_lines.append(str(v).strip())
                
            consignee_text = None
            if consignee_lines and len(consignee_lines) > 1:
                consignee_text = "\n".join(consignee_lines)
            else:
                consignee_text = find_block_below_keyword(ws, "consignee", 5)
                
            if consignee_text:
                self.address_textboxes["consignee"].delete("1.0", tk.END)
                self.address_textboxes["consignee"].insert("1.0", consignee_text)
                
            # Notify block (fallback keyword check)
            notify_text = find_block_below_keyword(ws, "notify", 4)
            if notify_text:
                self.address_textboxes["notify"].delete("1.0", tk.END)
                self.address_textboxes["notify"].insert("1.0", notify_text)
                
            # Invoice No & Date
            inv_val = ws["G3"].value or ws["H3"].value
            if not inv_val:
                inv_val = find_value_by_keywords(ws, ["invoice no", "inv no", "invoice number"], ["invoice", "date"])
            if inv_val:
                self.inputs["invoice_no_date"].set(str(inv_val).strip())
                self.billing_vars["INVOICE NO:&  DATE"].set(str(inv_val).strip())
            
            # Order No & Date
            order_val = ws["G6"].value or ws["H6"].value
            if not order_val:
                order_val = find_value_by_keywords(ws, ["buyer's order", "order no", "purchase order"], all_kws)
            if order_val: self.inputs["buyer_order_no_date"].set(str(order_val).strip())
            
            # Other Ref
            ref_val = ws["H8"].value or ws["I8"].value
            if not ref_val:
                ref_val = find_value_by_keywords(ws, ["other reference", "ref no", "reference"], all_kws)
            if ref_val: self.inputs["other_ref"].set(str(ref_val).strip())
            
            # Pre-Carriage
            if ws["B15"].value: self.inputs["pre_carriage"].set(str(ws["B15"].value).strip())
            if ws["D15"].value: self.inputs["place_of_receipt"].set(str(ws["D15"].value).strip())
            
            # Origin / Dest Countries
            orig = ws["G15"].value or ws["H15"].value
            if orig: self.inputs["country_of_origin"].set(str(orig).strip())
            dest = ws["J15"].value or ws["K15"].value
            if dest: self.inputs["country_of_dest"].set(str(dest).strip())
            
            # Vessel Details
            vessel = ws["B17"].value
            if not vessel:
                vessel = find_value_by_keywords(ws, ["vessel", "flight no", "voyage"], all_kws)
            if vessel:
                self.inputs["vessel_voyage"].set(str(vessel).strip())
                self.billing_vars["Vessel Details"].set(str(vessel).strip())
            
            pol = ws["D17"].value
            if not pol:
                pol = find_value_by_keywords(ws, ["port of loading", "pol"], all_kws)
            if pol: self.inputs["port_of_loading"].set(str(pol).strip())
            
            # Terms of payment
            terms_lines = []
            for r in range(17, 20):
                val = ws.cell(row=r, column=7).value or ws.cell(row=r, column=8).value
                if val: terms_lines.append(str(val).strip())
            if terms_lines:
                self.multiline_entries["terms_delivery_payment"].delete("1.0", tk.END)
                self.multiline_entries["terms_delivery_payment"].insert("1.0", "\n".join(terms_lines))
                
            # Port of discharge / destination
            pod = ws["B19"].value
            if not pod:
                pod = find_value_by_keywords(ws, ["port of discharge", "pod"], all_kws)
            if pod: self.inputs["port_of_discharge"].set(str(pod).strip())
            
            fdest = ws["D19"].value
            if not fdest:
                fdest = find_value_by_keywords(ws, ["final destination", "place of delivery"], all_kws)
            if fdest:
                self.inputs["final_destination"].set(str(fdest).strip())
                self.billing_vars["DESTINATION"].set(str(fdest).split("/")[0].strip())
                
            # Marks & Nos
            marks_lines = []
            for r in range(22, 24):
                v = ws.cell(row=r, column=2).value
                if v: marks_lines.append(str(v).strip())
            if marks_lines:
                self.multiline_entries["marks_nos"].delete("1.0", tk.END)
                self.multiline_entries["marks_nos"].insert("1.0", "\n".join(marks_lines))
                
            # Search for dimensions / weights
            for r in range(24, ws.max_row + 1):
                cell_b = ws.cell(row=r, column=2).value
                if cell_b:
                    cell_b_str = str(cell_b).upper()
                    if "DIMENSIONS" in cell_b_str:
                        dim = ws.cell(row=r, column=5).value
                        if dim: self.inputs["dimensions"].set(str(dim).strip())
                    if "GROSS WEIGHT" in cell_b_str:
                        gw = ws.cell(row=r, column=5).value
                        if gw:
                            self.inputs["gross_weight"].set(str(gw).strip())
                            self.billing_vars["GROSS WEIGHT"].set(str(gw).strip())
                    if "NET WEIGHT" in cell_b_str:
                        nw = ws.cell(row=r, column=5).value
                        if nw: self.inputs["net_weight"].set(str(nw).strip())
                        
            # Apply keyword fallbacks for weights / dimensions if still empty
            if not self.inputs["dimensions"].get():
                dim_val = find_value_by_keywords(ws, ["dimensions", "dimension"], all_kws)
                if dim_val: self.inputs["dimensions"].set(str(dim_val).strip())
            if not self.inputs["gross_weight"].get():
                gw_val = find_value_by_keywords(ws, ["gross weight", "gross wt", "gr. wt"], all_kws)
                if gw_val:
                    self.inputs["gross_weight"].set(str(gw_val).strip())
                    self.billing_vars["GROSS WEIGHT"].set(str(gw_val).strip())
            if not self.inputs["net_weight"].get():
                nw_val = find_value_by_keywords(ws, ["net weight", "net wt"], all_kws)
                if nw_val: self.inputs["net_weight"].set(str(nw_val).strip())
                
            # HS Code
            hs_val = find_value_by_keywords(ws, ["hs code", "h.s. code"], all_kws)
            if hs_val: self.inputs["hs_code"].set(str(hs_val).strip())
            
            # SB No
            sb_val = find_value_by_keywords(ws, ["shipping bill", "sb no", "sb date"], all_kws)
            if sb_val:
                self.inputs["sb_no_date"].set(str(sb_val).strip())
                self.billing_vars["SB. NO"].set(str(sb_val).strip())
                
            # Cargo description search
            desc_lines = []
            for r in range(24, ws.max_row + 1):
                desc_val = ws.cell(row=r, column=3).value
                cell_h = ws.cell(row=r, column=8).value
                if cell_h == "TOTAL" or (desc_val and "total" in str(desc_val).lower()):
                    break
                if desc_val:
                    desc_lines.append(str(desc_val).strip())
            if desc_lines:
                cargo_desc_str = "\n".join(desc_lines)
                self.multiline_entries["cargo_desc"].delete("1.0", tk.END)
                self.multiline_entries["cargo_desc"].insert("1.0", cargo_desc_str)
                self.billing_vars["DESCRIPTION"].set(desc_lines[0])
                
            self.update_idletasks()
            return True
        except Exception as e:
            messagebox.showerror("Parsing Error", f"Failed to parse Invoice spreadsheet: {e}")
            return False
            
    def upload_packing_list(self):
        filepath = filedialog.askopenfilename(
            filetypes=[("All Supported Files", "*.xlsx *.xls *.pdf"), ("Excel Files", "*.xlsx *.xls"), ("PDF Files", "*.pdf"), ("All Files", "*.*")],
            title="Select Packing List Document"
        )
        if not filepath:
            return
            
        self.packing_path_var.set(filepath)
        self.lbl_status_box.config(
            text=f"Loading and parsing packing list: {os.path.basename(filepath)}...",
            fg="#2563EB", bg="#EFF6FF"
        )
        self.update_idletasks()
        self.log_to_box(f"Uploaded Packing List: {os.path.basename(filepath)}")
        
        success = False
        if filepath.lower().endswith(".pdf"):
            try:
                data = pdf_parser.parse_pdf(filepath)
                if data:
                    self.apply_parsed_data(data)
                    success = True
            except Exception as e:
                self.log_to_box(f"  Error parsing PDF Packing List: {e}")
        else:
            success = self.parse_uploaded_packing_list(filepath)
            
        if success:
            self.lbl_status_box.config(
                text=f"SUCCESS: Containers & Package lists extracted from:\n{os.path.basename(filepath)}",
                fg="#16A34A", bg="#F0FDF4"
            )
            self.sync_billing_metadata()
        else:
            self.lbl_status_box.config(
                text=f"FAILED: Could not parse Packing List file:\n{os.path.basename(filepath)}",
                fg="#EF4444", bg="#FEF2F2"
            )
        self.update_idletasks()
            
    def parse_uploaded_packing_list(self, filepath):
        try:
            if filepath.lower().endswith(".xls"):
                wb = XlsWorkbookWrapper(filepath)
            else:
                wb = openpyxl.load_workbook(filepath, data_only=True)
            containers = []
            container_pattern = re.compile(r'\b[A-Z]{4}\d{7}\b')
            
            # Pattern-based search for containers on all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        if val:
                            val_str = str(val).strip()
                            match_obj = container_pattern.search(val_str)
                            if match_obj:
                                c_no = match_obj.group(0)
                                s_no = ""
                                pkgs = ""
                                net = 0.0
                                gross = 0.0
                                cbm = 0.0
                                
                                # Extract adjacent cells in the same row
                                s_val = ws.cell(row=r, column=c+1).value
                                if s_val is not None: s_no = str(s_val).strip()
                                p_val = ws.cell(row=r, column=c+2).value
                                if p_val is not None: pkgs = str(p_val).strip()
                                    
                                n_val = ws.cell(row=r, column=c+3).value
                                if n_val is not None:
                                    try: net = float(str(n_val).replace(",", "").strip())
                                    except: pass
                                    
                                g_val = ws.cell(row=r, column=c+4).value
                                if g_val is not None:
                                    try: gross = float(str(g_val).replace(",", "").strip())
                                    except: pass
                                    
                                c_val = ws.cell(row=r, column=c+5).value
                                if c_val is not None:
                                    try: cbm = float(str(c_val).strip())
                                    except: pass
                                    
                                containers.append({
                                    "container_no": c_no,
                                    "seal_no": s_no,
                                    "pkgs": pkgs,
                                    "net_wt": net,
                                    "gross_wt": gross,
                                    "cbm": cbm
                                })
            
            if containers:
                self.clear_container_rows()
                for c in containers:
                    self.add_container_row(
                        c_no=c["container_no"],
                        s_no=c["seal_no"],
                        pkgs=c["pkgs"],
                        net_wt=c["net_wt"],
                        gross_wt=c["gross_wt"],
                        cbm=c["cbm"]
                    )
                
                # Update billing fields
                c_nos = [c["container_no"] for c in containers]
                self.billing_vars["CONTAINER NUMBER "].set(", ".join(c_nos))
                
                # Packages sum
                total_pkgs = 0
                pkg_type = "BAGS"
                for c in containers:
                    try:
                        p_str = c["pkgs"].split()
                        total_pkgs += int(p_str[0])
                        if len(p_str) > 1:
                            pkg_type = p_str[1]
                    except:
                        pass
                if total_pkgs > 0:
                    self.billing_vars["NO OF PACKS"].set(f"{total_pkgs} PACKAGES")
                    self.inputs["total_pkgs"].set(f"{total_pkgs} {pkg_type}")
                    
                # Weights sum
                total_nw = sum(c["net_wt"] for c in containers)
                total_gw = sum(c["gross_wt"] for c in containers)
                if total_nw > 0: self.inputs["net_weight"].set(f"{total_nw:,.2f} KGS")
                if total_gw > 0:
                    self.inputs["gross_weight"].set(f"{total_gw:,.2f} Kgs")
                    self.billing_vars["GROSS WEIGHT"].set(f"{total_gw:,.3f} KGS")
                    
                # CBM sum
                total_cbm = sum(c["cbm"] for c in containers)
                if total_cbm > 0:
                    self.billing_vars["CBM"].set(f"{total_cbm:.3f}")
                    
                self.recalculate_totals()
                self.recalculate_billing()
                self.update_idletasks()
                return True
            else:
                messagebox.showinfo("Parsing Notice", "No container numbers matched pattern [A-Z]{4}\\d{7} in the sheet. You can enter them manually.")
                self.update_idletasks()
                return True
        except Exception as e:
            messagebox.showerror("Parsing Error", f"Failed to parse Packing List spreadsheet: {e}")
            return False
            
    def log_to_box(self, msg):
        if hasattr(self, "log_txt"):
            self.log_txt.insert(tk.END, msg + "\n")
            self.log_txt.see(tk.END)
            self.update_idletasks()
            
    def recalculate_totals(self):
        total_cnt_pkgs = 0
        total_cnt_net = 0.0
        total_cnt_gross = 0.0
        total_cnt_cbm = 0.0
        
        for row in self.container_rows:
            pkgs_str = row["pkgs_var"].get().strip()
            try:
                p_val = int(pkgs_str.split()[0])
                total_cnt_pkgs += p_val
            except Exception:
                pass
                
            try:
                net_val = float(row["net_ent"].get().strip())
                total_cnt_net += net_val
            except ValueError:
                pass
            try:
                gross_val = float(row["gross_ent"].get().strip())
                total_cnt_gross += gross_val
            except ValueError:
                pass
            try:
                cbm_val = float(row["cbm_ent"].get().strip())
                total_cnt_cbm += cbm_val
            except ValueError:
                pass
                
        self.total_container_pkgs.set(f"{total_cnt_pkgs} BAGS")
        self.total_container_net.set(total_cnt_net)
        self.total_container_gross.set(total_cnt_gross)
        self.total_container_cbm.set(total_cnt_cbm)
        
        self.lbl_container_totals.config(
            text=f"Total Containers: {len(self.container_rows)} | Total Pkgs: {total_cnt_pkgs} BAGS | Total Net: {total_cnt_net:,.2f} kg | Total Gross: {total_cnt_gross:,.2f} kg | Total CBM: {total_cnt_cbm:.3f}"
        )
        
    def recalculate_billing(self):
        total_taxable = 0.0
        total_cgst = 0.0
        total_sgst = 0.0
        
        for row in self.charge_rows:
            try:
                val_str = row["amount_ent"].get().strip()
                amount = float(val_str) if val_str else 0.0
            except ValueError:
                amount = 0.0
                
            try:
                cgst_str = row["cgst_rate_ent"].get().strip()
                cgst_rate = float(cgst_str) if cgst_str else 0.0
            except ValueError:
                cgst_rate = 0.0
                
            try:
                sgst_str = row["sgst_rate_ent"].get().strip()
                sgst_rate = float(sgst_str) if sgst_str else 0.0
            except ValueError:
                sgst_rate = 0.0
                
            row["cgst_rate_var"].set(cgst_rate)
            row["sgst_rate_var"].set(sgst_rate)
            
            cgst_amt = round((amount * cgst_rate) / 100.0, 2)
            sgst_amt = round((amount * sgst_rate) / 100.0, 2)
            row_total = round(amount + cgst_amt + sgst_amt, 2)
            
            row["cgst_amt_var"].set(cgst_amt)
            row["sgst_amt_var"].set(sgst_amt)
            row["total_var"].set(row_total)
            
            row["lbl_cgst_amt"].config(text=f"{cgst_amt:.2f}")
            row["lbl_sgst_amt"].config(text=f"{sgst_amt:.2f}")
            row["lbl_total"].config(text=f"{row_total:.2f}")
            
            total_taxable += amount
            total_cgst += cgst_amt
            total_sgst += sgst_amt
            
        grand_total = round(total_taxable + total_cgst + total_sgst, 2)
        tds = round(total_taxable * 0.02, 2)
        total_payable = round(grand_total - tds, 2)
        
        self.total_taxable_var.set(total_taxable)
        self.total_cgst_var.set(total_cgst)
        self.total_sgst_var.set(total_sgst)
        self.grand_total_var.set(grand_total)
        self.tds_var.set(tds)
        self.total_payable_var.set(total_payable)
        
        # Format variables
        self.fmt_taxable.set(f"Rs. {total_taxable:,.2f}")
        self.fmt_cgst.set(f"Rs. {total_cgst:,.2f}")
        self.fmt_sgst.set(f"Rs. {total_sgst:,.2f}")
        self.fmt_grand.set(f"Rs. {grand_total:,.2f}")
        self.fmt_tds.set(f"Rs. {tds:,.2f}")
        self.fmt_payable.set(f"Rs. {total_payable:,.2f}")

    def fetch_live_exchange_rate(self):
        try:
            url = "https://open.er-api.com/v6/latest/USD"
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=5) as response:
                data = json.loads(response.read().decode())
                if data and data.get("result") == "success":
                    inr_rate = data["rates"].get("INR")
                    if inr_rate:
                        self.after(0, lambda: self.of_ex_rate_var.set(f"{inr_rate:.2f}"))
        except Exception:
            pass
        
    def get_data_payload(self):
        payload = {}
        # Fetch standard inputs
        for k, v in self.inputs.items():
            payload[k] = v.get().strip()
        # Fetch address text boxes
        for k, txt in self.address_textboxes.items():
            payload[k] = txt.get("1.0", tk.END).strip()
        # Fetch multiline entries
        for k, txt in self.multiline_entries.items():
            payload[k] = txt.get("1.0", tk.END).strip()
            
        return payload
        
    def generate_all(self):
        self.recalculate_totals()
        self.recalculate_billing()
        dest_dir = self.export_path_var.get().strip()
        
        if not dest_dir:
            messagebox.showerror("Error", "Please specify a destination folder.")
            return
            
        os.makedirs(dest_dir, exist_ok=True)
        
        self.log_txt.delete("1.0", tk.END)
        self.log_to_box("--- Starting Logistics Document Generation ---")
        self.log_to_box(f"Destination: {dest_dir}\n")
        
        data = self.get_data_payload()
        
        # Get containers details
        containers = []
        for r in self.container_rows:
            c_no = r["c_no_var"].get().strip()
            s_no = r["s_no_var"].get().strip()
            pkgs = r["pkgs_var"].get().strip()
            net = r["net_ent"].get().strip()
            gross = r["gross_ent"].get().strip()
            cbm = r["cbm_ent"].get().strip()
            
            if c_no:
                containers.append({
                    "container_no": c_no,
                    "seal_no": s_no,
                    "pkgs": pkgs,
                    "net_wt": net,
                    "gross_wt": gross,
                    "cbm": cbm
                })
                
        # Define paths to templates
        data["template_bl_path"] = resource_path("BL INSTRUCTION CPS 001_template.xlsx")
        data["template_awb_path"] = resource_path("AWB Instructions_template.xlsx")
        
        # Verify templates exist
        for k in ["template_bl_path", "template_awb_path"]:
            p = data[k]
            if not os.path.exists(p):
                local_name = os.path.basename(p)
                local_p = os.path.abspath(local_name)
                if os.path.exists(local_p):
                    data[k] = local_p
                else:
                    self.log_to_box(f"ERROR: Template file {local_name} not found in bundle or local folder!")
                    messagebox.showerror("Template Missing", f"Template file {local_name} was not found!")
                    return
                    
        success_count = 0
        
        # Clean invoice number
        inv_no_raw = data.get("invoice_no_date", "")
        clean_inv = "REF"
        if inv_no_raw:
            part = inv_no_raw.split("DT")[0].split("dt")[0].strip()
            if "/" in part:
                parts = part.split("/")
                if len(parts[0]) > 2:
                    part = parts[0]
                else:
                    part = parts[0] + parts[1]
            if "-" in part:
                subparts = part.split("-")
                if len(subparts[0]) > 2:
                    part = subparts[0]
            clean_inv = re.sub(r'[^A-Za-z0-9]', '', part)
            
        bl_template = self.bl_template_var.get()
        if bl_template == "Cotton Home":
            bl_filename = f"COTTON_BL Instruction {clean_inv}.xlsx"
        elif bl_template == "PEMI":
            bl_filename = f"PEMI_BL Instruction {clean_inv}.xlsx"
        elif bl_template == "Sri Sakthi":
            bl_filename = f"SRI SAKTHI_BL Instruction {clean_inv}.xlsx"
        else:
            bl_filename = f"Generated_BL_Instructions_{clean_inv}.xlsx"
            
        # Determine dynamic billing name
        shipper_name = self.billing_vars["Shipper Name"].get().strip().upper()
        if "COTTON HOME" in shipper_name:
            shipper_clean = "COTTON HOME"
        elif "PEMI" in shipper_name:
            shipper_clean = "PEMI"
        elif "SAKTHI" in shipper_name:
            shipper_clean = "SRI SAKTHI"
        else:
            shipper_clean = shipper_name if shipper_name else "CLIENT"
            
        job_no = self.billing_vars["JOB NO:"].get().strip()
        job_no_clean = job_no.split("-")[0].strip() if "-" in job_no else job_no
        if not job_no_clean:
            job_no_clean = clean_inv
            
        pkgs = self.billing_vars["NO OF PACKS"].get().strip()
        pkgs_clean = pkgs if pkgs else "PACKAGES"
        
        container_text = ""
        c_nos = [r["c_no_var"].get().strip() for r in self.container_rows if r["c_no_var"].get().strip()]
        if c_nos:
            container_text = f"{len(c_nos)} X 40 FT_FCL" if "40" in self.bl_template_var.get() or "PEMI" in self.bl_template_var.get() or "Cotton" in self.bl_template_var.get() else f"{len(c_nos)} X 20 FT_FCL"
        else:
            container_text = "FCL"
            
        bill_filename = f"{shipper_clean}_JOB NO {job_no_clean} - {pkgs_clean} {container_text}.xlsx"
        bill_filename = re.sub(r'[\\/*?:"<>|]', "", bill_filename)
        
        awb_filename = f"Generated_AWB_Instructions_{clean_inv}.xlsx"
        
        # 1. BL Instructions
        if self.gen_bl.get():
            try:
                out_path = os.path.join(dest_dir, bl_filename)
                self.log_to_box(f"Generating BL Instructions: {os.path.basename(out_path)}...")
                self.generate_bl_instructions_file(data, containers, out_path)
                self.log_to_box(f"  Successfully generated BL Instructions: {os.path.basename(out_path)}!\n")
                success_count += 1
            except Exception as e:
                self.log_to_box(f"  Error generating BL Instructions: {e}\n")
                
        # 2. Billing Invoice
        if self.gen_bill.get():
            try:
                out_path = os.path.join(dest_dir, bill_filename)
                self.log_to_box(f"Generating Billing Invoice: {os.path.basename(out_path)}...")
                self.generate_billing_invoice_file(out_path)
                self.log_to_box(f"  Successfully generated Billing Invoice: {os.path.basename(out_path)}!\n")
                success_count += 1
            except Exception as e:
                self.log_to_box(f"  Error generating Billing Invoice: {e}\n")
                
        # 3. AWB Instructions
        if self.gen_awb.get():
            try:
                out_path = os.path.join(dest_dir, awb_filename)
                self.log_to_box(f"Generating AWB Instructions: {os.path.basename(out_path)}...")
                self.generate_awb_instructions_file(data, out_path)
                self.log_to_box(f"  Successfully generated AWB Instructions: {os.path.basename(out_path)}!\n")
                success_count += 1
            except Exception as e:
                self.log_to_box(f"  Error generating AWB Instructions: {e}\n")
                
        self.log_to_box(f"--- Export Finished: {success_count} documents compiled ---")
        if success_count > 0:
            messagebox.showinfo("Success", f"{success_count} documents successfully generated in:\n{dest_dir}")
            
    def generate_bl_instructions_file(self, data, containers, out_path):
        template_choice = self.bl_template_var.get()
        self.log_to_box(f"Selected BL Template: {template_choice}")
        
        app_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        templates_dir = os.path.join(app_dir, "Templates")
        if not os.path.exists(templates_dir):
            templates_dir = os.path.abspath("Templates")
        if not os.path.exists(templates_dir):
            templates_dir = "."
            
        if template_choice == "Cotton Home":
            template_file = os.path.join(templates_dir, "COTTON HOME_BL INSTRUCTION - CHX03_FCL 1X40.xlsx")
            self.generate_bl_cotton_home(data, containers, template_file, out_path)
        elif template_choice == "PEMI":
            template_file = os.path.join(templates_dir, "PEMI - BL INSTRUCTION - PE044 -  855 CTNS - GENOA.xlsx")
            self.generate_bl_pemi(data, containers, template_file, out_path)
        elif template_choice == "Sri Sakthi":
            template_file = os.path.join(templates_dir, "SRI SAKTHI - BL INSTRUCTION_ 432 PKGS_ SINGAPORE.xlsx")
            self.generate_bl_sri_sakthi(data, containers, template_file, out_path)
        else:
            # Standard CPS 001 template choice
            self.generate_bl_standard(data, containers, data["template_bl_path"], out_path)

    def generate_bl_cotton_home(self, data, containers, template_file, out_path):
        if not os.path.exists(template_file):
            template_file = os.path.basename(template_file)
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active
        ws.views.sheetView[0].showGridLines = True
        
        # Shipper
        write_address_block(ws, 6, 1, data.get("shipper", ""), max_rows=9)
        # Consignee
        write_address_block(ws, 16, 1, data.get("consignee", ""), max_rows=5)
        # Notify & Also Notify
        write_address_block(ws, 22, 1, data.get("notify", ""), max_rows=6)
        also_notify = data.get("also_notify", "")
        if also_notify and not also_notify.upper().startswith("ALSO NOTIFY"):
            also_notify = "ALSO NOTIFY:- " + also_notify
        write_address_block(ws, 21, 5, also_notify, max_rows=7)
        
        # Voyage Details
        ws["A28"] = data.get("vessel_voyage", "")
        ws["C28"] = data.get("port_of_loading", "")
        ws["A31"] = data.get("port_of_discharge", "")
        ws["C31"] = data.get("final_destination", "")
        
        # Marks & Nos
        ws["A37"] = "CTN NOS"
        pkg_from = data.get("package_from", 1)
        pkg_to = data.get("package_to", 1)
        try:
            from_str = f"{int(pkg_from):02d}"
        except:
            from_str = str(pkg_from)
        try:
            to_str = str(int(pkg_to))
        except:
            to_str = str(pkg_to)
        ws["A38"] = f"{from_str} TO {to_str}"
        
        # Total Packages
        total_pkgs = data.get("total_pkgs", "")
        try:
            val = int(total_pkgs.split()[0].replace(",", ""))
            ws["B35"] = val
        except:
            ws["B35"] = total_pkgs
            
        # Standard Container Text
        ws["C35"] = f"{len(containers)} X 40'HC FCL / FCL - CONTAINER ONLY"
        try:
            val = int(total_pkgs.split()[0].replace(",", ""))
            ws["C37"] = f"[{num_to_words(val).replace('SAY USD ', '').replace(' ONLY', '')} CARTONS ONLY]"
        except:
            ws["C37"] = f"[{total_pkgs} ONLY]"
            
        # Cargo Description
        cargo_desc = data.get("cargo_desc", "")
        if len(cargo_desc.split("\n")) > 6 and data.get("cargo_items"):
            categories = []
            for item in data.get("cargo_items", []):
                desc = item.get("description", "")
                if "-" in desc:
                    cat = desc.split("-")[0].strip()
                else:
                    cat = " ".join(desc.split()[:4])
                # Remove parentheses content and multiple spaces
                cat = re.sub(r'\(.*?\)', '', cat).strip()
                cat = re.sub(r'\s+', ' ', cat)
                if cat and cat.upper() not in [c.upper() for c in categories]:
                    categories.append(cat)
            if categories:
                if len(categories) > 1:
                    cargo_desc = ", ".join(categories[:-1]) + " AND " + categories[-1]
                else:
                    cargo_desc = categories[0]
        write_wrapped_text_to_rows(ws, 39, 3, cargo_desc, max_rows=4, limit=65)
        
        # Weights
        ws["H35"] = clean_weight(data.get("gross_weight", ""))
        ws["H42"] = clean_weight(data.get("net_weight", ""))
        
        # CBM
        total_cbm = 0.0
        for c in containers:
            if c.get("cbm"):
                try:
                    total_cbm += float(str(c["cbm"]).replace(",", "").strip())
                except:
                    pass
        ws["I35"] = total_cbm if total_cbm > 0 else None
        
        # Containers and Seals
        c_nos = [c["container_no"] for c in containers if c["container_no"]]
        s_nos = [c["seal_no"] for c in containers if c["seal_no"]]
        ws["A52"] = ", ".join(c_nos)
        ws["A55"] = ", ".join(s_nos)
        
        # References
        ws["C52"] = "INV NO: " + data.get("invoice_no_date", "")
        ws["C53"] = "S.B.NO: " + data.get("sb_no_date", "")
        
        # Unique HS Codes
        unique_hs = []
        for item in data.get("cargo_items", []):
            hs = item.get("hs_code")
            if hs and hs not in unique_hs:
                unique_hs.append(hs)
        if not unique_hs and data.get("hs_code"):
            unique_hs = [data.get("hs_code")]
        hs_text = "HS CODE : " + ", ".join(unique_hs)
        write_wrapped_text_to_rows(ws, 47, 3, hs_text, max_rows=3, limit=65)
        
        ws["C55"] = '"FREIGHT ' + data.get("payment_mode", "COLLECT") + '"'
        
        wb.save(out_path)

    def generate_bl_pemi(self, data, containers, template_file, out_path):
        if not os.path.exists(template_file):
            template_file = os.path.basename(template_file)
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active
        ws.views.sheetView[0].showGridLines = True
        
        # Shipper
        write_address_block(ws, 6, 1, data.get("shipper", ""), max_rows=7)
        # Consignee
        write_address_block(ws, 14, 1, data.get("consignee", ""), max_rows=7)
        # Notify
        write_address_block(ws, 22, 1, data.get("notify", ""), max_rows=7)
        
        ws["A30"] = data.get("vessel_voyage", "")
        ws["C30"] = data.get("port_of_loading", "")
        ws["A33"] = data.get("port_of_discharge", "")
        ws["C33"] = data.get("final_destination", "")
        
        ws["A37"] = data.get("marks_nos", "")
        pkg_from = data.get("package_from", 1)
        pkg_to = data.get("package_to", 1)
        ws["A40"] = f"{pkg_from} TO {pkg_to}"
        
        total_pkgs = data.get("total_pkgs", "")
        try:
            val = int(total_pkgs.split()[0].replace(",", ""))
            ws["B37"] = val
        except:
            ws["B37"] = total_pkgs
            
        ws["C37"] = f"{len(containers)} X 40'HC - FCL / FCL"
        try:
            val = int(total_pkgs.split()[0].replace(",", ""))
            ws["C39"] = f"[{num_to_words(val).replace('SAY USD ', '').replace(' ONLY', '')} CARTONS ONLY]"
        except:
            ws["C39"] = f"[{total_pkgs} ONLY]"
            
        # Cargo Description
        cargo_desc = data.get("cargo_desc", "")
        if len(cargo_desc.split("\n")) > 6 and data.get("cargo_items"):
            categories = []
            for item in data.get("cargo_items", []):
                desc = item.get("description", "")
                if "-" in desc:
                    cat = desc.split("-")[0].strip()
                else:
                    cat = " ".join(desc.split()[:4])
                # Remove parentheses content and multiple spaces
                cat = re.sub(r'\(.*?\)', '', cat).strip()
                cat = re.sub(r'\s+', ' ', cat)
                if cat and cat.upper() not in [c.upper() for c in categories]:
                    categories.append(cat)
            if categories:
                if len(categories) > 1:
                    cargo_desc = ", ".join(categories[:-1]) + " AND " + categories[-1]
                else:
                    cargo_desc = categories[0]
        write_wrapped_text_to_rows(ws, 42, 3, cargo_desc, max_rows=6, limit=65)
        
        # Weights
        ws["E37"] = clean_weight(data.get("gross_weight", ""))
        net_wt = clean_weight(data.get("net_weight", ""))
        ws["C54"] = f"NET WEIGHT: {net_wt:.3f} KGS" if isinstance(net_wt, float) else f"NET WEIGHT: {net_wt}"
        
        # CBM
        total_cbm = 0.0
        for c in containers:
            if c.get("cbm"):
                try:
                    total_cbm += float(str(c["cbm"]).replace(",", "").strip())
                except:
                    pass
        ws["F37"] = total_cbm if total_cbm > 0 else None
        
        c_nos = [c["container_no"] for c in containers if c["container_no"]]
        s_nos = [c["seal_no"] for c in containers if c["seal_no"]]
        ws["A49"] = ", ".join(c_nos)
        ws["A54"] = ", ".join(s_nos)
        
        ws["C52"] = "INV No: " + data.get("invoice_no_date", "")
        ws["C53"] = "SB NO: " + data.get("sb_no_date", "")
        ws["C56"] = '"FREIGHT ' + data.get("payment_mode", "COLLECT") + '"'
        
        wb.save(out_path)

    def generate_bl_sri_sakthi(self, data, containers, template_file, out_path):
        if not os.path.exists(template_file):
            template_file = os.path.basename(template_file)
        wb = openpyxl.load_workbook(template_file)
        ws = wb.active
        ws.views.sheetView[0].showGridLines = True
        
        # Shipper
        write_address_block(ws, 6, 1, data.get("shipper", ""), max_rows=9)
        # Consignee
        write_address_block(ws, 17, 1, data.get("consignee", ""), max_rows=8)
        # Notify
        write_address_block(ws, 26, 1, data.get("notify", ""), max_rows=8)
        
        ws["A35"] = data.get("vessel_voyage", "")
        ws["C35"] = data.get("port_of_loading", "")
        ws["A38"] = data.get("port_of_discharge", "")
        ws["C38"] = data.get("final_destination", "")
        
        write_address_block(ws, 42, 1, data.get("marks_nos", ""), max_rows=5)
        total_pkgs = data.get("total_pkgs", "")
        try:
            val = int(total_pkgs.split()[0].replace(",", ""))
            ws["B42"] = val
        except:
            ws["B42"] = total_pkgs
            
        ws["C42"] = f"{len(containers)} X 20' GP / FCL - FCL"
        try:
            val = int(total_pkgs.split()[0].replace(",", ""))
            ws["C44"] = f"({num_to_words(val).replace('SAY USD ', '').replace(' ONLY', '')} PACKAGES ONLY)"
        except:
            ws["C44"] = f"({total_pkgs} ONLY)"
            
        # Cargo Description
        cargo_desc = data.get("cargo_desc", "")
        if len(cargo_desc.split("\n")) > 6 and data.get("cargo_items"):
            categories = []
            for item in data.get("cargo_items", []):
                desc = item.get("description", "")
                if "-" in desc:
                    cat = desc.split("-")[0].strip()
                else:
                    cat = " ".join(desc.split()[:4])
                # Remove parentheses content and multiple spaces
                cat = re.sub(r'\(.*?\)', '', cat).strip()
                cat = re.sub(r'\s+', ' ', cat)
                if cat and cat.upper() not in [c.upper() for c in categories]:
                    categories.append(cat)
            if categories:
                if len(categories) > 1:
                    cargo_desc = ", ".join(categories[:-1]) + " AND " + categories[-1]
                else:
                    cargo_desc = categories[0]
        write_wrapped_text_to_rows(ws, 48, 3, cargo_desc, max_rows=5, limit=65)
        
        # Unique HS Codes
        unique_hs = []
        for item in data.get("cargo_items", []):
            hs = item.get("hs_code")
            if hs and hs not in unique_hs:
                unique_hs.append(hs)
        if not unique_hs and data.get("hs_code"):
            unique_hs = [data.get("hs_code")]
        hs_text = "HS CODE : " + ", ".join(unique_hs)
        write_wrapped_text_to_rows(ws, 46, 3, hs_text, max_rows=2, limit=65)
        
        # Weights
        ws["F42"] = clean_weight(data.get("gross_weight", ""))
        net_wt = clean_weight(data.get("net_weight", ""))
        ws["C57"] = f"NT.WT: {net_wt:.3f} KGS" if isinstance(net_wt, float) else f"NT.WT: {net_wt}"
        
        # CBM
        total_cbm = 0.0
        for c in containers:
            if c.get("cbm"):
                try:
                    total_cbm += float(str(c["cbm"]).replace(",", "").strip())
                except:
                    pass
        ws["G42"] = total_cbm if total_cbm > 0 else None
        
        c_nos = [c["container_no"] for c in containers if c["container_no"]]
        s_nos = [c["seal_no"] for c in containers if c["seal_no"]]
        ws["A52"] = ", ".join(c_nos)
        ws["A55"] = ", ".join(s_nos)
        
        ws["C55"] = "INV NO: " + data.get("invoice_no_date", "")
        ws["C56"] = "S.B.NO: " + data.get("sb_no_date", "")
        ws["C60"] = '"FREIGHT ' + data.get("payment_mode", "PREPAID") + '"'
        
        wb.save(out_path)

    def generate_bl_standard(self, data, containers, template_file, out_path):
        wb = openpyxl.load_workbook(template_file)
        ws = wb["Table 1"]
        ws.views.sheetView[0].showGridLines = True
        
        # Shipper
        ws.cell(row=2, column=1, value="SHIPPER:\n" + data.get("shipper", ""))
        # CHA
        ws.cell(row=2, column=4, value="CHA\n" + data.get("cha", ""))
        # Consignee
        ws.cell(row=3, column=1, value="CONSIGNEE: (Complete Address and fax#)\n" + data.get("consignee", ""))
        # Forwarding Agent
        ws.cell(row=3, column=4, value="FORWARDING AGENT ADDRESS:\n" + data.get("forwarding_agent", ""))
        # Notify
        ws.cell(row=4, column=1, value="NOTIFY:       (Complete Address and Fax#)\n" + data.get("notify", ""))
        # Also Notify
        ws.cell(row=4, column=4, value="ALSO NOTIFY (Complete Address and Fax#)\n" + data.get("also_notify", ""))
        
        # Voyage details
        ws.cell(row=5, column=1, value="VESSEL AND VOYAGE NO :  " + data.get("vessel_voyage", ""))
        ws.cell(row=5, column=4, value="PORT OF LOADING: " + data.get("port_of_loading", ""))
        ws.cell(row=6, column=1, value="PORT OF DISCHARGE: " + data.get("port_of_discharge", ""))
        ws.cell(row=6, column=4, value="FINAL DESTINATION: " + data.get("final_destination", ""))
        
        # Marks & Nos, Packages
        ws.cell(row=8, column=1, value=data.get("marks_nos", ""))
        ws.cell(row=8, column=2, value=data.get("type_of_packing", ""))
        
        # Container totals & texts
        total_cbm = sum(float(c["cbm"]) for c in containers if c["cbm"])
        
        # Total pkgs parsing
        try:
            pkgs_sum = sum(int(c["pkgs"].split()[0]) for c in containers if c["pkgs"] and c["pkgs"].split()[0].isdigit())
            pkgs_text = f"{pkgs_sum} PP BAGS"
        except Exception:
            pkgs_text = data.get("total_pkgs", "")
            
        ws.cell(row=8, column=3, value=f"{len(containers)}X20' FCL CONTAINER SAID TO CONTAIN\nTOTAL: {pkgs_text}")
        ws.cell(row=9, column=3, value=data.get("cargo_desc", ""))
        
        ws.cell(row=11, column=3, value="INV NOS :")
        ws.cell(row=12, column=3, value="  " + data.get("invoice_no_date", ""))
        ws.cell(row=14, column=3, value="HS CODE : " + data.get("hs_code", ""))
        ws.cell(row=16, column=3, value="SB NO : " + data.get("sb_no_date", ""))
        
        # Weights
        ws.cell(row=8, column=5, value=data.get("gross_weight", ""))
        ws.cell(row=10, column=5, value="NET. WT.")
        ws.cell(row=11, column=5, value=data.get("net_weight", ""))
        ws.cell(row=8, column=6, value=f"{total_cbm:.3f} CBM" if total_cbm > 0 else "")
        
        # Container table
        first_item_row = 20
        template_rows = 4
        
        if len(containers) > template_rows:
            rows_to_insert = len(containers) - template_rows
            for _ in range(rows_to_insert):
                insert_row_copy_style(ws, 24, 23)
                
        for idx, c in enumerate(containers):
            r = first_item_row + idx
            ws.cell(row=r, column=1, value=c["container_no"])
            ws.cell(row=r, column=2, value=c["seal_no"])
            ws.cell(row=r, column=3, value=c["pkgs"])
            
            try:
                net_val = float(c["net_wt"])
                ws.cell(row=r, column=4, value=net_val)
                ws.cell(row=r, column=4).number_format = "#,##0"
            except ValueError:
                ws.cell(row=r, column=4, value=c["net_wt"])
                
            try:
                gross_val = float(c["gross_wt"])
                ws.cell(row=r, column=5, value=gross_val)
                ws.cell(row=r, column=5).number_format = "#,##0"
            except ValueError:
                ws.cell(row=r, column=5, value=c["gross_wt"])
                
            try:
                cbm_val = float(c["cbm"])
                ws.cell(row=r, column=6, value=cbm_val)
                ws.cell(row=r, column=6).number_format = "0.00"
            except ValueError:
                ws.cell(row=r, column=6, value=c["cbm"])
                
        # Clear leftover rows
        if len(containers) < template_rows:
            for r in range(first_item_row + len(containers), first_item_row + template_rows):
                ws.cell(row=r, column=1, value=None)
                ws.cell(row=r, column=2, value=None)
                ws.cell(row=r, column=3, value=None)
                ws.cell(row=r, column=4, value=None)
                ws.cell(row=r, column=5, value=None)
                ws.cell(row=r, column=6, value=None)
                
        wb.save(out_path)
        
    def generate_billing_invoice_file(self, out_path):
        app_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(app_dir, "Billing", "Billing_Invoice.xlsx")
        if not os.path.exists(template_path):
            template_path = os.path.join("Billing", "Billing_Invoice.xlsx")
        if not os.path.exists(template_path):
            template_path = os.path.join(app_dir, "Billing_Invoice.xlsx")
            
        if not os.path.exists(template_path):
            self.log_to_box(f"Warning: Billing_Invoice.xlsx template not found at {template_path}. Generating from scratch.")
            self.generate_billing_invoice_file_scratch(out_path)
            return

        self.log_to_box(f"Using template: {template_path}")
        wb = openpyxl.load_workbook(template_path)
        ws = wb["Invoice Details"]
        ws.views.sheetView[0].showGridLines = True
        
        # Gather meta and calculations
        total_taxable = self.total_taxable_var.get()
        total_cgst = self.total_cgst_var.get()
        total_sgst = self.total_sgst_var.get()
        grand_total = self.grand_total_var.get()
        tds = self.tds_var.get()
        total_payable = self.total_payable_var.get()
        
        # Write Shipper Name in E2
        ws["E2"] = self.billing_vars["Shipper Name"].get()
        
        # Write Metadata
        metadata_fields = [
            "JOB NO:", "BOOKING REF.  NO", "MBL ", "DESTINATION", "NO OF PACKS",
            "GROSS WEIGHT", "CHARGABLE WEIGHT", "DESCRIPTION", "INVOICE NO:&  DATE",
            "SHIPPING LINE", "SB. NO", "Account Name", "Vessel Details",
            "CONTAINER NUMBER ", "CBM", "Remarks"
        ]
        for idx, key in enumerate(metadata_fields):
            row_num = 4 + idx
            val = self.billing_vars[key].get()
            if key == "BOOKING REF.  NO" and str(val).strip().isdigit():
                ws.cell(row=row_num, column=5, value=int(str(val).strip()))
            else:
                ws.cell(row=row_num, column=5, value=str(val).strip())
                
        # Handle rows adjustment
        N = len(self.charge_rows)
        if N < 15:
            ws.delete_rows(21 + N, 15 - N)
        elif N > 15:
            ws.insert_rows(35, N - 15)
            # Copy style from row 34 to inserted rows
            for r in range(35, 21 + N):
                for col in range(3, 8):
                    src_cell = ws.cell(row=34, column=col)
                    dest_cell = ws.cell(row=r, column=col)
                    dest_cell.font = openpyxl.styles.Font(
                        name=src_cell.font.name, size=src_cell.font.size,
                        bold=src_cell.font.bold, italic=src_cell.font.italic,
                        color=src_cell.font.color, underline=src_cell.font.underline
                    )
                    dest_cell.border = openpyxl.styles.Border(
                        left=src_cell.border.left, right=src_cell.border.right,
                        top=src_cell.border.top, bottom=src_cell.border.bottom
                    )
                    dest_cell.alignment = openpyxl.styles.Alignment(
                        horizontal=src_cell.alignment.horizontal,
                        vertical=src_cell.alignment.vertical,
                        wrap_text=src_cell.alignment.wrap_text
                    )
                    dest_cell.number_format = src_cell.number_format
                    
        # Populate charges
        for idx, row_data in enumerate(self.charge_rows):
            r = 21 + idx
            desc = row_data["desc_var"].get()
            try:
                amt_str = row_data["amount_ent"].get().strip()
                amount = float(amt_str) if amt_str else 0.0
            except:
                amount = 0.0
                
            cgst_val = round((amount * row_data["cgst_rate_var"].get()) / 100.0, 2)
            sgst_val = round((amount * row_data["sgst_rate_var"].get()) / 100.0, 2)
            tax_val = cgst_val + sgst_val
            row_total = amount + tax_val
            
            ws.cell(row=r, column=3, value=idx + 1)
            ws.cell(row=r, column=4, value=desc)
            ws.cell(row=r, column=5, value=amount)
            ws.cell(row=r, column=6, value=tax_val)
            ws.cell(row=r, column=7, value=row_total)
            
            # Format numbers
            ws.cell(row=r, column=5).number_format = "#,##0.00"
            ws.cell(row=r, column=6).number_format = "#,##0.00"
            ws.cell(row=r, column=7).number_format = "#,##0.00"
            
        # Overwrite the totals at rows 21+N to 26+N
        ws.cell(row=21 + N, column=7, value=total_taxable)
        ws.cell(row=22 + N, column=7, value=total_cgst)
        ws.cell(row=23 + N, column=7, value=total_sgst)
        ws.cell(row=24 + N, column=7, value=grand_total)
        ws.cell(row=25 + N, column=7, value=tds)
        ws.cell(row=26 + N, column=7, value=total_payable)
        
        # Ensure number formatting for totals
        for idx in range(6):
            ws.cell(row=21 + N + idx, column=7).number_format = "#,##0.00"
            
        wb.save(out_path)

    def generate_billing_invoice_file_scratch(self, out_path):
        # Gather meta and calculations
        total_taxable = self.total_taxable_var.get()
        total_cgst = self.total_cgst_var.get()
        total_sgst = self.total_sgst_var.get()
        grand_total = self.grand_total_var.get()
        tds = self.tds_var.get()
        total_payable = self.total_payable_var.get()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice Details"
        ws.views.sheetView[0].showGridLines = True
        
        # Estimate CONTENT column width
        max_desc_len = 35.0
        for r in self.charge_rows:
            d = r["desc_var"].get()
            if d: max_desc_len = max(max_desc_len, len(d))
        col_d_width = min(45.0, max(35.0, max_desc_len + 3))
        
        column_widths = {
            'A': 6.1, 'B': 13.0, 'C': 8.0, 'D': col_d_width, 'E': 18.0, 'F': 18.0, 'G': 20.0
        }
        for col, w in column_widths.items():
            ws.column_dimensions[col].width = w
            
        font_title = Font(name="Calibri", size=16, bold=True)
        font_metadata_val = Font(name="Calibri", size=14)
        font_metadata_lbl = Font(name="Calibri", size=14)
        font_metadata_idx = Font(name="Calibri", size=11, bold=True)
        
        font_header = Font(name="Calibri", size=16, bold=True)
        font_header_sino = Font(name="Cambria", size=12, bold=True)
        font_cell = Font(name="Calibri", size=16)
        font_cell_red = Font(name="Calibri", size=16, color="FF0000")
        font_cell_bold = Font(name="Calibri", size=16, bold=True)
        font_cell_bold_red = Font(name="Calibri", size=16, bold=True, color="FF0000")
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        thin_border_side = Side(border_style="thin", color="D1D5DB")
        medium_border_side = Side(border_style="medium", color="000000")
        
        # Shipper Name Box
        shipper_name = self.billing_vars["Shipper Name"].get()
        shipper_lines = max(1, len(shipper_name) // 50 + (1 if len(shipper_name) % 50 > 0 else 0))
        ws.row_dimensions[2].height = 21.6 * shipper_lines
        
        ws.merge_cells("C2:D2")
        ws["C2"].value = "SHIPPER NAME "
        ws["C2"].font = font_title
        ws["C2"].alignment = align_center
        
        ws.merge_cells("E2:G2")
        ws["E2"].value = shipper_name
        ws["E2"].font = font_title
        ws["E2"].alignment = align_left
        
        # Apply borders row 2
        for col_idx in range(3, 8):
            cell = ws.cell(row=2, column=col_idx)
            cell.border = Border(
                top=medium_border_side, bottom=medium_border_side,
                left=medium_border_side if col_idx in (3, 5) else None,
                right=medium_border_side if col_idx in (4, 7) else None
            )
            
        ws.row_dimensions[3].height = 7.5
        
        # Metadata rows 4 to 19
        metadata_fields = [
            "JOB NO:", "BOOKING REF.  NO", "MBL ", "DESTINATION", "NO OF PACKS",
            "GROSS WEIGHT", "CHARGABLE WEIGHT", "DESCRIPTION", "INVOICE NO:&  DATE",
            "SHIPPING LINE", "SB. NO", "Account Name", "Vessel Details",
            "CONTAINER NUMBER ", "CBM", "Remarks"
        ]
        
        for idx, key in enumerate(metadata_fields):
            row_num = 4 + idx
            val = self.billing_vars[key].get()
            val_str = str(val)
            val_lines = max(1, len(val_str) // 50 + (1 if len(val_str) % 50 > 0 else 0))
            ws.row_dimensions[row_num].height = 18.0 * val_lines
            
            # C index
            ws.cell(row=row_num, column=3, value=idx + 1).font = font_metadata_idx
            ws.cell(row=row_num, column=3).alignment = align_center
            
            # D label
            ws.cell(row=row_num, column=4, value=key).font = font_metadata_lbl
            ws.cell(row=row_num, column=4).alignment = align_left
            
            # E:G Merged value
            ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=7)
            cell_e = ws.cell(row=row_num, column=5)
            if key == "BOOKING REF.  NO" and val_str.isdigit():
                cell_e.value = int(val_str)
            else:
                cell_e.value = val_str
            cell_e.font = font_metadata_val
            cell_e.alignment = align_left
            
            for c_idx in range(3, 8):
                cell = ws.cell(row=row_num, column=c_idx)
                cell.border = Border(
                    top=medium_border_side if row_num == 4 else thin_border_side,
                    bottom=medium_border_side if row_num == 19 else thin_border_side,
                    left=medium_border_side if c_idx == 3 else thin_border_side,
                    right=medium_border_side if c_idx == 7 else thin_border_side
                )
                
        # Header Row 20
        ws.row_dimensions[20].height = 18.0
        headers_config = [
            ("C20", "Si.No", font_header_sino, align_center),
            ("D20", "CONTENT", font_header, align_center),
            ("E20", "TAXABLE VALUE", font_header, align_center),
            ("F20", "TAX AMOUNT", font_header, align_center),
            ("G20", "TOTAL AMOUNT", font_header, align_center)
        ]
        for coord, txt, font_s, align in headers_config:
            ws[coord].value = txt
            ws[coord].font = font_s
            ws[coord].alignment = align
            
        for c_idx in range(3, 8):
            cell = ws.cell(row=20, column=c_idx)
            cell.border = Border(
                top=thin_border_side, bottom=thin_border_side,
                left=medium_border_side if c_idx == 3 else thin_border_side,
                right=medium_border_side if c_idx == 7 else thin_border_side
            )
            
        # Charges rows 21 onwards
        current_row = 21
        for idx, row_data in enumerate(self.charge_rows):
            desc = row_data["desc_var"].get()
            desc_limit = int(col_d_width - 3)
            desc_lines = max(1, len(desc) // desc_limit + (1 if len(desc) % desc_limit > 0 else 0))
            ws.row_dimensions[current_row].height = 21.0 * desc_lines
            
            try:
                amt_str = row_data["amount_ent"].get().strip()
                amount = float(amt_str) if amt_str else 0.0
            except:
                amount = 0.0
                
            ws.cell(row=current_row, column=3, value=idx + 1).font = font_cell
            ws.cell(row=current_row, column=3).alignment = align_center
            
            ws.cell(row=current_row, column=4, value=desc).font = font_cell
            ws.cell(row=current_row, column=4).alignment = align_left
            
            ws.cell(row=current_row, column=5, value=amount).font = font_cell_red
            ws.cell(row=current_row, column=5).alignment = align_right
            ws.cell(row=current_row, column=5).number_format = "#,##0.00"
            
            cgst_val = round((amount * row_data["cgst_rate_var"].get()) / 100.0, 2)
            sgst_val = round((amount * row_data["sgst_rate_var"].get()) / 100.0, 2)
            tax_val = cgst_val + sgst_val
            row_total = amount + tax_val
            
            ws.cell(row=current_row, column=6, value=tax_val).font = font_cell_red
            ws.cell(row=current_row, column=6).alignment = align_right
            ws.cell(row=current_row, column=6).number_format = "#,##0.00"
            
            ws.cell(row=current_row, column=7, value=row_total).font = font_cell_red
            ws.cell(row=current_row, column=7).alignment = align_right
            ws.cell(row=current_row, column=7).number_format = "#,##0.00"
            
            for c_idx in range(3, 8):
                cell = ws.cell(row=current_row, column=c_idx)
                cell.border = Border(
                    top=thin_border_side, bottom=thin_border_side,
                    left=medium_border_side if c_idx == 3 else thin_border_side,
                    right=medium_border_side if c_idx == 7 else thin_border_side
                )
            current_row += 1
            
        # Summary Rows
        summary_start = current_row
        summary_configs = [
            ("Total Taxable Value", total_taxable),
            ("Total CGST", total_cgst),
            ("Total SGST", total_sgst),
            ("Grand Total with Taxes", grand_total),
            ("TDS @ 2%", tds), 
            ("Total Payable Amount", total_payable)
        ]
        
        for idx, (label, val) in enumerate(summary_configs):
            row_num = summary_start + idx
            ws.row_dimensions[row_num].height = 21.0
            
            ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=6)
            cell_lbl = ws.cell(row=row_num, column=3, value=label)
            cell_lbl.font = font_cell_bold
            cell_lbl.alignment = align_right
            
            cell_val = ws.cell(row=row_num, column=7, value=val)
            if label in ["Grand Total with Taxes", "Total Payable Amount"]:
                cell_lbl.font = font_cell_bold
                cell_val.font = font_cell_bold_red
            else:
                cell_val.font = font_cell_red
                
            cell_val.alignment = align_right
            cell_val.number_format = "#,##0.00"
            
            is_last = (idx == len(summary_configs) - 1)
            for c_idx in range(3, 8):
                cell = ws.cell(row=row_num, column=c_idx)
                cell.border = Border(
                    top=thin_border_side,
                    bottom=medium_border_side if is_last else thin_border_side,
                    left=medium_border_side if c_idx == 3 else thin_border_side,
                    right=medium_border_side if c_idx == 7 else thin_border_side
                )
                
        wb.save(out_path)
        
    def generate_awb_instructions_file(self, data, out_path):
        wb = openpyxl.load_workbook(data["template_awb_path"])
        ws = wb["Sheet1 (2)"]
        ws.views.sheetView[0].showGridLines = True
        
        # Final destination
        ws.cell(row=6, column=3, value=data["final_destination"])
        # Payment Mode
        ws.cell(row=6, column=8, value=data["payment_mode"])
        
        # Shipper
        shipper_lines = data["shipper"].split("\n")
        for idx, line in enumerate(shipper_lines[:4]):
            ws.cell(row=9 + idx, column=1, value=line)
            
        # Payment mode
        ws.cell(row=8, column=8, value=data["payment_mode"])
        # AWB / Invoice fallback
        ws.cell(row=9, column=8, value=data["invoice_no_date"])
        # Pkgs
        ws.cell(row=11, column=7, value=data["total_pkgs"])
        ws.cell(row=11, column=9, value=data["type_of_packing"])
        
        # Consignee
        consignee_lines = data["consignee"].split("\n")
        for idx, line in enumerate(consignee_lines[:4]):
            ws.cell(row=14 + idx, column=1, value=line)
            
        # Codes
        ws.cell(row=13, column=8, value=data["rbi_code"])
        ws.cell(row=14, column=8, value=data["ie_code"])
        ws.cell(row=15, column=8, value=data["gross_weight"])
        
        # Notify
        notify_lines = data["notify"].split("\n")
        for idx, line in enumerate(notify_lines[:4]):
            ws.cell(row=19 + idx, column=1, value=line)
            
        # Dimensions
        dim_lines = data["dimensions"].split("\n")
        for idx, line in enumerate(dim_lines[:2]):
            ws.cell(row=19 + idx, column=6, value=line)
            
        # GR No
        ws.cell(row=22, column=6, value=data["gr_no_date"])
        
        # Account
        acc_lines = data["account"].split("\n")
        for idx, line in enumerate(acc_lines[:4]):
            ws.cell(row=24 + idx, column=1, value=line)
            
        # Marks
        marks_lines = data["marks_nos"].split("\n")
        for idx, line in enumerate(marks_lines[:3]):
            ws.cell(row=24 + idx, column=6, value=line)
            
        # Cargo desc
        cargo_lines = data["cargo_desc"].split("\n")
        for idx, line in enumerate(cargo_lines[:5]):
            ws.cell(row=28 + idx, column=6, value=line)
            
        ws.cell(row=29, column=1, value=data["invoice_no_date"])
        ws.cell(row=32, column=1, value=data["contents"])
        
        bank_lines = data["bank_ac_drawback"].split("\n")
        for idx, line in enumerate(bank_lines[:2]):
            ws.cell(row=34 + idx, column=1, value=line)
            
        spec_lines = data["special_instructions"].split("\n")
        for idx, line in enumerate(spec_lines[:2]):
            ws.cell(row=34 + idx, column=6, value=line)
            
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['H'].width = 17
        
        wb.save(out_path)

if __name__ == "__main__":
    app = UnifiedGeneratorApp()
    app.mainloop()
