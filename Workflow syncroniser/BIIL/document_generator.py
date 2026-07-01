import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def num_to_words(num):
    """ Convert numeric amount to USD words format """
    units = ["", "ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN", "EIGHT", "NINE", "TEN", 
             "ELEVEN", "TWELVE", "THIRTEEN", "FOURTEEN", "FIFTEEN", "SIXTEEN", "SEVENTEEN", "EIGHTEEN", "NINETEEN"]
    tens = ["", "", "TWENTY", "THIRTY", "FORTY", "FIFTY", "SIXTY", "SEVENTY", "EIGHTY", "NINETY"]
    
    def helper(n):
        if n < 20:
            return units[n]
        elif n < 100:
            return tens[n // 10] + (" " + units[n % 10] if n % 10 != 0 else "")
        elif n < 1000:
            return units[n // 100] + " HUNDRED" + (" AND " + helper(n % 100) if n % 100 != 0 else "")
        elif n < 100000:
            return helper(n // 1000) + " THOUSAND" + (" " + helper(n % 1000) if n % 1000 != 0 else "")
        else:
            return str(n)
            
    if num == 0:
        return "SAY USD ZERO ONLY"
        
    parts = f"{num:.2f}".split(".")
    dollars = int(parts[0])
    cents = int(parts[1])
    
    words = "SAY USD " + helper(dollars)
    if cents > 0:
        words += " AND CENTS " + helper(cents) + " ONLY"
    else:
        words += " ONLY"
    return words

def insert_row_copy_style(ws, row_idx, copy_from_row):
    """ Insert a row in openpyxl and copy cell styles from a reference row """
    ws.insert_rows(row_idx)
    for col_idx in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=copy_from_row, column=col_idx)
        dest_cell = ws.cell(row=row_idx, column=col_idx)
        if src_cell.has_style:
            dest_cell.font = Font(
                name=src_cell.font.name,
                size=src_cell.font.size,
                bold=src_cell.font.bold,
                italic=src_cell.font.italic,
                color=src_cell.font.color,
                underline=src_cell.font.underline
            )
            dest_cell.border = Border(
                left=src_cell.border.left,
                right=src_cell.border.right,
                top=src_cell.border.top,
                bottom=src_cell.border.bottom
            )
            dest_cell.fill = PatternFill(
                fill_type=src_cell.fill.fill_type,
                start_color=src_cell.fill.start_color,
                end_color=src_cell.fill.end_color
            )
            dest_cell.alignment = Alignment(
                horizontal=src_cell.alignment.horizontal,
                vertical=src_cell.alignment.vertical,
                wrap_text=src_cell.alignment.wrap_text
            )
            dest_cell.number_format = src_cell.number_format

class LogisticsApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Logistics Document Generator")
        self.geometry("1150x800")
        self.configure(bg="#F1F5F9")
        
        # Styles
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure(".", background="#F1F5F9", font=("Segoe UI", 10))
        self.style.configure("TLabel", background="#F1F5F9", foreground="#334155")
        self.style.configure("Header.TLabel", font=("Segoe UI", 16, "bold"), foreground="#1E293B", background="#F1F5F9")
        self.style.configure("Card.TFrame", background="#FFFFFF", relief="flat", borderwidth=1)
        self.style.configure("CardTitle.TLabel", font=("Segoe UI", 12, "bold"), foreground="#1E293B", background="#FFFFFF")
        
        self.style.configure("TNotebook", background="#F1F5F9", borderwidth=0)
        self.style.configure("TNotebook.Tab", background="#CBD5E1", foreground="#475569", padding=[15, 6], font=("Segoe UI", 10, "bold"))
        self.style.map("TNotebook.Tab", background=[("selected", "#FFFFFF")], foreground=[("selected", "#2563EB")])
        
        self.style.configure("Primary.TButton", background="#2563EB", foreground="#FFFFFF", borderwidth=0, font=("Segoe UI", 10, "bold"), padding=[12, 6])
        self.style.map("Primary.TButton", background=[("active", "#1D4ED8"), ("pressed", "#1E40AF")])
        
        self.style.configure("Secondary.TButton", background="#64748B", foreground="#FFFFFF", borderwidth=0, font=("Segoe UI", 10), padding=[10, 4])
        self.style.map("Secondary.TButton", background=[("active", "#475569"), ("pressed", "#334155")])
        
        self.style.configure("Danger.TButton", background="#EF4444", foreground="#FFFFFF", borderwidth=0, font=("Segoe UI", 9, "bold"), padding=[4, 2])
        self.style.map("Danger.TButton", background=[("active", "#DC2626"), ("pressed", "#B91C1C")])

        # Variables holding text inputs
        self.inputs = {}
        
        # Lists for cargo items and container rows
        self.cargo_rows = []
        self.container_rows = []
        
        # Summary variables
        self.total_cargo_qty = tk.IntVar(value=0)
        self.total_cargo_amount = tk.DoubleVar(value=0.0)
        self.total_cargo_net = tk.DoubleVar(value=0.0)
        self.total_cargo_gross = tk.DoubleVar(value=0.0)
        
        self.total_container_pkgs = tk.StringVar(value="0 BAGS")
        self.total_container_net = tk.DoubleVar(value=0.0)
        self.total_container_gross = tk.DoubleVar(value=0.0)
        self.total_container_cbm = tk.DoubleVar(value=0.0)
        
        # Document Selection Variables
        self.gen_invoice = tk.BooleanVar(value=True)
        self.gen_packing = tk.BooleanVar(value=True)
        self.gen_bl = tk.BooleanVar(value=True)
        self.gen_awb = tk.BooleanVar(value=True)
        
        self.export_path_var = tk.StringVar()
        default_export = r"c:\Users\vkmsd\OneDrive\Desktop\B,I,I,L\Generated"
        self.export_path_var.set(default_export)
        
        self.init_variables()
        self.create_widgets()
        self.load_defaults()
        self.recalculate_totals()
        
    def init_variables(self):
        fields = [
            "shipper", "consignee", "notify", "also_notify", "cha", "forwarding_agent",
            "vessel_voyage", "port_of_loading", "port_of_discharge", "final_destination",
            "country_of_origin", "country_of_dest", "terms_delivery_payment",
            "invoice_no_date", "buyer_order_no_date", "pre_carriage", "place_of_receipt",
            "sb_no_date", "hs_code", "payment_mode", "dimensions", "gross_weight",
            "net_weight", "special_instructions", "rbi_code", "ie_code", "account",
            "gr_no_date", "other_ref", "buyer_if_other", "marks_nos", "contents",
            "bank_ac_drawback", "type_of_packing", "total_pkgs", "cargo_desc", "invoice_value"
        ]
        for f in fields:
            self.inputs[f] = tk.StringVar()
            
    def create_widgets(self):
        # Top Title Bar
        title_frame = tk.Frame(self, bg="#1E293B", height=60)
        title_frame.pack(fill="x", side="top")
        title_frame.pack_propagate(False)
        
        title_lbl = tk.Label(title_frame, text="Logistics Document Suite", font=("Segoe UI", 18, "bold"), fg="#F8FAFC", bg="#1E293B")
        title_lbl.pack(side="left", padx=20, pady=12)
        
        subtitle_lbl = tk.Label(title_frame, text="Excel Invoice & Shipping Instructions Generator", font=("Segoe UI", 10), fg="#94A3B8", bg="#1E293B")
        subtitle_lbl.pack(side="left", padx=10, pady=18)
        
        # Tabs container
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=20, pady=15)
        
        self.tab_shipper = ttk.Frame(self.notebook)
        self.tab_voyage = ttk.Frame(self.notebook)
        self.tab_cargo = ttk.Frame(self.notebook)
        self.tab_container = ttk.Frame(self.notebook)
        self.tab_export = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab_shipper, text="1. Shipper & Addresses")
        self.notebook.add(self.tab_voyage, text="2. Voyage & References")
        self.notebook.add(self.tab_cargo, text="3. Cargo Items (Inv/PL)")
        self.notebook.add(self.tab_container, text="4. Containers (BL/PL)")
        self.notebook.add(self.tab_export, text="5. Export & Generate")
        
        self.build_shipper_tab()
        self.build_voyage_tab()
        self.build_cargo_tab()
        self.build_container_tab()
        self.build_export_tab()
        
    def build_shipper_tab(self):
        card = ttk.Frame(self.tab_shipper, style="Card.TFrame")
        card.pack(fill="both", expand=True, padx=20, pady=20)
        
        lbl = ttk.Label(card, text="Entity Names & Addresses", style="CardTitle.TLabel")
        lbl.grid(row=0, column=0, columnspan=2, sticky="w", padx=20, pady=15)
        
        # Let's use Grid to lay out multiline Text boxes
        address_fields = [
            ("Shipper Details", "shipper"),
            ("Consignee Details", "consignee"),
            ("Notify Party Details", "notify"),
            ("Also Notify Details", "also_notify"),
            ("CHA Address/Details", "cha"),
            ("Forwarding Agent", "forwarding_agent")
        ]
        
        self.address_textboxes = {}
        for idx, (label, key) in enumerate(address_fields):
            r = 1 + (idx // 2)
            c = (idx % 2) * 2
            
            lbl_w = tk.Label(card, text=label, font=("Segoe UI", 10, "bold"), fg="#475569", bg="#FFFFFF", anchor="w")
            lbl_w.grid(row=r*2-1, column=c, padx=20, pady=(10, 2), sticky="w")
            
            txt = tk.Text(card, font=("Segoe UI", 10), width=45, height=5, relief="solid", bd=1)
            txt.grid(row=r*2, column=c, columnspan=2, padx=20, pady=(2, 10), sticky="ew")
            self.address_textboxes[key] = txt
            
        btn = ttk.Button(card, text="Continue to Voyage & Ref →", style="Primary.TButton", command=lambda: self.notebook.select(1))
        btn.grid(row=8, column=3, sticky="e", padx=20, pady=20)
        
    def build_voyage_tab(self):
        # We need a scrollable container for this tab since there are a lot of fields
        canvas_frame = tk.Frame(self.tab_voyage, bg="#F1F5F9")
        canvas_frame.pack(fill="both", expand=True)
        
        canvas = tk.Canvas(canvas_frame, bg="#F1F5F9", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="#F1F5F9")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=20, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        
        card = ttk.Frame(scrollable_frame, style="Card.TFrame")
        card.pack(fill="both", expand=True, padx=5, pady=5)
        
        lbl = ttk.Label(card, text="Voyage, Payment & Reference Details", style="CardTitle.TLabel")
        lbl.grid(row=0, column=0, columnspan=4, sticky="w", padx=20, pady=15)
        
        # Setup entries grid
        entries_config = [
            ("Vessel / Flight No:", "vessel_voyage", 1, 0),
            ("Port of Loading:", "port_of_loading", 1, 2),
            ("Port of Discharge:", "port_of_discharge", 2, 0),
            ("Final Destination:", "final_destination", 2, 2),
            ("Country of Origin:", "country_of_origin", 3, 0),
            ("Country of Final Dest:", "country_of_dest", 3, 2),
            ("Invoice No & Date:", "invoice_no_date", 4, 0),
            ("Buyer's Order No & Date:", "buyer_order_no_date", 4, 2),
            ("Pre-Carriage by:", "pre_carriage", 5, 0),
            ("Place of Receipt:", "place_of_receipt", 5, 2),
            ("Shipping Bill No & Date:", "sb_no_date", 6, 0),
            ("HS Code:", "hs_code", 6, 2),
            ("Payment Mode (Prepaid/Collect):", "payment_mode", 7, 0),
            ("RBI Code No:", "rbi_code", 7, 2),
            ("IE Code No:", "ie_code", 8, 0),
            ("Invoice Value (Currency/Val):", "invoice_value", 8, 2),
            ("G.R. No. / Date:", "gr_no_date", 9, 0),
            ("Other Reference(s):", "other_ref", 9, 2),
            ("Type of Packing:", "type_of_packing", 10, 0),
            ("Total Pkgs (Text e.g. 432 PKGS):", "total_pkgs", 10, 2),
            ("Gross Weight (e.g. 500 KGS):", "gross_weight", 11, 0),
            ("Net Weight (e.g. 480 KGS):", "net_weight", 11, 2),
            ("Dimensions (CMS):", "dimensions", 12, 0),
            ("Contents:", "contents", 12, 2),
        ]
        
        for label_text, key, r, c in entries_config:
            lbl_w = ttk.Label(card, text=label_text, background="#FFFFFF", font=("Segoe UI", 10, "bold"), width=30, anchor="w")
            lbl_w.grid(row=r, column=c, padx=(20, 5), pady=8, sticky="w")
            
            ent = ttk.Entry(card, textvariable=self.inputs[key], font=("Segoe UI", 10), width=32)
            ent.grid(row=r, column=c + 1, padx=(5, 20), pady=8, sticky="ew")
            
        # Add a couple of textareas for multiline strings
        textareas = [
            ("Terms of Delivery & Payment:", "terms_delivery_payment", 13, 0),
            ("Special Instructions:", "special_instructions", 14, 0),
            ("Bank Account Details:", "bank_ac_drawback", 15, 0),
            ("Buyer (If other than consignee):", "buyer_if_other", 16, 0),
            ("Marks & Numbers:", "marks_nos", 17, 0),
            ("Cargo Description:", "cargo_desc", 18, 0)
        ]
        
        self.multiline_entries = {}
        for idx, (label_text, key, r, c) in enumerate(textareas):
            lbl_w = tk.Label(card, text=label_text, font=("Segoe UI", 10, "bold"), fg="#475569", bg="#FFFFFF", anchor="w")
            lbl_w.grid(row=r, column=0, padx=20, pady=(10, 2), sticky="w")
            
            txt = tk.Text(card, font=("Segoe UI", 10), width=85, height=3, relief="solid", bd=1)
            txt.grid(row=r, column=1, columnspan=3, padx=20, pady=(2, 10), sticky="ew")
            self.multiline_entries[key] = txt
            
        btn = ttk.Button(card, text="Continue to Cargo Items →", style="Primary.TButton", command=lambda: self.notebook.select(2))
        btn.grid(row=20, column=3, sticky="e", padx=20, pady=20)
        
    def build_cargo_tab(self):
        card = ttk.Frame(self.tab_cargo, style="Card.TFrame")
        card.pack(fill="both", expand=True, padx=20, pady=20)
        
        top_ctrl = tk.Frame(card, bg="#FFFFFF")
        top_ctrl.pack(fill="x", padx=15, pady=10)
        
        lbl = ttk.Label(top_ctrl, text="Cargo Items (For Invoice and Packing List)", style="CardTitle.TLabel")
        lbl.pack(side="left")
        
        btn_add = ttk.Button(top_ctrl, text="+ Add Cargo Item", style="Primary.TButton", command=self.add_cargo_row)
        btn_add.pack(side="right", padx=5)
        
        # Table Header
        headers_frame = tk.Frame(card, bg="#F8FAFC", height=30)
        headers_frame.pack(fill="x", padx=15, pady=(5, 0))
        
        headers = [
            ("Sl.No", 50, "center"),
            ("DESCRIPTION OF GOODS", 350, "w"),
            ("QTY (NOS)", 100, "e"),
            ("RATE (USD)", 100, "e"),
            ("NET WT (KGS)", 100, "e"),
            ("GROSS WT (KGS)", 100, "e"),
            ("TOTAL AMOUNT", 120, "e"),
            ("ACTION", 80, "center")
        ]
        
        for name, width, anchor in headers:
            lbl = tk.Label(headers_frame, text=name, font=("Segoe UI", 9, "bold"), fg="#475569", bg="#F8FAFC", width=width, anchor=anchor)
            lbl.pack(side="left", padx=5, fill="y")
            lbl.pack_propagate(False)
            lbl.config(width=int(width/8))
            
        # Scrollable container
        canvas_frame = tk.Frame(card, bg="#FFFFFF")
        canvas_frame.pack(fill="both", expand=True, padx=15, pady=5)
        
        self.cargo_canvas = tk.Canvas(canvas_frame, bg="#FFFFFF", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.cargo_canvas.yview)
        
        self.cargo_scroll_frame = tk.Frame(self.cargo_canvas, bg="#FFFFFF")
        self.cargo_scroll_frame.bind(
            "<Configure>",
            lambda e: self.cargo_canvas.configure(scrollregion=self.cargo_canvas.bbox("all"))
        )
        
        self.cargo_window = self.cargo_canvas.create_window((0, 0), window=self.cargo_scroll_frame, anchor="nw")
        self.cargo_canvas.configure(yscrollcommand=scrollbar.set)
        
        self.cargo_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.cargo_canvas.bind('<Configure>', lambda e: self.cargo_canvas.itemconfig(self.cargo_window, width=e.width))
        
        # Totals bar
        self.cargo_totals_frame = tk.Frame(card, bg="#F8FAFC", height=40)
        self.cargo_totals_frame.pack(fill="x", padx=15, pady=(5, 15))
        
        self.lbl_cargo_totals = tk.Label(
            self.cargo_totals_frame, 
            text="Total Qty: 0 | Total Amount: $0.00 | Total Net: 0.00 kg | Total Gross: 0.00 kg", 
            font=("Segoe UI", 10, "bold"), 
            fg="#1E293B", 
            bg="#F8FAFC", 
            anchor="e"
        )
        self.lbl_cargo_totals.pack(fill="both", expand=True, padx=20)
        
        self.bind_all("<KeyRelease>", lambda e: self.recalculate_totals())
        
    def add_cargo_row(self, desc="", qty=0, rate=0.0, net_wt=0.0, gross_wt=0.0):
        row_frame = tk.Frame(self.cargo_scroll_frame, bg="#FFFFFF", pady=4)
        row_frame.pack(fill="x", expand=True)
        
        idx = len(self.cargo_rows) + 1
        
        # Row variables
        desc_var = tk.StringVar(value=desc)
        qty_var = tk.IntVar(value=qty)
        rate_var = tk.DoubleVar(value=rate)
        net_var = tk.DoubleVar(value=net_wt)
        gross_var = tk.DoubleVar(value=gross_wt)
        
        widgets = []
        
        # Sl.No
        lbl_idx = tk.Label(row_frame, text=str(idx), font=("Segoe UI", 10), bg="#FFFFFF", fg="#64748B", width=6, anchor="center")
        lbl_idx.pack(side="left", padx=5)
        widgets.append(lbl_idx)
        
        # Description
        ent_desc = ttk.Entry(row_frame, textvariable=desc_var, font=("Segoe UI", 10), width=42)
        ent_desc.pack(side="left", padx=5)
        widgets.append(ent_desc)
        
        # Qty
        ent_qty = ttk.Entry(row_frame, font=("Segoe UI", 10), width=10, justify="right")
        ent_qty.insert(0, str(qty))
        ent_qty.pack(side="left", padx=5)
        widgets.append(ent_qty)
        
        # Rate
        ent_rate = ttk.Entry(row_frame, font=("Segoe UI", 10), width=10, justify="right")
        ent_rate.insert(0, f"{rate:.2f}")
        ent_rate.pack(side="left", padx=5)
        widgets.append(ent_rate)
        
        # Net Wt
        ent_net = ttk.Entry(row_frame, font=("Segoe UI", 10), width=10, justify="right")
        ent_net.insert(0, f"{net_wt:.2f}")
        ent_net.pack(side="left", padx=5)
        widgets.append(ent_net)
        
        # Gross Wt
        ent_gross = ttk.Entry(row_frame, font=("Segoe UI", 10), width=10, justify="right")
        ent_gross.insert(0, f"{gross_wt:.2f}")
        ent_gross.pack(side="left", padx=5)
        widgets.append(ent_gross)
        
        # Total Amt
        lbl_total = tk.Label(row_frame, text="0.00", font=("Segoe UI", 10, "bold"), bg="#FFFFFF", fg="#0F172A", width=14, anchor="e")
        lbl_total.pack(side="left", padx=5)
        widgets.append(lbl_total)
        
        # Remove button
        btn_del = ttk.Button(row_frame, text="Remove", style="Danger.TButton", command=lambda: self.remove_cargo_row(row_frame))
        btn_del.pack(side="left", padx=5)
        widgets.append(btn_del)
        
        row_data = {
            "frame": row_frame,
            "widgets": widgets,
            "desc_var": desc_var,
            "qty_ent": ent_qty,
            "rate_ent": ent_rate,
            "net_ent": ent_net,
            "gross_ent": ent_gross,
            "lbl_total": lbl_total,
            "lbl_idx": lbl_idx
        }
        self.cargo_rows.append(row_data)
        self.recalculate_totals()
        
    def remove_cargo_row(self, frame):
        for idx, row in enumerate(self.cargo_rows):
            if row["frame"] == frame:
                for w in row["widgets"]:
                    w.destroy()
                row["frame"].destroy()
                self.cargo_rows.pop(idx)
                break
                
        for idx, row in enumerate(self.cargo_rows):
            row["lbl_idx"].config(text=str(idx+1))
            
        self.recalculate_totals()
        
    def build_container_tab(self):
        card = ttk.Frame(self.tab_container, style="Card.TFrame")
        card.pack(fill="both", expand=True, padx=20, pady=20)
        
        top_ctrl = tk.Frame(card, bg="#FFFFFF")
        top_ctrl.pack(fill="x", padx=15, pady=10)
        
        lbl = ttk.Label(top_ctrl, text="Container & Package List (For BL Instructions and Packing List)", style="CardTitle.TLabel")
        lbl.pack(side="left")
        
        btn_add = ttk.Button(top_ctrl, text="+ Add Container", style="Primary.TButton", command=self.add_container_row)
        btn_add.pack(side="right", padx=5)
        
        # Table Header
        headers_frame = tk.Frame(card, bg="#F8FAFC", height=30)
        headers_frame.pack(fill="x", padx=15, pady=(5, 0))
        
        headers = [
            ("Sl.No", 50, "center"),
            ("CONTAINER NUMBER", 180, "w"),
            ("SEAL NUMBER", 180, "w"),
            ("NO. OF PKGS (e.g. 400 BAGS)", 180, "w"),
            ("NET WT (KGS)", 110, "e"),
            ("GROSS WT (KGS)", 110, "e"),
            ("CBM", 100, "e"),
            ("ACTION", 80, "center")
        ]
        
        for name, width, anchor in headers:
            lbl = tk.Label(headers_frame, text=name, font=("Segoe UI", 9, "bold"), fg="#475569", bg="#F8FAFC", width=width, anchor=anchor)
            lbl.pack(side="left", padx=5, fill="y")
            lbl.pack_propagate(False)
            lbl.config(width=int(width/8))
            
        # Scrollable container
        canvas_frame = tk.Frame(card, bg="#FFFFFF")
        canvas_frame.pack(fill="both", expand=True, padx=15, pady=5)
        
        self.container_canvas = tk.Canvas(canvas_frame, bg="#FFFFFF", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.container_canvas.yview)
        
        self.container_scroll_frame = tk.Frame(self.container_canvas, bg="#FFFFFF")
        self.container_scroll_frame.bind(
            "<Configure>",
            lambda e: self.container_canvas.configure(scrollregion=self.container_canvas.bbox("all"))
        )
        
        self.container_window = self.container_canvas.create_window((0, 0), window=self.container_scroll_frame, anchor="nw")
        self.container_canvas.configure(yscrollcommand=scrollbar.set)
        
        self.container_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.container_canvas.bind('<Configure>', lambda e: self.container_canvas.itemconfig(self.container_window, width=e.width))
        
        # Totals Bar
        self.container_totals_frame = tk.Frame(card, bg="#F8FAFC", height=40)
        self.container_totals_frame.pack(fill="x", padx=15, pady=(5, 15))
        
        self.lbl_container_totals = tk.Label(
            self.container_totals_frame, 
            text="Total Containers: 0 | Total Pkgs: 0 BAGS | Total Net: 0.00 kg | Total Gross: 0.00 kg | Total CBM: 0.000", 
            font=("Segoe UI", 10, "bold"), 
            fg="#1E293B", 
            bg="#F8FAFC", 
            anchor="e"
        )
        self.lbl_container_totals.pack(fill="both", expand=True, padx=20)
        
    def add_container_row(self, c_no="", s_no="", pkgs="", net_wt=0.0, gross_wt=0.0, cbm=0.0):
        row_frame = tk.Frame(self.container_scroll_frame, bg="#FFFFFF", pady=4)
        row_frame.pack(fill="x", expand=True)
        
        idx = len(self.container_rows) + 1
        
        # Row variables
        c_no_var = tk.StringVar(value=c_no)
        s_no_var = tk.StringVar(value=s_no)
        pkgs_var = tk.StringVar(value=pkgs)
        
        widgets = []
        
        # Sl.No
        lbl_idx = tk.Label(row_frame, text=str(idx), font=("Segoe UI", 10), bg="#FFFFFF", fg="#64748B", width=6, anchor="center")
        lbl_idx.pack(side="left", padx=5)
        widgets.append(lbl_idx)
        
        # Container No
        ent_c_no = ttk.Entry(row_frame, textvariable=c_no_var, font=("Segoe UI", 10), width=20)
        ent_c_no.pack(side="left", padx=5)
        widgets.append(ent_c_no)
        
        # Seal No
        ent_s_no = ttk.Entry(row_frame, textvariable=s_no_var, font=("Segoe UI", 10), width=20)
        ent_s_no.pack(side="left", padx=5)
        widgets.append(ent_s_no)
        
        # Pkgs
        ent_pkgs = ttk.Entry(row_frame, textvariable=pkgs_var, font=("Segoe UI", 10), width=20)
        ent_pkgs.pack(side="left", padx=5)
        widgets.append(ent_pkgs)
        
        # Net
        ent_net = ttk.Entry(row_frame, font=("Segoe UI", 10), width=12, justify="right")
        ent_net.insert(0, f"{net_wt:.2f}")
        ent_net.pack(side="left", padx=5)
        widgets.append(ent_net)
        
        # Gross
        ent_gross = ttk.Entry(row_frame, font=("Segoe UI", 10), width=12, justify="right")
        ent_gross.insert(0, f"{gross_wt:.2f}")
        ent_gross.pack(side="left", padx=5)
        widgets.append(ent_gross)
        
        # CBM
        ent_cbm = ttk.Entry(row_frame, font=("Segoe UI", 10), width=10, justify="right")
        ent_cbm.insert(0, f"{cbm:.3f}")
        ent_cbm.pack(side="left", padx=5)
        widgets.append(ent_cbm)
        
        # Remove button
        btn_del = ttk.Button(row_frame, text="Remove", style="Danger.TButton", command=lambda: self.remove_container_row(row_frame))
        btn_del.pack(side="left", padx=5)
        widgets.append(btn_del)
        
        row_data = {
            "frame": row_frame,
            "widgets": widgets,
            "c_no_var": c_no_var,
            "s_no_var": s_no_var,
            "pkgs_var": pkgs_var,
            "net_ent": ent_net,
            "gross_ent": ent_gross,
            "cbm_ent": ent_cbm,
            "lbl_idx": lbl_idx
        }
        self.container_rows.append(row_data)
        self.recalculate_totals()
        
    def remove_container_row(self, frame):
        for idx, row in enumerate(self.container_rows):
            if row["frame"] == frame:
                for w in row["widgets"]:
                    w.destroy()
                row["frame"].destroy()
                self.container_rows.pop(idx)
                break
                
        for idx, row in enumerate(self.container_rows):
            row["lbl_idx"].config(text=str(idx+1))
            
        self.recalculate_totals()
        
    def recalculate_totals(self):
        # 1. Recalculate Cargo Items Totals
        total_c_qty = 0
        total_c_amt = 0.0
        total_c_net = 0.0
        total_c_gross = 0.0
        
        for row in self.cargo_rows:
            try:
                qty_val = int(row["qty_ent"].get().strip())
            except ValueError:
                qty_val = 0
            try:
                rate_val = float(row["rate_ent"].get().strip())
            except ValueError:
                rate_val = 0.0
            try:
                net_val = float(row["net_ent"].get().strip())
            except ValueError:
                net_val = 0.0
            try:
                gross_val = float(row["gross_ent"].get().strip())
            except ValueError:
                gross_val = 0.0
                
            amt = round(qty_val * rate_val, 2)
            row["lbl_total"].config(text=f"{amt:.2f}")
            
            total_c_qty += qty_val
            total_c_amt += amt
            total_c_net += net_val
            total_c_gross += gross_val
            
        self.total_cargo_qty.set(total_c_qty)
        self.total_cargo_amount.set(total_c_amt)
        self.total_cargo_net.set(total_c_net)
        self.total_cargo_gross.set(total_c_gross)
        
        self.lbl_cargo_totals.config(
            text=f"Total Qty: {total_c_qty} | Total Amount: ${total_c_amt:,.2f} | Total Net: {total_c_net:,.2f} kg | Total Gross: {total_c_gross:,.2f} kg"
        )
        
        # 2. Recalculate Containers Totals
        total_cnt_pkgs = 0
        total_cnt_net = 0.0
        total_cnt_gross = 0.0
        total_cnt_cbm = 0.0
        
        for row in self.container_rows:
            pkgs_str = row["pkgs_var"].get().strip()
            # Extract number if possible
            try:
                p_val = int(pkgs_str.split()[0])
            except Exception:
                p_val = 0
                
            try:
                net_val = float(row["net_ent"].get().strip())
            except ValueError:
                net_val = 0.0
            try:
                gross_val = float(row["gross_ent"].get().strip())
            except ValueError:
                gross_val = 0.0
            try:
                cbm_val = float(row["cbm_ent"].get().strip())
            except ValueError:
                cbm_val = 0.0
                
            total_cnt_pkgs += p_val
            total_cnt_net += net_val
            total_cnt_gross += gross_val
            total_cnt_cbm += cbm_val
            
        self.total_container_pkgs.set(f"{total_cnt_pkgs} BAGS")
        self.total_container_net.set(total_cnt_net)
        self.total_container_gross.set(total_cnt_gross)
        self.total_container_cbm.set(total_cnt_cbm)
        
        self.lbl_container_totals.config(
            text=f"Total Containers: {len(self.container_rows)} | Total Pkgs: {total_cnt_pkgs} BAGS | Total Net: {total_cnt_net:,.2f} kg | Total Gross: {total_cnt_gross:,.2f} kg | Total CBM: {total_cnt_cbm:.3f}"
        )
        
    def build_export_tab(self):
        card = ttk.Frame(self.tab_export, style="Card.TFrame")
        card.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Selection on left, logging/actions on right
        panel_left = tk.Frame(card, bg="#FFFFFF")
        panel_left.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        
        lbl_select = ttk.Label(panel_left, text="Select Documents to Generate", style="CardTitle.TLabel")
        lbl_select.pack(anchor="w", pady=(0, 15))
        
        chk_style = {"background": "#FFFFFF", "font": ("Segoe UI", 11), "anchor": "w"}
        cb_inv = tk.Checkbutton(panel_left, text="1. Invoice Spreadsheet (INVOICE TEMPLATE)", variable=self.gen_invoice, **chk_style)
        cb_inv.pack(fill="x", pady=5)
        
        cb_pack = tk.Checkbutton(panel_left, text="2. Packing List Spreadsheet (Packinglist)", variable=self.gen_packing, **chk_style)
        cb_pack.pack(fill="x", pady=5)
        
        cb_bl = tk.Checkbutton(panel_left, text="3. Bill of Lading Instructions (BL INSTRUCTION)", variable=self.gen_bl, **chk_style)
        cb_bl.pack(fill="x", pady=5)
        
        cb_awb = tk.Checkbutton(panel_left, text="4. Air Waybill Instructions (AWB Instructions)", variable=self.gen_awb, **chk_style)
        cb_awb.pack(fill="x", pady=5)
        
        lbl_dest = tk.Label(panel_left, text="Destination Folder Path:", font=("Segoe UI", 10, "bold"), fg="#1E293B", bg="#FFFFFF")
        lbl_dest.pack(anchor="w", pady=(20, 5))
        
        path_frame = tk.Frame(panel_left, bg="#FFFFFF")
        path_frame.pack(fill="x", pady=5)
        
        ent_path = ttk.Entry(path_frame, textvariable=self.export_path_var, font=("Segoe UI", 10))
        ent_path.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        btn_browse = ttk.Button(path_frame, text="Browse...", style="Secondary.TButton", command=self.browse_export_path)
        btn_browse.pack(side="right")
        
        btn_export = ttk.Button(panel_left, text="Generate Selected Sheets", style="Primary.TButton", command=self.export_documents)
        btn_export.pack(fill="x", pady=30)
        
        # Log Panel on Right
        panel_right = tk.Frame(card, bg="#FFFFFF")
        panel_right.pack(side="right", fill="both", expand=True, padx=20, pady=20)
        
        lbl_log = ttk.Label(panel_right, text="Generation Progress Log", style="CardTitle.TLabel")
        lbl_log.pack(anchor="w", pady=(0, 10))
        
        self.log_txt = tk.Text(panel_right, font=("Consolas", 9), height=25, width=45, bg="#F8FAFC", relief="solid", bd=1)
        self.log_txt.pack(fill="both", expand=True)
        
    def browse_export_path(self):
        folder = filedialog.askdirectory(title="Select Destination Folder")
        if folder:
            self.export_path_var.set(folder)
            
    def load_defaults(self):
        # Default text inputs
        defaults = {
            "shipper": "CREATIVE PACKAGING SYSTEMS\nNo.195-A, SERVICE ROAD , SOMJI BHAI NAGAR\nKALLIKUPPAM, CHENNAI - 600 053.\nPh: 044-6133 4666 MAIL: bm@creativepackaging.in",
            "consignee": "FOSHAN BOGAL PACKING MACHINERY CO.,LTD\nNO.65,Road 2,Rongxing Industrial Zone ,\nWuzhuang,Luocun Town,Naihai  District,Guangdong\npostcode:528000\nEMAIL ID:amy@sierac.cn\nTAX ID :91440605MA4W0ERU95",
            "notify": "GLOBAL OCEAN TRADING LIMITED,\nBLK 808 FRENCH ROAD,\n06-151 KITCHENER COMPLEX\nSINGAPORE 200208",
            "also_notify": "SAME AS CONSIGNEE",
            "cha": "CREATIVE CHA SERVICES",
            "forwarding_agent": "LEADKING AIR SERVICES PVT. LTD",
            "vessel_voyage": "HUA DA 617 V. 006W",
            "port_of_loading": "CHENNAI / INDIA",
            "port_of_discharge": "SHANGHAI",
            "final_destination": "SHANGHAI / CHINA",
            "country_of_origin": "INDIA",
            "country_of_dest": "CHINA",
            "terms_delivery_payment": "SAMPLE FOR TESTING THE PACKAGING MACHINERY\nNO COMMERCIAL VALUE & NOT FOR SALE\nNO FOREIGN EXCHANGE IS INVOLVED",
            "invoice_no_date": "CPS/001 DT. 08.04.2026",
            "buyer_order_no_date": "Order dt. 05.04.2026",
            "pre_carriage": "NA",
            "place_of_receipt": "NA",
            "sb_no_date": "2421435 DT. 16/04/2026",
            "hs_code": "10051000",
            "payment_mode": "PREPAID",
            "dimensions": "28 X 25.5 X 25.5 CMS",
            "gross_weight": "4.66 Kgs",
            "net_weight": "4.00 Kgs",
            "special_instructions": "DELIVER TO SHANGHAI SHED 3",
            "rbi_code": "RBI-12345",
            "ie_code": "IE-54321",
            "account": "A/C NO. 9876543210 - STATE BANK OF INDIA",
            "gr_no_date": "GR-777 DT. 08.04.2026",
            "other_ref": "REF-001",
            "buyer_if_other": "",
            "marks_nos": "1 CARTON BOX\nContainer no.",
            "contents": "PACKAGING MACHINERY SPARES",
            "bank_ac_drawback": "SBI BANK A/C: 9876543210, IFSC: SBIN0000001",
            "type_of_packing": "PP BAGS",
            "total_pkgs": "1600 PP BAGS",
            "cargo_desc": "HYBRID MAIZE SEEDS (C.P.811)",
            "invoice_value": "USD 120.00"
        }
        
        for k, v in defaults.items():
            if k in self.inputs:
                self.inputs[k].set(v)
            if k in self.address_textboxes:
                self.address_textboxes[k].delete("1.0", tk.END)
                self.address_textboxes[k].insert("1.0", v)
            if k in self.multiline_entries:
                self.multiline_entries[k].delete("1.0", tk.END)
                self.multiline_entries[k].insert("1.0", v)
                
        # Default cargo items
        self.add_cargo_row("TOOTH BRUSH ( PLASTIC )", 10, 1.50, 0.1, 0.12)
        self.add_cargo_row("COMB ( PLASTIC )", 10, 1.00, 0.08, 0.1)
        self.add_cargo_row("TOOTH PASTE", 10, 2.00, 0.15, 0.18)
        self.add_cargo_row("SHAVING RAZOR", 20, 3.00, 0.05, 0.07)
        self.add_cargo_row("PLASTIC ROLL", 1, 12.00, 1.2, 1.3)
        self.add_cargo_row("PAPER ROLL", 1, 15.00, 1.5, 1.6)
        
        # Default containers
        self.add_container_row("REGU3290956", "INMAA2605783", "400 BAGS", 20000.0, 20200.0, 20.0)
        self.add_container_row("CAIU3743948", "INMAA2605787", "400 BAGS", 20000.0, 20200.0, 20.0)
        self.add_container_row("TGBU3715873", "INMAA2605738", "400 BAGS", 20000.0, 20200.0, 20.0)
        self.add_container_row("REGU3265634", "INMAA2605754", "400 BAGS", 20000.0, 20200.0, 20.0)
        
    def log(self, msg):
        self.log_txt.insert(tk.END, msg + "\n")
        self.log_txt.see(tk.END)
        self.update_idletasks()
        
    def get_data_payload(self):
        payload = {}
        # Fetch from inputs
        for k, v in self.inputs.items():
            payload[k] = v.get().strip()
        # Fetch from addresses
        for k, txt in self.address_textboxes.items():
            payload[k] = txt.get("1.0", tk.END).strip()
        # Fetch from textareas
        for k, txt in self.multiline_entries.items():
            payload[k] = txt.get("1.0", tk.END).strip()
            
        return payload
        
    def export_documents(self):
        self.recalculate_totals()
        dest_dir = self.export_path_var.get().strip()
        
        if not dest_dir:
            messagebox.showerror("Error", "Please specify a destination folder.")
            return
            
        os.makedirs(dest_dir, exist_ok=True)
        
        self.log_txt.delete("1.0", tk.END)
        self.log("--- Starting Excel Document Generation ---")
        self.log(f"Destination: {dest_dir}\n")
        
        data = self.get_data_payload()
        
        # Get items payload
        items = []
        for r in self.cargo_rows:
            desc = r["desc_var"].get().strip()
            try:
                qty = int(r["qty_ent"].get().strip())
            except ValueError:
                qty = 0
            try:
                rate = float(r["rate_ent"].get().strip())
            except ValueError:
                rate = 0.0
            try:
                net = float(r["net_ent"].get().strip())
            except ValueError:
                net = 0.0
            try:
                gross = float(r["gross_ent"].get().strip())
            except ValueError:
                gross = 0.0
                
            if desc:
                items.append({"desc": desc, "qty": qty, "rate": rate, "net_wt": net, "gross_wt": gross})
                
        # Get containers payload
        containers = []
        for r in self.container_rows:
            c_no = r["c_no_var"].get().strip()
            s_no = r["s_no_var"].get().strip()
            pkgs = r["pkgs_var"].get().strip()
            net = r["net_ent"].get().strip()
            gross = r["gross_ent"].get().strip()
            cbm = r["cbm_ent"].get().strip()
            
            if c_no:
                containers.append({"container_no": c_no, "seal_no": s_no, "pkgs": pkgs, "net_wt": net, "gross_wt": gross, "cbm": cbm})
                
        # Define paths to internal templates
        data["template_invoice_path"] = resource_path("INVOICE TEMPLATE_template.xlsx")
        data["template_packing_path"] = resource_path("Packinglist_template.xlsx")
        data["template_bl_path"] = resource_path("BL INSTRUCTION CPS 001_template.xlsx")
        data["template_awb_path"] = resource_path("AWB Instructions_template.xlsx")
        
        # Verify internal templates exist
        for k in ["template_invoice_path", "template_packing_path", "template_bl_path", "template_awb_path"]:
            p = data[k]
            if not os.path.exists(p):
                # Fallback to local execution directory template files
                local_name = os.path.basename(p)
                local_p = os.path.abspath(local_name)
                if os.path.exists(local_p):
                    data[k] = local_p
                else:
                    self.log(f"CRITICAL ERROR: Template file {local_name} not found in bundle or current directory!")
                    messagebox.showerror("Error", f"Template file {local_name} not found!")
                    return
                    
        success_count = 0
        
        # 1. Export Invoice
        if self.gen_invoice.get():
            try:
                out_path = os.path.join(dest_dir, "Generated_Invoice.xlsx")
                self.log(f"Generating Invoice: {os.path.basename(out_path)}...")
                self.generate_invoice_file(data, items, out_path)
                self.log("  Successfully generated Invoice!\n")
                success_count += 1
            except Exception as e:
                self.log(f"  Error generating Invoice: {e}\n")
                
        # 2. Export Packing List
        if self.gen_packing.get():
            try:
                out_path = os.path.join(dest_dir, "Generated_Packinglist.xlsx")
                self.log(f"Generating Packing List: {os.path.basename(out_path)}...")
                self.generate_packing_list_file(data, items, out_path)
                self.log("  Successfully generated Packing List!\n")
                success_count += 1
            except Exception as e:
                self.log(f"  Error generating Packing List: {e}\n")
                
        # 3. Export BL Instructions
        if self.gen_bl.get():
            try:
                out_path = os.path.join(dest_dir, "Generated_BL_Instructions.xlsx")
                self.log(f"Generating BL Instructions: {os.path.basename(out_path)}...")
                self.generate_bl_instructions_file(data, containers, out_path)
                self.log("  Successfully generated BL Instructions!\n")
                success_count += 1
            except Exception as e:
                self.log(f"  Error generating BL Instructions: {e}\n")
                
        # 4. Export AWB Instructions
        if self.gen_awb.get():
            try:
                out_path = os.path.join(dest_dir, "Generated_AWB_Instructions.xlsx")
                self.log(f"Generating Air Waybill Instructions: {os.path.basename(out_path)}...")
                self.generate_awb_instructions_file(data, out_path)
                self.log("  Successfully generated AWB Instructions!\n")
                success_count += 1
            except Exception as e:
                self.log(f"  Error generating AWB Instructions: {e}\n")
                
        self.log(f"--- Export Finished: {success_count} documents generated successfully ---")
        if success_count > 0:
            messagebox.showinfo("Success", f"{success_count} documents successfully generated in:\n{dest_dir}")
            
    def generate_invoice_file(self, data, items, out_path):
        wb = openpyxl.load_workbook(data["template_invoice_path"])
        ws = wb["Invoice"]
        ws.views.sheetView[0].showGridLines = True
        
        # Shipper
        shipper_lines = data["shipper"].split("\n")
        for idx, line in enumerate(shipper_lines[:4]):
            ws.cell(row=3 + idx, column=2, value=line)
            
        # Invoice No & Date
        ws.cell(row=3, column=7, value=data["invoice_no_date"])
        
        # Buyer's Order No & Date
        ws.cell(row=6, column=7, value=data["buyer_order_no_date"])
        
        # Consignee
        consignee_lines = data["consignee"].split("\n")
        for idx, line in enumerate(consignee_lines[:6]):
            ws.cell(row=8 + idx, column=2, value=line)
            
        # Other Reference(s)
        ws.cell(row=8, column=8, value=data["other_ref"])
        
        # Buyer if other than consignee
        buyer_lines = data["buyer_if_other"].split("\n")
        for idx, line in enumerate(buyer_lines[:3]):
            ws.cell(row=10 + idx, column=8, value=line)
            
        # Shipping Details
        ws.cell(row=15, column=2, value=data["pre_carriage"])
        ws.cell(row=15, column=4, value=data["place_of_receipt"])
        ws.cell(row=15, column=7, value=data["country_of_origin"])
        ws.cell(row=15, column=10, value=data["country_of_dest"])
        
        ws.cell(row=17, column=2, value=data["vessel_voyage"])
        ws.cell(row=17, column=4, value=data["port_of_loading"])
        
        terms_lines = data["terms_delivery_payment"].split("\n")
        for idx, line in enumerate(terms_lines[:3]):
            ws.cell(row=17 + idx, column=7, value=line)
            
        ws.cell(row=19, column=2, value=data["port_of_discharge"])
        ws.cell(row=19, column=4, value=data["final_destination"])
        
        # Marks & Nos
        marks_lines = data["marks_nos"].split("\n")
        for idx, line in enumerate(marks_lines[:2]):
            ws.cell(row=22 + idx, column=2, value=line)
            
        # Fill Cargo Items
        first_item_row = 24
        template_rows = 6
        
        if len(items) > template_rows:
            rows_to_insert = len(items) - template_rows
            for _ in range(rows_to_insert):
                insert_row_copy_style(ws, 30, 29)
                
        total_qty = 0
        total_amt = 0.0
        for idx, item in enumerate(items):
            r = first_item_row + idx
            ws.cell(row=r, column=3, value=item["desc"])
            ws.cell(row=r, column=9, value=item["qty"])
            ws.cell(row=r, column=10, value=item["rate"])
            
            amt = round(item["qty"] * item["rate"], 2)
            ws.cell(row=r, column=11, value=amt)
            ws.cell(row=r, column=11).number_format = "$#,##0.00"
            
            total_qty += item["qty"]
            total_amt += amt
            
        # Clear leftovers
        if len(items) < template_rows:
            for r in range(first_item_row + len(items), first_item_row + template_rows):
                ws.cell(row=r, column=3, value=None)
                ws.cell(row=r, column=9, value=None)
                ws.cell(row=r, column=10, value=None)
                ws.cell(row=r, column=11, value=None)
                
        # Find totals & labels
        total_row = None
        dim_row = None
        gw_row = None
        words_row = None
        
        for r in range(first_item_row, ws.max_row + 1):
            cell_h = ws.cell(row=r, column=8).value
            cell_b = ws.cell(row=r, column=2).value
            
            if cell_h == "TOTAL":
                total_row = r
            if cell_b == "DIMENSIONS IN CMS : ":
                dim_row = r
            if cell_b == "GROSS WEIGHT:":
                gw_row = r
            if cell_b == "(Amount in words)":
                words_row = r + 1
                
        if total_row:
            ws.cell(row=total_row, column=9, value=total_qty)
            ws.cell(row=total_row, column=11, value=total_amt)
            ws.cell(row=total_row, column=11).number_format = "$#,##0.00"
        if dim_row:
            ws.cell(row=dim_row, column=5, value=data["dimensions"])
        if gw_row:
            ws.cell(row=gw_row, column=5, value=data["gross_weight"])
        if words_row:
            ws.cell(row=words_row, column=2, value=num_to_words(total_amt))
            
        wb.save(out_path)
        
    def generate_packing_list_file(self, data, items, out_path):
        wb = openpyxl.load_workbook(data["template_packing_path"])
        ws = wb["Invoice"]
        ws.views.sheetView[0].showGridLines = True
        
        # Shipper
        shipper_lines = data["shipper"].split("\n")
        for idx, line in enumerate(shipper_lines[:4]):
            ws.cell(row=3 + idx, column=2, value=line)
            
        # Invoice No & Date
        ws.cell(row=3, column=7, value=data["invoice_no_date"])
        
        # Buyer's Order No & Date
        ws.cell(row=6, column=7, value=data["buyer_order_no_date"])
        
        # Consignee
        consignee_lines = data["consignee"].split("\n")
        for idx, line in enumerate(consignee_lines[:6]):
            ws.cell(row=8 + idx, column=2, value=line)
            
        # Other Reference(s)
        ws.cell(row=8, column=8, value=data["other_ref"])
        
        # Buyer if other than consignee
        buyer_lines = data["buyer_if_other"].split("\n")
        for idx, line in enumerate(buyer_lines[:3]):
            ws.cell(row=10 + idx, column=8, value=line)
            
        # Shipping details
        ws.cell(row=14, column=2, value=data["pre_carriage"])
        ws.cell(row=14, column=4, value=data["place_of_receipt"])
        ws.cell(row=14, column=7, value=data["country_of_origin"])
        ws.cell(row=14, column=10, value=data["country_of_dest"])
        
        ws.cell(row=16, column=2, value=data["vessel_voyage"])
        ws.cell(row=16, column=4, value=data["port_of_loading"])
        
        terms_lines = data["terms_delivery_payment"].split("\n")
        for idx, line in enumerate(terms_lines[:3]):
            ws.cell(row=16 + idx, column=7, value=line)
            
        ws.cell(row=18, column=2, value=data["port_of_discharge"])
        ws.cell(row=18, column=4, value=data["final_destination"])
        
        # Marks & Nos
        marks_lines = data["marks_nos"].split("\n")
        for idx, line in enumerate(marks_lines[:2]):
            ws.cell(row=21 + idx, column=2, value=line)
            
        # Cargo Items
        first_item_row = 23
        template_rows = 16
        
        if len(items) > template_rows:
            rows_to_insert = len(items) - template_rows
            for _ in range(rows_to_insert):
                insert_row_copy_style(ws, 39, 38)
                
        total_qty = 0
        total_net = 0.0
        total_gross = 0.0
        for idx, item in enumerate(items):
            r = first_item_row + idx
            ws.cell(row=r, column=4, value=item["desc"])
            ws.cell(row=r, column=9, value=item["qty"])
            ws.cell(row=r, column=10, value=item["net_wt"])
            ws.cell(row=r, column=11, value=item["gross_wt"])
            
            total_qty += item["qty"]
            total_net += item["net_wt"]
            total_gross += item["gross_wt"]
            
        # Clear leftover
        if len(items) < template_rows:
            for r in range(first_item_row + len(items), first_item_row + template_rows):
                ws.cell(row=r, column=4, value=None)
                ws.cell(row=r, column=9, value=None)
                ws.cell(row=r, column=10, value=None)
                ws.cell(row=r, column=11, value=None)
                
        # Find totals & labels
        total_row = None
        dim_row = None
        
        for r in range(first_item_row, ws.max_row + 1):
            cell_h = ws.cell(row=r, column=8).value
            cell_b = ws.cell(row=r, column=2).value
            
            if cell_h == "TOTAL":
                total_row = r
            if cell_b == "Dimensions :":
                dim_row = r
                
        if total_row:
            ws.cell(row=total_row, column=9, value=total_qty)
            ws.cell(row=total_row, column=10, value=total_net)
            ws.cell(row=total_row, column=11, value=total_gross)
        if dim_row:
            ws.cell(row=dim_row, column=5, value=data["dimensions"])
            
        wb.save(out_path)
        
    def generate_bl_instructions_file(self, data, containers, out_path):
        wb = openpyxl.load_workbook(data["template_bl_path"])
        ws = wb["Table 1"]
        ws.views.sheetView[0].showGridLines = True
        
        # Shipper
        ws.cell(row=2, column=1, value="SHIPPER:\n" + data["shipper"])
        # CHA
        ws.cell(row=2, column=4, value="CHA\n" + data["cha"])
        # Consignee
        ws.cell(row=3, column=1, value="CONSIGNEE: (Complete Address and fax#)\n" + data["consignee"])
        # Forwarding Agent
        ws.cell(row=3, column=4, value="FORWARDING AGENT ADDRESS:\n" + data["forwarding_agent"])
        # Notify
        ws.cell(row=4, column=1, value="NOTIFY:       (Complete Address and Fax#)\n" + data["notify"])
        # Also Notify
        ws.cell(row=4, column=4, value="ALSO NOTIFY (Complete Address and Fax#)\n" + data["also_notify"])
        
        # Voyage details
        ws.cell(row=5, column=1, value="VESSEL AND VOYAGE NO :  " + data["vessel_voyage"])
        ws.cell(row=5, column=4, value="PORT OF LOADING: " + data["port_of_loading"])
        ws.cell(row=6, column=1, value="PORT OF DISCHARGE: " + data["port_of_discharge"])
        ws.cell(row=6, column=4, value="FINAL DESTINATION: " + data["final_destination"])
        
        # Marks & Nos, Packages
        ws.cell(row=8, column=1, value=data["marks_nos"])
        ws.cell(row=8, column=2, value=data["type_of_packing"])
        
        # Container totals & texts
        total_cbm = sum(float(c["cbm"]) for c in containers if c["cbm"])
        
        # Total pkgs parsing
        try:
            pkgs_sum = sum(int(c["pkgs"].split()[0]) for c in containers if c["pkgs"] and c["pkgs"].split()[0].isdigit())
            pkgs_text = f"{pkgs_sum} PP BAGS"
        except Exception:
            pkgs_text = data["total_pkgs"]
            
        ws.cell(row=8, column=3, value=f"{len(containers)}X20' FCL CONTAINER SAID TO CONTAIN\nTOTAL: {pkgs_text}")
        ws.cell(row=9, column=3, value=data["cargo_desc"])
        
        ws.cell(row=11, column=3, value="INV NOS :")
        ws.cell(row=12, column=3, value="  " + data["invoice_no_date"])
        ws.cell(row=14, column=3, value="HS CODE : " + data["hs_code"])
        ws.cell(row=16, column=3, value="SB NO : " + data["sb_no_date"])
        
        # Weights
        ws.cell(row=8, column=5, value=data["gross_weight"])
        ws.cell(row=10, column=5, value="NET. WT.")
        ws.cell(row=11, column=5, value=data["net_weight"])
        ws.cell(row=8, column=6, value=f"{total_cbm:.3f} CBM" if total_cbm > 0 else "")
        
        # Container table
        first_item_row = 20
        template_rows = 4
        
        if len(containers) > template_rows:
            rows_to_insert = len(containers) - template_rows
            for _ in range(rows_to_insert):
                insert_row_copy_style(ws, 24, 23)
                
        for idx, c in enumerate(containers):
            r = first_item_row + idx
            ws.cell(row=r, column=1, value=c["container_no"])
            ws.cell(row=r, column=2, value=c["seal_no"])
            ws.cell(row=r, column=3, value=c["pkgs"])
            
            try:
                net_val = float(c["net_wt"])
                ws.cell(row=r, column=4, value=net_val)
                ws.cell(row=r, column=4).number_format = "#,##0"
            except ValueError:
                ws.cell(row=r, column=4, value=c["net_wt"])
                
            try:
                gross_val = float(c["gross_wt"])
                ws.cell(row=r, column=5, value=gross_val)
                ws.cell(row=r, column=5).number_format = "#,##0"
            except ValueError:
                ws.cell(row=r, column=5, value=c["gross_wt"])
                
            try:
                cbm_val = float(c["cbm"])
                ws.cell(row=r, column=6, value=cbm_val)
                ws.cell(row=r, column=6).number_format = "0.00"
            except ValueError:
                ws.cell(row=r, column=6, value=c["cbm"])
                
        # Clear leftover rows
        if len(containers) < template_rows:
            for r in range(first_item_row + len(containers), first_item_row + template_rows):
                ws.cell(row=r, column=1, value=None)
                ws.cell(row=r, column=2, value=None)
                ws.cell(row=r, column=3, value=None)
                ws.cell(row=r, column=4, value=None)
                ws.cell(row=r, column=5, value=None)
                ws.cell(row=r, column=6, value=None)
                
        wb.save(out_path)
        
    def generate_awb_instructions_file(self, data, out_path):
        wb = openpyxl.load_workbook(data["template_awb_path"])
        ws = wb["Sheet1 (2)"]
        ws.views.sheetView[0].showGridLines = True
        
        # Destination
        ws.cell(row=6, column=3, value=data["final_destination"])
        # Collect / Prepaid
        ws.cell(row=6, column=8, value=data["payment_mode"])
        
        # Shipper
        shipper_lines = data["shipper"].split("\n")
        for idx, line in enumerate(shipper_lines[:4]):
            ws.cell(row=9 + idx, column=1, value=line)
            
        # Mode of payment
        ws.cell(row=8, column=8, value=data["payment_mode"])
        
        # AWB / BL No (fallback to invoice)
        ws.cell(row=9, column=8, value=data["invoice_no_date"])
        
        # Packages
        ws.cell(row=11, column=7, value=data["total_pkgs"])
        ws.cell(row=11, column=9, value=data["type_of_packing"])
        
        # Consignee
        consignee_lines = data["consignee"].split("\n")
        for idx, line in enumerate(consignee_lines[:4]):
            ws.cell(row=14 + idx, column=1, value=line)
            
        # RBI / IE Codes, Gross Weight
        ws.cell(row=13, column=8, value=data["rbi_code"])
        ws.cell(row=14, column=8, value=data["ie_code"])
        ws.cell(row=15, column=8, value=data["gross_weight"])
        
        # Notify
        notify_lines = data["notify"].split("\n")
        for idx, line in enumerate(notify_lines[:4]):
            ws.cell(row=19 + idx, column=1, value=line)
            
        # Dimensions
        dim_lines = data["dimensions"].split("\n")
        for idx, line in enumerate(dim_lines[:2]):
            ws.cell(row=19 + idx, column=6, value=line)
            
        # G.R No
        ws.cell(row=22, column=6, value=data["gr_no_date"])
        
        # Account
        account_lines = data["account"].split("\n")
        for idx, line in enumerate(account_lines[:4]):
            ws.cell(row=24 + idx, column=1, value=line)
            
        # Marks & Nos
        marks_lines = data["marks_nos"].split("\n")
        for idx, line in enumerate(marks_lines[:3]):
            ws.cell(row=24 + idx, column=6, value=line)
            
        # Goods Description
        goods_lines = data["cargo_desc"].split("\n")
        for idx, line in enumerate(goods_lines[:5]):
            ws.cell(row=28 + idx, column=6, value=line)
            
        # Invoice No
        ws.cell(row=29, column=1, value=data["invoice_no_date"])
        
        # Contents
        ws.cell(row=32, column=1, value=data["contents"])
        
        # Bank Account
        bank_lines = data["bank_ac_drawback"].split("\n")
        for idx, line in enumerate(bank_lines[:2]):
            ws.cell(row=34 + idx, column=1, value=line)
            
        # Special Instructions
        special_lines = data["special_instructions"].split("\n")
        for idx, line in enumerate(special_lines[:2]):
            ws.cell(row=34 + idx, column=6, value=line)
            
        # Specific layout adjustments for AWB Instructions print safety
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['H'].width = 17
        wb.save(out_path)

if __name__ == "__main__":
    app = LogisticsApp()
    app.mainloop()
