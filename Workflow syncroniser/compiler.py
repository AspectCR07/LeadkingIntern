import os
import sys
import json
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import pdf_parser

# Helpers copied from unified_generator.py
def num_to_words(num):
    units = ["", "ONE", "TWO", "THREE", "FOUR", "FIVE", "SIX", "SEVEN", "EIGHT", "NINE", "TEN", 
             "ELEVEN", "TWELVE", "THIRTEEN", "FOURTEEN", "FIFTEEN", "SIXTEEN", "SEVENTEEN", "EIGHTEEN", "NINETEEN"]
    tens = ["", "", "TWENTY", "THIRTY", "FORTY", "FIFTY", "SIXTY", "SEVENTY", "EIGHTY", "NINETY"]
    
    def helper(n):
        if n < 20:
            return units[n]
        elif n < 100:
            return tens[n // 10] + (" " + units[n % 10] if n % 10 != 0 else "")
        elif n < 1000:
            return units[n // 100] + " HUNDRED" + (" AND " + helper(n % 100) if n % 100 != 0 else "")
        elif n < 100000:
            return helper(n // 1000) + " THOUSAND" + (" " + helper(n % 1000) if n % 1000 != 0 else "")
        else:
            return str(n)
            
    if num == 0:
        return "SAY USD ZERO ONLY"
        
    parts = f"{num:.2f}".split(".")
    dollars = int(parts[0])
    cents = int(parts[1])
    
    words = "SAY USD " + helper(dollars)
    if cents > 0:
        words += " AND CENTS " + helper(cents) + " ONLY"
    else:
        words += " ONLY"
    return words

def write_address_block(ws, start_row, col_idx, text, max_rows=10):
    lines = [line.strip() for line in str(text).split("\n") if line.strip()]
    for i in range(max_rows):
        r = start_row + i
        if i < len(lines):
            ws.cell(row=r, column=col_idx).value = lines[i]
        else:
            ws.cell(row=r, column=col_idx).value = None

def write_wrapped_text_to_rows(ws, start_row, col_idx, text, max_rows, limit=65):
    raw_lines = [line.strip() for line in str(text).split("\n") if line.strip()]
    wrapped_lines = []
    for line in raw_lines:
        if len(line) <= limit:
            wrapped_lines.append(line)
        else:
            words = line.split()
            current = []
            for w in words:
                if len(" ".join(current + [w])) <= limit:
                    current.append(w)
                else:
                    wrapped_lines.append(" ".join(current))
                    current = [w]
            if current:
                wrapped_lines.append(" ".join(current))
    for i in range(max_rows):
        r = start_row + i
        if i < len(wrapped_lines):
            ws.cell(row=r, column=col_idx).value = wrapped_lines[i]
        else:
            ws.cell(row=r, column=col_idx).value = None

def clean_weight(val):
    if not val:
        return None
    cleaned = re.sub(r'(?i)kgs|kg|packages|packages|pkg|pkgs', '', str(val)).replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return val

def insert_row_copy_style(ws, row_idx, copy_from_row):
    ws.insert_rows(row_idx)
    for col_idx in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=copy_from_row, column=col_idx)
        dest_cell = ws.cell(row=row_idx, column=col_idx)
        if src_cell.has_style:
            dest_cell.font = Font(
                name=src_cell.font.name, size=src_cell.font.size,
                bold=src_cell.font.bold, italic=src_cell.font.italic,
                color=src_cell.font.color, underline=src_cell.font.underline
            )
            dest_cell.border = Border(
                left=src_cell.border.left, right=src_cell.border.right,
                top=src_cell.border.top, bottom=src_cell.border.bottom
            )
            dest_cell.fill = openpyxl.styles.PatternFill(
                fill_type=src_cell.fill.fill_type,
                start_color=src_cell.fill.start_color,
                end_color=src_cell.fill.end_color
            )
            dest_cell.alignment = Alignment(
                horizontal=src_cell.alignment.horizontal,
                vertical=src_cell.alignment.vertical,
                wrap_text=src_cell.alignment.wrap_text
            )
            dest_cell.number_format = src_cell.number_format

# BL Generation functions adapted from unified_generator.py
def generate_bl_cotton_home(data, containers, template_file, out_path):
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    write_address_block(ws, 6, 1, data.get("shipper", ""), max_rows=9)
    write_address_block(ws, 16, 1, data.get("consignee", ""), max_rows=5)
    write_address_block(ws, 22, 1, data.get("notify", ""), max_rows=6)
    
    also_notify = data.get("also_notify", "")
    if also_notify and not also_notify.upper().startswith("ALSO NOTIFY"):
        also_notify = "ALSO NOTIFY:- " + also_notify
    write_address_block(ws, 21, 5, also_notify, max_rows=7)
    
    ws["A28"] = data.get("vessel_voyage", "")
    ws["C28"] = data.get("port_of_loading", "")
    ws["A31"] = data.get("port_of_discharge", "")
    ws["C31"] = data.get("final_destination", "")
    
    ws["A37"] = "CTN NOS"
    pkg_from = data.get("package_from", 1)
    pkg_to = data.get("package_to", 1)
    ws["A38"] = f"{pkg_from} TO {pkg_to}"
    
    total_pkgs = data.get("total_pkgs", "")
    try:
        val = int(total_pkgs.split()[0].replace(",", ""))
        ws["B35"] = val
    except:
        ws["B35"] = total_pkgs
        
    ws["C35"] = f"{len(containers)} X 40'HC FCL / FCL - CONTAINER ONLY"
    try:
        val = int(total_pkgs.split()[0].replace(",", ""))
        ws["C37"] = f"[{num_to_words(val).replace('SAY USD ', '').replace(' ONLY', '')} CARTONS ONLY]"
    except:
        ws["C37"] = f"[{total_pkgs} ONLY]"
        
    cargo_desc = data.get("cargo_desc", "")
    write_wrapped_text_to_rows(ws, 39, 3, cargo_desc, max_rows=4, limit=65)
    
    ws["H35"] = clean_weight(data.get("gross_weight", ""))
    ws["H42"] = clean_weight(data.get("net_weight", ""))
    
    total_cbm = sum(float(c.get("cbm", 0)) for c in containers)
    ws["I35"] = total_cbm if total_cbm > 0 else None
    
    c_nos = [c["container_no"] for c in containers if c.get("container_no")]
    s_nos = [c["seal_no"] for c in containers if c.get("seal_no")]
    ws["A52"] = ", ".join(c_nos)
    ws["A55"] = ", ".join(s_nos)
    
    ws["C52"] = "INV NO: " + data.get("invoice_no_date", "")
    ws["C53"] = "S.B.NO: " + data.get("sb_no_date", "")
    
    hs = data.get("hs_code", "")
    hs_text = f"HS CODE : {hs}" if hs else ""
    write_wrapped_text_to_rows(ws, 47, 3, hs_text, max_rows=3, limit=65)
    
    ws["C55"] = '"FREIGHT ' + data.get("payment_mode", "COLLECT") + '"'
    
    wb.save(out_path)

def generate_bl_pemi(data, containers, template_file, out_path):
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    write_address_block(ws, 6, 1, data.get("shipper", ""), max_rows=7)
    write_address_block(ws, 14, 1, data.get("consignee", ""), max_rows=7)
    write_address_block(ws, 22, 1, data.get("notify", ""), max_rows=7)
    
    ws["A30"] = data.get("vessel_voyage", "")
    ws["C30"] = data.get("port_of_loading", "")
    ws["A33"] = data.get("port_of_discharge", "")
    ws["C33"] = data.get("final_destination", "")
    
    ws["A37"] = data.get("marks_nos", "")
    pkg_from = data.get("package_from", 1)
    pkg_to = data.get("package_to", 1)
    ws["A40"] = f"{pkg_from} TO {pkg_to}"
    
    total_pkgs = data.get("total_pkgs", "")
    try:
        val = int(total_pkgs.split()[0].replace(",", ""))
        ws["B37"] = val
    except:
        ws["B37"] = total_pkgs
        
    ws["C37"] = f"{len(containers)} X 40'HC - FCL / FCL"
    try:
        val = int(total_pkgs.split()[0].replace(",", ""))
        ws["C39"] = f"[{num_to_words(val).replace('SAY USD ', '').replace(' ONLY', '')} CARTONS ONLY]"
    except:
        ws["C39"] = f"[{total_pkgs} ONLY]"
        
    cargo_desc = data.get("cargo_desc", "")
    write_wrapped_text_to_rows(ws, 42, 3, cargo_desc, max_rows=6, limit=65)
    
    ws["E37"] = clean_weight(data.get("gross_weight", ""))
    net_wt = clean_weight(data.get("net_weight", ""))
    ws["C54"] = f"NET WEIGHT: {net_wt:.3f} KGS" if isinstance(net_wt, float) else f"NET WEIGHT: {net_wt}"
    
    total_cbm = sum(float(c.get("cbm", 0)) for c in containers)
    ws["F37"] = total_cbm if total_cbm > 0 else None
    
    c_nos = [c["container_no"] for c in containers if c.get("container_no")]
    s_nos = [c["seal_no"] for c in containers if c.get("seal_no")]
    ws["A49"] = ", ".join(c_nos)
    ws["A54"] = ", ".join(s_nos)
    
    ws["C52"] = "INV No: " + data.get("invoice_no_date", "")
    ws["C53"] = "SB NO: " + data.get("sb_no_date", "")
    ws["C56"] = '"FREIGHT ' + data.get("payment_mode", "COLLECT") + '"'
    
    wb.save(out_path)

def generate_bl_sri_sakthi(data, containers, template_file, out_path):
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    write_address_block(ws, 6, 1, data.get("shipper", ""), max_rows=9)
    write_address_block(ws, 17, 1, data.get("consignee", ""), max_rows=8)
    write_address_block(ws, 26, 1, data.get("notify", ""), max_rows=8)
    
    ws["A35"] = data.get("vessel_voyage", "")
    ws["C35"] = data.get("port_of_loading", "")
    ws["A38"] = data.get("port_of_discharge", "")
    ws["C38"] = data.get("final_destination", "")
    
    write_address_block(ws, 42, 1, data.get("marks_nos", ""), max_rows=5)
    total_pkgs = data.get("total_pkgs", "")
    try:
        val = int(total_pkgs.split()[0].replace(",", ""))
        ws["B42"] = val
    except:
        ws["B42"] = total_pkgs
        
    ws["C42"] = f"{len(containers)} X 20' GP / FCL - FCL"
    try:
        val = int(total_pkgs.split()[0].replace(",", ""))
        ws["C44"] = f"({num_to_words(val).replace('SAY USD ', '').replace(' ONLY', '')} PACKAGES ONLY)"
    except:
        ws["C44"] = f"({total_pkgs} ONLY)"
        
    cargo_desc = data.get("cargo_desc", "")
    write_wrapped_text_to_rows(ws, 48, 3, cargo_desc, max_rows=5, limit=65)
    
    hs = data.get("hs_code", "")
    hs_text = f"HS CODE : {hs}" if hs else ""
    write_wrapped_text_to_rows(ws, 46, 3, hs_text, max_rows=2, limit=65)
    
    ws["F42"] = clean_weight(data.get("gross_weight", ""))
    net_wt = clean_weight(data.get("net_weight", ""))
    ws["C57"] = f"NT.WT: {net_wt:.3f} KGS" if isinstance(net_wt, float) else f"NT.WT: {net_wt}"
    
    total_cbm = sum(float(c.get("cbm", 0)) for c in containers)
    ws["G42"] = total_cbm if total_cbm > 0 else None
    
    c_nos = [c["container_no"] for c in containers if c.get("container_no")]
    s_nos = [c["seal_no"] for c in containers if c.get("seal_no")]
    ws["A52"] = ", ".join(c_nos)
    ws["A55"] = ", ".join(s_nos)
    
    ws["C55"] = "INV NO: " + data.get("invoice_no_date", "")
    ws["C56"] = "S.B.NO: " + data.get("sb_no_date", "")
    ws["C60"] = '"FREIGHT ' + data.get("payment_mode", "PREPAID") + '"'
    
    wb.save(out_path)

def generate_bl_standard(data, containers, template_file, out_path):
    wb = openpyxl.load_workbook(template_file)
    ws = wb["Table 1"]
    ws.views.sheetView[0].showGridLines = True
    
    ws.cell(row=2, column=1, value="SHIPPER:\n" + data.get("shipper", ""))
    ws.cell(row=2, column=4, value="CHA\n" + data.get("cha", ""))
    ws.cell(row=3, column=1, value="CONSIGNEE: (Complete Address and fax#)\n" + data.get("consignee", ""))
    ws.cell(row=3, column=4, value="FORWARDING AGENT ADDRESS:\n" + data.get("forwarding_agent", ""))
    ws.cell(row=4, column=1, value="NOTIFY:       (Complete Address and Fax#)\n" + data.get("notify", ""))
    ws.cell(row=4, column=4, value="ALSO NOTIFY (Complete Address and Fax#)\n" + data.get("also_notify", ""))
    
    ws.cell(row=5, column=1, value="VESSEL AND VOYAGE NO :  " + data.get("vessel_voyage", ""))
    ws.cell(row=5, column=4, value="PORT OF LOADING: " + data.get("port_of_loading", ""))
    ws.cell(row=6, column=1, value="PORT OF DISCHARGE: " + data.get("port_of_discharge", ""))
    ws.cell(row=6, column=4, value="FINAL DESTINATION: " + data.get("final_destination", ""))
    
    ws.cell(row=8, column=1, value=data.get("marks_nos", ""))
    ws.cell(row=8, column=2, value=data.get("type_of_packing", ""))
    
    total_cbm = sum(float(c.get("cbm", 0)) for c in containers)
    pkgs_text = data.get("total_pkgs", "")
    
    ws.cell(row=8, column=3, value=f"{len(containers)}X20' FCL CONTAINER SAID TO CONTAIN\nTOTAL: {pkgs_text}")
    ws.cell(row=9, column=3, value=data.get("cargo_desc", ""))
    
    ws.cell(row=11, column=3, value="INV NOS :")
    ws.cell(row=12, column=3, value="  " + data.get("invoice_no_date", ""))
    ws.cell(row=14, column=3, value="HS CODE : " + data.get("hs_code", ""))
    ws.cell(row=16, column=3, value="SB NO : " + data.get("sb_no_date", ""))
    
    ws.cell(row=8, column=5, value=data.get("gross_weight", ""))
    ws.cell(row=10, column=5, value="NET. WT.")
    ws.cell(row=11, column=5, value=data.get("net_weight", ""))
    ws.cell(row=8, column=6, value=f"{total_cbm:.3f} CBM" if total_cbm > 0 else "")
    
    first_item_row = 20
    template_rows = 4
    
    if len(containers) > template_rows:
        rows_to_insert = len(containers) - template_rows
        for _ in range(rows_to_insert):
            insert_row_copy_style(ws, 24, 23)
            
    for idx, c in enumerate(containers):
        r = first_item_row + idx
        ws.cell(row=r, column=1, value=c.get("container_no", ""))
        ws.cell(row=r, column=2, value=c.get("seal_no", ""))
        ws.cell(row=r, column=3, value=c.get("pkgs", ""))
        
        try:
            net_val = float(c.get("net_wt", 0))
            ws.cell(row=r, column=4, value=net_val)
            ws.cell(row=r, column=4).number_format = "#,##0"
        except:
            ws.cell(row=r, column=4, value=c.get("net_wt", ""))
            
        try:
            gross_val = float(c.get("gross_wt", 0))
            ws.cell(row=r, column=5, value=gross_val)
            ws.cell(row=r, column=5).number_format = "#,##0"
        except:
            ws.cell(row=r, column=5, value=c.get("gross_wt", ""))
            
        try:
            cbm_val = float(c.get("cbm", 0))
            ws.cell(row=r, column=6, value=cbm_val)
            ws.cell(row=r, column=6).number_format = "0.00"
        except:
            ws.cell(row=r, column=6, value=c.get("cbm", ""))
            
    if len(containers) < template_rows:
        for r in range(first_item_row + len(containers), first_item_row + template_rows):
            ws.cell(row=r, column=1, value=None)
            ws.cell(row=r, column=2, value=None)
            ws.cell(row=r, column=3, value=None)
            ws.cell(row=r, column=4, value=None)
            ws.cell(row=r, column=5, value=None)
            ws.cell(row=r, column=6, value=None)
            
    wb.save(out_path)

def generate_awb_instructions(data, template_file, out_path):
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True

    # 1. Destination
    ws["B6"] = data.get("final_destination", "")

    # 2. Collect/Prepaid
    pmode = str(data.get("payment_mode", "")).upper()
    if "COLL" in pmode:
        ws["H6"] = "COLLECT"
        ws["G8"] = "COLLECT"
    else:
        ws["H6"] = "PREPAID"
        ws["G8"] = "PREPAID"

    # 3. Address Blocks
    write_address_block(ws, 8, 2, data.get("shipper", ""), max_rows=4)
    write_address_block(ws, 13, 2, data.get("consignee", ""), max_rows=4)
    write_address_block(ws, 18, 2, data.get("notify", ""), max_rows=4)
    write_address_block(ws, 23, 2, data.get("buyer_if_other", ""), max_rows=4)

    # 4. Invoice No & Date
    ws["B28"] = data.get("invoice_no_date", "")

    # 5. AWB Number (MAWB/HAWB combined)
    mawb = data.get("mawb_no_date", "")
    hawb = data.get("hawb_no_date", "")
    awb_str = ""
    if mawb and hawb:
        awb_str = f"MAWB: {mawb} / HAWB: {hawb}"
    elif mawb:
        awb_str = f"MAWB: {mawb}"
    elif hawb:
        awb_str = f"HAWB: {hawb}"
    ws["G9"] = awb_str

    # 6. No of Pkgs / Packing Type
    total_pkgs = data.get("total_pkgs", "")
    if total_pkgs:
        parts = total_pkgs.split(None, 1)
        try:
            ws["G11"] = int(parts[0].replace(",", ""))
        except:
            ws["G11"] = parts[0]
        if len(parts) > 1:
            ws["H11"] = parts[1]

    # 7. GSTN / AD Codes
    ws["G13"] = data.get("gstn", "")
    ws["G14"] = data.get("ad_code", "")

    # 8. Gross Weight
    ws["G15"] = clean_weight(data.get("gross_weight", ""))

    # 9. Dimensions
    ws["G18"] = data.get("dimensions", "")

    # 10. G.R. No/Date (S.B. No & Date)
    ws["G21"] = data.get("sb_no_date", "")

    # 11. Marks & Nos
    write_wrapped_text_to_rows(ws, 23, 7, data.get("marks_nos", ""), max_rows=3, limit=35)

    # 12. Description of Goods
    cargo_desc = data.get("cargo_desc", "")
    hs = data.get("hs_code", "")
    if hs:
        cargo_desc = f"{cargo_desc}\nHS CODE: {hs}"
    write_wrapped_text_to_rows(ws, 27, 7, cargo_desc, max_rows=5, limit=35)

    # 13. Flight/Airline
    flight_airline = ""
    vessel_voyage = data.get("vessel_voyage", "")
    carrier = data.get("carrier", "")
    if vessel_voyage and carrier:
        flight_airline = f"FLIGHT: {vessel_voyage}\nAIRLINE: {carrier}"
    elif vessel_voyage:
        flight_airline = f"FLIGHT: {vessel_voyage}"
    elif carrier:
        flight_airline = f"AIRLINE: {carrier}"
    write_wrapped_text_to_rows(ws, 33, 7, flight_airline, max_rows=3, limit=35)

    # 14. Bank details
    write_address_block(ws, 33, 2, data.get("bank_ac_drawback", ""), max_rows=2)

    # 15. Invoice Value (Exchange rate/FOB)
    rate_str = f"Exchange Rate: {data.get('exchange_rate', '')}" if data.get('exchange_rate') else ""
    ws["B36"] = rate_str

    wb.save(out_path)

def generate_billing_invoice(data, charges, template_path, out_path):
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Invoice Details"]
    ws.views.sheetView[0].showGridLines = True
    
    # Calculate tax totals
    total_taxable = 0.0
    for charge in charges:
        total_taxable += float(charge.get("amount", 0))
        
    total_cgst = round(total_taxable * 0.09, 2)
    total_sgst = round(total_taxable * 0.09, 2)
    grand_total = round(total_taxable + total_cgst + total_sgst, 2)
    tds = round(total_taxable * 0.02, 2)
    total_payable = round(grand_total - tds, 2)
    
    # Shipper Name
    ws["E2"] = data.get("shipper", "").split("\n")[0].strip()
    
    # Metadata map
    metadata = {
        "JOB NO:": data.get("job_no", ""),
        "BOOKING REF.  NO": data.get("buyer_order_no_date", "").split("DT")[0].split("dt")[0].strip() or "N/A",
        "MBL ": "N/A",
        "DESTINATION": data.get("final_destination", "").split("/")[0].strip(),
        "NO OF PACKS": data.get("total_pkgs", ""),
        "GROSS WEIGHT": data.get("gross_weight", ""),
        "CHARGABLE WEIGHT": data.get("chargeable_weight", ""),
        "DESCRIPTION": data.get("cargo_desc", "").split("\n")[0].strip(),
        "INVOICE NO:&  DATE": data.get("invoice_no_date", ""),
        "SHIPPING LINE": data.get("carrier", ""),
        "SB. NO": data.get("sb_no_date", ""),
        "Account Name": "",
        "Vessel Details": data.get("vessel_voyage", ""),
        "CONTAINER NUMBER ": ", ".join([c.get("container_no", "") for c in data.get("containers", []) if c.get("container_no")]),
        "CBM": data.get("cbm", ""),
        "Remarks": ""
    }
    
    metadata_fields = [
        "JOB NO:", "BOOKING REF.  NO", "MBL ", "DESTINATION", "NO OF PACKS",
        "GROSS WEIGHT", "CHARGABLE WEIGHT", "DESCRIPTION", "INVOICE NO:&  DATE",
        "SHIPPING LINE", "SB. NO", "Account Name", "Vessel Details",
        "CONTAINER NUMBER ", "CBM", "Remarks"
    ]
    for idx, key in enumerate(metadata_fields):
        row_num = 4 + idx
        val = metadata.get(key, "")
        ws.cell(row=row_num, column=5, value=str(val).strip())
        
    N = len(charges)
    if N < 15:
        ws.delete_rows(21 + N, 15 - N)
    elif N > 15:
        ws.insert_rows(35, N - 15)
        # Copy styles
        for r in range(35, 21 + N):
            for col in range(3, 8):
                src_cell = ws.cell(row=34, column=col)
                dest_cell = ws.cell(row=r, column=col)
                dest_cell.font = Font(
                    name=src_cell.font.name, size=src_cell.font.size,
                    bold=src_cell.font.bold, italic=src_cell.font.italic,
                    color=src_cell.font.color, underline=src_cell.font.underline
                )
                dest_cell.border = Border(
                    left=src_cell.border.left, right=src_cell.border.right,
                    top=src_cell.border.top, bottom=src_cell.border.bottom
                )
                dest_cell.alignment = Alignment(
                    horizontal=src_cell.alignment.horizontal,
                    vertical=src_cell.alignment.vertical,
                    wrap_text=src_cell.alignment.wrap_text
                )
                dest_cell.number_format = src_cell.number_format
                
    for idx, charge in enumerate(charges):
        r = 21 + idx
        amount = float(charge.get("amount", 0))
        cgst_rate = 9.0
        sgst_rate = 9.0
        cgst_val = round((amount * cgst_rate) / 100.0, 2)
        sgst_val = round((amount * sgst_rate) / 100.0, 2)
        tax_val = cgst_val + sgst_val
        row_total = amount + tax_val
        
        ws.cell(row=r, column=3, value=idx + 1)
        ws.cell(row=r, column=4, value=charge.get("description", ""))
        ws.cell(row=r, column=5, value=amount)
        ws.cell(row=r, column=6, value=tax_val)
        ws.cell(row=r, column=7, value=row_total)
        
        ws.cell(row=r, column=5).number_format = "#,##0.00"
        ws.cell(row=r, column=6).number_format = "#,##0.00"
        ws.cell(row=r, column=7).number_format = "#,##0.00"
        
    ws.cell(row=21 + N, column=7, value=total_taxable)
    ws.cell(row=22 + N, column=7, value=total_cgst)
    ws.cell(row=23 + N, column=7, value=total_sgst)
    ws.cell(row=24 + N, column=7, value=grand_total)
    ws.cell(row=25 + N, column=7, value=tds)
    ws.cell(row=26 + N, column=7, value=total_payable)
    
    for idx in range(6):
        ws.cell(row=21 + N + idx, column=7).number_format = "#,##0.00"
        
    wb.save(out_path)

def main():
    if len(sys.argv) < 3:
        print("Usage: python compiler.py <parse/compile> <args...>")
        sys.exit(1)
        
    mode = sys.argv[1]
    
    if mode == "parse":
        pdf_path = sys.argv[2]
        try:
            parsed_data = pdf_parser.parse_pdf(pdf_path)
            print(json.dumps(parsed_data, indent=2))
        except Exception as e:
            print(json.dumps({"error": str(e)}))
            sys.exit(1)
            
    elif mode == "compile":
        dest_dir = sys.argv[2]
        try:
            # Read JSON payload from stdin
            payload_str = sys.stdin.read()
            payload = json.loads(payload_str)
            
            os.makedirs(dest_dir, exist_ok=True)
            
            # Extract lists
            data = payload.get("info", {})
            containers = data.get("containers", [])
            charges = data.get("billing_charges", [])
            
            # Exporter template detection
            shipper_name = data.get("shipper", "").upper()
            template_name = "Standard"
            job_type = data.get("type", "sea_export")
            is_air_export = job_type == "air_export"
            
            templates_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")
            
            if is_air_export:
                template_file = os.path.join(templates_dir, "AWB_Instructions_template.xlsx")
                template_name = "AWB"
            elif "COTTON HOME" in shipper_name:
                template_file = os.path.join(templates_dir, "COTTON HOME_BL INSTRUCTION - CHX03_FCL 1X40.xlsx")
                template_name = "Cotton Home"
            elif "PEMI" in shipper_name:
                template_file = os.path.join(templates_dir, "PEMI - BL INSTRUCTION - PE044 -  855 CTNS - GENOA.xlsx")
                template_name = "PEMI"
            elif "SAKTHI" in shipper_name:
                template_file = os.path.join(templates_dir, "SRI SAKTHI - BL INSTRUCTION_ 432 PKGS_ SINGAPORE.xlsx")
                template_name = "Sri Sakthi"
            else:
                template_file = os.path.join(templates_dir, "BL INSTRUCTION CPS 001_template.xlsx")
                template_name = "Standard"
                
            # Clean filename names
            clean_inv = "REF"
            inv_no_raw = data.get("invoice_no_date", "")
            if inv_no_raw:
                part = inv_no_raw.split("DT")[0].split("dt")[0].strip()
                clean_inv = re.sub(r'[^A-Za-z0-9]', '', part)
                
            if is_air_export:
                bl_filename = f"Generated_AWB_Instructions_{clean_inv}.xlsx"
            else:
                bl_filename = f"Generated_BL_Instructions_{clean_inv}.xlsx"
                if template_name == "Cotton Home":
                    bl_filename = f"COTTON_BL_Instruction_{clean_inv}.xlsx"
                elif template_name == "PEMI":
                    bl_filename = f"PEMI_BL_Instruction_{clean_inv}.xlsx"
                elif template_name == "Sri Sakthi":
                    bl_filename = f"SRI_SAKTHI_BL_Instruction_{clean_inv}.xlsx"
                
            bl_out_path = os.path.join(dest_dir, bl_filename)
            
            # Generate BL/AWB
            if template_name == "AWB":
                generate_awb_instructions(data, template_file, bl_out_path)
            elif template_name == "Cotton Home":
                generate_bl_cotton_home(data, containers, template_file, bl_out_path)
            elif template_name == "PEMI":
                generate_bl_pemi(data, containers, template_file, bl_out_path)
            elif template_name == "Sri Sakthi":
                generate_bl_sri_sakthi(data, containers, template_file, bl_out_path)
            else:
                generate_bl_standard(data, containers, template_file, bl_out_path)
                
            # Generate Billing
            shipper_clean = "CLIENT"
            if "COTTON HOME" in shipper_name: shipper_clean = "COTTON_HOME"
            elif "PEMI" in shipper_name: shipper_clean = "PEMI"
            elif "SAKTHI" in shipper_name: shipper_clean = "SRI_SAKTHI"
            else:
                s_parts = shipper_name.split("\n")[0].split()
                if s_parts: shipper_clean = re.sub(r'[^A-Za-z0-9]', '', s_parts[0])
                
            job_no = data.get("job_no", clean_inv)
            job_no_clean = re.sub(r'[^A-Za-z0-9]', '', job_no)
            
            bill_filename = f"{shipper_clean}_JOB_NO_{job_no_clean}.xlsx"
            bill_out_path = os.path.join(dest_dir, bill_filename)
            
            billing_template_path = os.path.join(templates_dir, "Billing_Invoice.xlsx")
            generate_billing_invoice(data, charges, billing_template_path, bill_out_path)
            
            # Return filenames to node
            response = {
                "success": True,
                "bl_file": f"/downloads/{bl_filename}",
                "billing_file": f"/downloads/{bill_filename}"
            }
            print(json.dumps(response))
        except Exception as e:
            import traceback
            print(json.dumps({"success": False, "error": str(e), "traceback": traceback.format_exc()}))
            sys.exit(1)

if __name__ == "__main__":
    main()
